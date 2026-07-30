

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_cookie
from django.contrib import messages
from django.http import HttpResponse
from django.http import JsonResponse
from django.db.models import Q, Sum, Count
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io
import csv

from eleves.models import Eleve
from .models import AbonnementBus
from .forms import AbonnementBusForm
from utilisateurs.utils import user_is_admin, user_is_superadmin, filter_by_user_school
from utilisateurs.permissions import can_delete_subscriptions
from ecole_moderne.security_decorators import require_school_object
from ecole_moderne.pdf_utils import draw_logo_watermark
from paiements.twilio_utils import send_message_async


@login_required
def tableau_bord(request):
    qs = AbonnementBus.objects.select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')

    today = timezone.localdate()
    total = qs.count()
    # Optimisation : éviter de charger tous les objets en mémoire
    exp = qs.filter(date_expiration__lt=today).count()
    proche = 0
    for date_exp, alerte_jours in qs.values_list('date_expiration', 'alerte_avant_jours'):
        if date_exp and date_exp >= today:
            delta = (date_exp - today).days
            if 0 <= delta <= (alerte_jours or 7):
                proche += 1

    context = {
        'titre_page': 'Abonnements Bus',
        'total': total,
        'exp': exp,
        'proche': proche,
    }
    return render(request, 'bus/tableau_bord.html', context)


@login_required
def liste_abonnements(request):
    qs = AbonnementBus.objects.select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')

    q = (request.GET.get('q') or '').strip()
    filtre = (request.GET.get('filtre') or '').strip().lower()
    if q:
        qs = qs.filter(
            Q(eleve__nom__icontains=q) |
            Q(eleve__prenom__icontains=q) |
            Q(eleve__matricule__icontains=q) |
            Q(zone__icontains=q) |
            Q(point_arret__icontains=q) |
            Q(contact_parent__icontains=q)
        )

    # Appliquer le filtre de statut/échéance
    today = timezone.localdate()
    if filtre == 'expires':
        qs = qs.filter(statut=AbonnementBus.Statut.EXPIRE)
    elif filtre == 'suspendus':
        qs = qs.filter(statut=AbonnementBus.Statut.SUSPENDU)
    elif filtre == 'depassees':
        qs = qs.filter(date_expiration__lt=today)
    elif filtre == 'proches':
        # Sélectionner par logique métier est_proche_expiration (fenêtre d'alerte)
        ids_proches = []
        for a_id, date_exp, alerte_jours in qs.values_list('id', 'date_expiration', 'alerte_avant_jours'):
            if date_exp:
                if date_exp >= today:
                    delta = (date_exp - today).days
                    if 0 <= delta <= (alerte_jours or 7):
                        ids_proches.append(a_id)
        qs = qs.filter(id__in=ids_proches)

    # Aggregates for dashboard (sur le queryset filtré)
    total_count = qs.count()
    agg = qs.aggregate(
        total_montant=Sum('montant'),
        nb_actifs=Count('id', filter=Q(statut=AbonnementBus.Statut.ACTIF)),
        nb_expires=Count('id', filter=Q(statut=AbonnementBus.Statut.EXPIRE)),
        nb_suspendus=Count('id', filter=Q(statut=AbonnementBus.Statut.SUSPENDU)),
        montant_actifs=Sum('montant', filter=Q(statut=AbonnementBus.Statut.ACTIF)),
        montant_expires=Sum('montant', filter=Q(statut=AbonnementBus.Statut.EXPIRE)),
        montant_suspendus=Sum('montant', filter=Q(statut=AbonnementBus.Statut.SUSPENDU)),
    )

    # Counts for expiration proximity using only necessary fields
    nb_expiration_proche = 0
    nb_expiration_depassee = 0
    for date_exp, alerte_jours in qs.values_list('date_expiration', 'alerte_avant_jours'):
        if date_exp:
            if date_exp < today:
                nb_expiration_depassee += 1
            else:
                delta = (date_exp - today).days
                if 0 <= delta <= (int(alerte_jours) if alerte_jours else 7):
                    nb_expiration_proche += 1

    # Periodicite breakdown
    choices_map = dict(AbonnementBus.Periodicite.choices)
    periodicite_rows = []
    for row in qs.values('periodicite').annotate(nb=Count('id'), montant=Sum('montant')):
        code = row['periodicite']
        periodicite_rows.append({
            'code': code,
            'label': choices_map.get(code, code or '-'),
            'nb': row['nb'] or 0,
            'montant': row['montant'] or 0,
        })

    # Top zones breakdown (limit 10)
    zone_rows = []
    for row in qs.values('zone').annotate(nb=Count('id'), montant=Sum('montant')).order_by('-nb')[:10]:
        zone_rows.append({
            'zone': row['zone'] or '-',
            'nb': row['nb'] or 0,
            'montant': row['montant'] or 0,
        })

    context = {
        'titre_page': 'Abonnements Bus - Liste',
        'abonnements': qs.order_by('-updated_at')[:500],
        'q': q,
        'filtre': filtre,
        # Dashboard context
        'total_count': total_count,
        'total_montant': agg.get('total_montant') or 0,
        'nb_actifs': agg.get('nb_actifs') or 0,
        'nb_expires': agg.get('nb_expires') or 0,
        'nb_suspendus': agg.get('nb_suspendus') or 0,
        'nb_expiration_proche': nb_expiration_proche,
        'nb_expiration_depassee': nb_expiration_depassee,
        'montant_actifs': agg.get('montant_actifs') or 0,
        'montant_expires': agg.get('montant_expires') or 0,
        'montant_suspendus': agg.get('montant_suspendus') or 0,
        # Breakdowns
        'periodicite_rows': periodicite_rows,
        'zone_rows': zone_rows,
    }
    return render(request, 'bus/liste.html', context)


@login_required
def abonnement_create(request):
    initial = {}
    eleve_id = request.GET.get('eleve')
    if eleve_id:
        try:
            initial['eleve'] = Eleve.objects.get(id=int(eleve_id))
        except Exception:
            pass
    if request.method == 'POST':
        form = AbonnementBusForm(request.POST)
        if form.is_valid():
            abo = form.save()
            messages.success(request, "Abonnement bus créé avec succès.")
            return redirect('bus:liste')
    else:
        form = AbonnementBusForm(initial=initial)

    # Sécurité : filtrer les élèves par école de l'utilisateur
    if not user_is_superadmin(request.user):
        form.fields['eleve'].queryset = filter_by_user_school(
            Eleve.objects.all(),
            request.user,
            'classe__ecole'
        )

    return render(request, 'bus/form.html', {'form': form, 'titre_page': 'Nouvel abonnement Bus'})


@login_required
@require_school_object(model=AbonnementBus, pk_kwarg='abo_id', field_path='eleve__classe__ecole')
def abonnement_edit(request, abo_id):
    abo = get_object_or_404(AbonnementBus, id=abo_id)
    if request.method == 'POST':
        form = AbonnementBusForm(request.POST, instance=abo)
        if form.is_valid():
            form.save()
            messages.success(request, "Abonnement mis à jour.")
            return redirect('bus:liste')
    else:
        form = AbonnementBusForm(instance=abo)

    # Sécurité : filtrer les élèves par école de l'utilisateur
    if not user_is_superadmin(request.user):
        form.fields['eleve'].queryset = filter_by_user_school(
            Eleve.objects.all(),
            request.user,
            'classe__ecole'
        )

    return render(request, 'bus/form.html', {'form': form, 'titre_page': 'Modifier abonnement Bus'})


@login_required
@can_delete_subscriptions
@require_school_object(model=AbonnementBus, pk_kwarg='abo_id', field_path='eleve__classe__ecole')
def supprimer_abonnement_bus(request, abo_id):
    """Supprimer définitivement un abonnement bus"""
    abonnement = get_object_or_404(AbonnementBus, id=abo_id)
    
    if request.method == 'POST':
        eleve_nom = str(abonnement.eleve)
        abonnement.delete()
        messages.success(request, f"Abonnement bus supprimé définitivement pour {eleve_nom}")
        return redirect('bus:liste')
    
    context = {
        'titre_page': 'Supprimer Abonnement Bus',
        'abonnement': abonnement,
    }
    return render(request, 'bus/confirmer_suppression.html', context)


@login_required
def relances(request):
    qs = AbonnementBus.objects.select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')

    a_relancer = [a for a in qs if a.est_expire or a.est_proche_expiration or a.statut != AbonnementBus.Statut.ACTIF]

    context = {
        'titre_page': 'Relances Abonnements Bus',
        'abonnements': a_relancer,
    }
    return render(request, 'bus/relances.html', context)


@login_required
def export_relances_excel(request):
    qs = AbonnementBus.objects.select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')

    data = [a for a in qs if a.est_expire or a.est_proche_expiration or a.statut != AbonnementBus.Statut.ACTIF]

    wb = Workbook(); ws = wb.active; ws.title = 'Relances Bus'
    headers = ['Élève', 'Classe', 'École', 'Périodicité', 'Montant', 'Début', 'Expiration', 'Statut', 'Zone', "Point d'arrêt", 'Contact parent']
    ws.append(headers)
    for a in data:
        el = a.eleve
        ws.append([
            f"{el.prenom} {el.nom} ({el.matricule})",
            getattr(el.classe, 'nom', ''),
            getattr(getattr(el.classe, 'ecole', None), 'nom', ''),
            a.get_periodicite_display(),
            int(a.montant or 0),
            a.date_debut.strftime('%d/%m/%Y') if a.date_debut else '',
            a.date_expiration.strftime('%d/%m/%Y') if a.date_expiration else '',
            a.get_statut_display(),
            a.zone,
            a.point_arret,
            a.contact_parent,
        ])
    widths = [30, 16, 22, 16, 14, 14, 14, 12, 16, 18, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    import io
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    resp = HttpResponse(stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="relances_bus.xlsx"'
    return resp


@login_required
@vary_on_cookie
@cache_page(60 * 10)
@require_school_object(model=AbonnementBus, pk_kwarg='abo_id', field_path='eleve__classe__ecole')
def generer_recu_abonnement_pdf(request, abo_id):
    """Génère un reçu simple pour un abonnement bus"""
    abo = get_object_or_404(AbonnementBus.objects.select_related('eleve', 'eleve__classe', 'eleve__classe__ecole'), id=abo_id)
    try:
        import io
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from django.contrib.staticfiles import finders
    except Exception:
        messages.error(request, "ReportLab requis (pip install reportlab)")
        return redirect('bus:liste')

    buffer = io.BytesIO(); c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Filigrane standardisé (logo centré, rotation légère, opacité 4%) — dynamique par école
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        ecole_obj = getattr(getattr(abo.eleve, 'classe', None), 'ecole', None)
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=ecole_obj)
    except Exception:
        pass

    # Logo de l'école en haut à gauche
    try:
        from reportlab.lib.utils import ImageReader
        import os
        logo_path = None
        ecole_obj = getattr(getattr(abo.eleve, 'classe', None), 'ecole', None)
        
        # Essayer d'abord le logo de l'école
        if ecole_obj and hasattr(ecole_obj, 'logo'):
            school_logo_path = getattr(getattr(ecole_obj, 'logo', None), 'path', None)
            if school_logo_path and os.path.exists(school_logo_path):
                logo_path = school_logo_path
        
        # Fallback vers le logo statique global
        if not logo_path:
            logo_path = finders.find('logos/logo.png')
        
        if logo_path:
            try:
                logo_img = ImageReader(logo_path)
                logo_w, logo_h = 60, 60
                c.drawImage(logo_img, 40, height - 100, width=logo_w, height=logo_h, preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
    except Exception:
        pass

    # Titre
    c.setFont('Helvetica-Bold', 18)
    title = 'REÇU ABONNEMENT BUS SCOLAIRE'
    tw = c.stringWidth(title, 'Helvetica-Bold', 18)
    c.drawString((width - tw)/2, height - 50, title)
    
    # Nom de l'école sous le titre
    try:
        ecole_nom = getattr(ecole_obj, 'nom', '')
        if ecole_nom:
            c.setFont('Helvetica-Bold', 12)
            tw_ecole = c.stringWidth(ecole_nom, 'Helvetica-Bold', 12)
            c.drawString((width - tw_ecole)/2, height - 70, ecole_nom)
    except Exception:
        pass

    # Photo de l'élève en haut à droite
    try:
        from reportlab.lib.utils import ImageReader
        import os
        img_drawn = False
        img_w, img_h = 100, 100
        x_img = width - 40 - img_w
        y_img = height - 40 - img_h
        
        el = abo.eleve
        photo_path = getattr(getattr(el, 'photo', None), 'path', None)
        
        if photo_path and os.path.exists(photo_path):
            try:
                img = ImageReader(photo_path)
                c.drawImage(img, x_img, y_img, width=img_w, height=img_h, preserveAspectRatio=True, mask='auto')
                img_drawn = True
            except Exception:
                img_drawn = False
        
        if not img_drawn:
            # Dessiner un placeholder avec initiales
            nom_complet = f"{getattr(el, 'prenom', '')} {getattr(el, 'nom', '')}".strip()
            initiales = ''.join([p[0].upper() for p in nom_complet.split()[:2]]) or 'E'
            c.setLineWidth(1)
            try:
                c.roundRect(x_img, y_img, img_w, img_h, 8)
            except Exception:
                c.rect(x_img, y_img, img_w, img_h)
            c.setFont('Helvetica-Bold', 24)
            c.drawCentredString(x_img + img_w/2, y_img + img_h/2 - 8, initiales)
            c.setFont('Helvetica', 8)
            c.drawCentredString(x_img + img_w/2, y_img + 6, "Pas de photo")
        
        # Afficher le nom de l'élève sous l'image/placeholder
        try:
            nom_aff = f"{getattr(el, 'prenom', '')} {getattr(el, 'nom', '')}".strip()
            if nom_aff:
                c.setFont('Helvetica-Bold', 10)
                c.drawCentredString(x_img + img_w/2, y_img - 12, nom_aff)
        except Exception:
            pass
    except Exception:
        # En cas de problème, ne pas bloquer la génération du reçu
        pass

    # Corps
    y = height - 110
    c.setFont('Helvetica', 12)
    def line(lbl, val):
        nonlocal y
        c.setFont('Helvetica-Bold', 12); c.drawString(40, y, f"{lbl} :")
        c.setFont('Helvetica', 12); c.drawString(200, y, str(val)); y -= 20

    el = abo.eleve
    line('Élève', f"{el.prenom} {el.nom} ({el.matricule})")
    line('Classe', getattr(el.classe, 'nom', ''))
    line('École', getattr(getattr(el.classe, 'ecole', None), 'nom', ''))
    line('Périodicité', abo.get_periodicite_display())
    line('Montant', f"{int(abo.montant):,}".replace(',', ' ') + ' GNF')
    line('Début', abo.date_debut.strftime('%d/%m/%Y') if abo.date_debut else '')
    line('Expiration', abo.date_expiration.strftime('%d/%m/%Y') if abo.date_expiration else '')
    
    # Calcul et affichage de la durée en jours
    if abo.date_debut and abo.date_expiration:
        duree_jours = (abo.date_expiration - abo.date_debut).days
        line('Durée', f"{duree_jours} jours")
    line('Zone', abo.zone)
    line("Point d'arrêt", abo.point_arret)
    line('Contact parent', abo.contact_parent)

    c.showPage(); c.save(); pdf = buffer.getvalue(); buffer.close()
    resp = HttpResponse(content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename=recu_abonnement_{abo.id}.pdf'
    resp.write(pdf)
    return resp


@login_required
def envoyer_relances_bus(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)

    try:
        ids = request.POST.getlist('abo_ids')
        message_type = (request.POST.get('message_type') or 'sms').lower()
        message_personnalise = (request.POST.get('message_personnalise') or '').strip()
    except Exception:
        return JsonResponse({'success': False, 'error': 'Requête invalide'}, status=400)

    if not ids:
        return JsonResponse({'success': False, 'error': 'Aucun abonnement sélectionné'}, status=400)

    channel = 'whatsapp' if message_type == 'whatsapp' else 'sms'
    envoyes = 0

    abonnements = AbonnementBus.objects.select_related('eleve', 'eleve__classe', 'eleve__classe__ecole').filter(id__in=ids)
    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if not user_is_superadmin(request.user):
        abonnements = filter_by_user_school(abonnements, request.user, 'eleve__classe__ecole')

    for abo in abonnements:
        el = abo.eleve

        # Préparer destinataires (principal + secondaire)
        destinataires = []
        rp = getattr(el, 'responsable_principal', None)
        rs = getattr(el, 'responsable_secondaire', None)
        if rp and getattr(rp, 'telephone', None):
            destinataires.append((getattr(rp, 'nom', 'Responsable'), rp.telephone))
        if rs and getattr(rs, 'telephone', None):
            destinataires.append((getattr(rs, 'nom', 'Responsable'), rs.telephone))

        if not destinataires:
            continue

        # Construire message
        if not message_personnalise:
            base_msg = (
                "Bonjour {nom_responsable},\n\n"
                "L'abonnement bus de {prenom_eleve} {nom_eleve} ({classe}) {etat} le {date_expiration}.\n"
                "Montant: {montant} GNF. Zone: {zone}. Point d'arrêt: {arret}.\n\n"
                "Merci de procéder au renouvellement.\n"
                "École {nom_ecole}"
            )
        else:
            base_msg = message_personnalise

        etat = 'a expiré' if abo.est_expire else 'arrive à expiration'
        date_exp = abo.date_expiration.strftime('%d/%m/%Y') if abo.date_expiration else ''
        montant_txt = f"{int(abo.montant):,}".replace(',', ' ')
        classe_nom = getattr(el.classe, 'nom', 'Non définie')
        ecole_nom = getattr(getattr(el.classe, 'ecole', None), 'nom', 'École')

        for nom_resp, numero in destinataires:
            try:
                msg = base_msg.format(
                    nom_responsable=nom_resp or 'Responsable',
                    prenom_eleve=getattr(el, 'prenom', ''),
                    nom_eleve=getattr(el, 'nom', ''),
                    classe=classe_nom,
                    etat=etat,
                    date_expiration=date_exp,
                    montant=montant_txt,
                    zone=abo.zone or '-',
                    arret=abo.point_arret or '-',
                    nom_ecole=ecole_nom,
                )
            except Exception:
                msg = base_msg

            # Envoi
            try:
                send_message_async(to_number=numero, body=msg, channel=channel)
                envoyes += 1
                abo.derniere_relance = timezone.now()
                abo.save(update_fields=['derniere_relance'])
            except Exception:
                # Continuer autres destinataires
                pass

    return JsonResponse({'success': True, 'message': f'{envoyes} message(s) envoyé(s)'})


@login_required
def export_abonnements_breakdown_csv(request, kind: str):
    """Exporte au format CSV les répartitions par périodicité ou par zone.

    kind: 'periodicite' ou 'zone'
    Optionnellement respecte le filtre de recherche ?q= comme la liste.
    """
    kind = (kind or '').lower()
    if kind not in ('periodicite', 'zone'):
        return HttpResponse('Type invalide', status=400)

    qs = AbonnementBus.objects.select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')

    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(eleve__nom__icontains=q) |
            Q(eleve__prenom__icontains=q) |
            Q(eleve__matricule__icontains=q) |
            Q(zone__icontains=q) |
            Q(point_arret__icontains=q) |
            Q(contact_parent__icontains=q)
        )

    buffer = io.StringIO()
    writer = csv.writer(buffer)

    if kind == 'periodicite':
        writer.writerow(['Périodicité', 'Nombre', 'Montant (GNF)'])
        label_map = dict(AbonnementBus.Periodicite.choices)
        for row in qs.values('periodicite').annotate(nb=Count('id'), montant=Sum('montant')).order_by('periodicite'):
            code = row['periodicite']
            label = label_map.get(code, code or '-')
            writer.writerow([label, row['nb'] or 0, int(row['montant'] or 0)])
        filename = 'repartition_periodicite.csv'
    else:
        writer.writerow(['Zone', 'Nombre', 'Montant (GNF)'])
        for row in qs.values('zone').annotate(nb=Count('id'), montant=Sum('montant')).order_by('-nb'):
            writer.writerow([row['zone'] or '-', row['nb'] or 0, int(row['montant'] or 0)])
        filename = 'repartition_zones.csv'

    content = buffer.getvalue().encode('utf-8-sig')
    resp = HttpResponse(content, content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
