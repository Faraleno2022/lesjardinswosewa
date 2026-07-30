"""
Système de Bulletin Intelligent avec Calculs Automatiques
Exports PDF (avec filigrane) et Excel
Conforme au système guinéen (40% cours + 60% composition)
"""

from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render
from django.http import HttpResponse
from django.db.models import Avg, Count
from decimal import Decimal
import io
from datetime import datetime

# ReportLab pour PDF
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.lib.utils import ImageReader
from PIL import Image

# OpenPyXL pour Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from eleves.models import Eleve, Classe
from .models import ClasseNote, MatiereNote, Evaluation, NoteEleve
# Import du module centralisé pour garantir la cohérence
from .calculs_moyennes import (
    calculer_moyenne_generale_eleve,
    calculer_classement_classe,
    obtenir_mention_intelligente,
    obtenir_appreciation_intelligente,
    formater_rang_intelligent
)
from .calculs import (
    calculer_moyenne_devoirs,
    calculer_moyenne_periode,
    calculer_moyenne_cours_mensuels,
    calculer_rang
)
from .classifier import classify
from utilisateurs.utils import filter_by_user_school, user_school
from ecole_moderne.security_decorators import require_school_object


class CalculateurBulletinIntelligent:
    """Calculateur intelligent de bulletin conforme au système guinéen"""
    
    def __init__(self, eleve, classe_note, periode, systeme='TRIMESTRE'):
        self.eleve = eleve
        self.classe_note = classe_note
        self.periode = periode
        self.systeme = systeme
        self.niveau, self.serie, self.section = self._detecter_niveau_intelligent()
        
    def _detecter_niveau_intelligent(self):
        """Détecte intelligemment le niveau, la série et la section à partir du nom de classe"""
        # Essayer d'abord le champ niveau_enseignement si disponible et cohérent
        niveau_db = getattr(self.classe_note, 'niveau_enseignement', None)
        
        # Classification intelligente à partir du nom
        niveau_auto, serie, section = classify(self.classe_note.nom)
        
        # Si le niveau DB existe et est cohérent, on le garde
        if niveau_db in ['PRIMAIRE', 'MATERNELLE']:
            niveau_final = 'PRIMAIRE'
        elif niveau_db in ['COLLEGE', 'LYCEE']:
            niveau_final = 'SECONDAIRE'
        else:
            # Sinon on utilise la détection automatique
            niveau_final = niveau_auto
        
        return niveau_final, serie, section
    
    def _obtenir_mois_periode(self):
        """Retourne les mois d'une période"""
        mapping = {
            'TRIMESTRE_1': ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE'],
            'TRIMESTRE_2': ['JANVIER', 'FEVRIER', 'MARS'],
            'TRIMESTRE_3': ['AVRIL', 'MAI', 'JUIN'],
            'SEMESTRE_1': ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER'],
            'SEMESTRE_2': ['MARS', 'AVRIL', 'MAI', 'JUIN'],
        }
        return mapping.get(self.periode, [])
    
    def calculer_notes_matiere(self, matiere):
        """Calcule les notes d'une matière pour la période"""
        mois_periode = self._obtenir_mois_periode()
        
        # Récupérer les notes mensuelles
        notes_mensuelles = {}
        for mois in mois_periode:
            notes_mois_qs = NoteEleve.objects.filter(
                eleve=self.eleve,
                evaluation__matiere=matiere,
                evaluation__periode=mois,
                evaluation__type_eval='DEVOIR'
            )

            notes_mois = []
            for note_obj in notes_mois_qs:
                if note_obj.absent or note_obj.note is None:
                    notes_mois.append(Decimal('0'))
                else:
                    notes_mois.append(Decimal(str(note_obj.note)))

            if notes_mois:
                notes_mensuelles[mois.lower()] = notes_mois
        
        # Récupérer la composition
        composition = None
        compo_obj = NoteEleve.objects.filter(
            eleve=self.eleve,
            evaluation__matiere=matiere,
            evaluation__periode=self.periode,
            evaluation__type_eval='COMPOSITION'
        ).first()
        
        if compo_obj:
            if compo_obj.absent or compo_obj.note is None:
                composition = Decimal('0')
            else:
                composition = Decimal(str(compo_obj.note))
        
        # Calculer selon le niveau
        if self.niveau == 'PRIMAIRE':
            # Primaire : composition uniquement
            moyenne_periode = composition
        else:
            # Secondaire : 40% cours + 60% composition
            moyenne_cours = None
            if notes_mensuelles:
                moyenne_cours = calculer_moyenne_cours_mensuels(notes_mensuelles)
            
            moyenne_periode = calculer_moyenne_periode(
                moyenne_cours,
                composition,
                niveau=self.niveau
            )
        
        return {
            'matiere': matiere.nom,
            'coefficient': matiere.coefficient,
            'notes_mensuelles': notes_mensuelles,
            'composition': composition,
            'moyenne': moyenne_periode
        }
    
    def generer_bulletin(self):
        """Génère le bulletin complet pour la période en utilisant le module centralisé"""
        from .calculs_moyennes import detecter_niveau_scolaire
        
        # Détecter si c'est une classe de maternelle
        niveau_detecte = detecter_niveau_scolaire(self.classe_note.nom)
        est_maternelle = (niveau_detecte == 'MATERNELLE')
        
        # Récupérer toutes les matières
        matieres = MatiereNote.objects.filter(classe=self.classe_note)
        
        # Pour la maternelle : utiliser les appréciations
        if est_maternelle:
            return self._generer_bulletin_maternelle(matieres)
        
        # UTILISER LE MODULE CENTRALISÉ (SOURCE UNIQUE)
        system_type = 'trimestre' if 'TRIMESTRE' in self.periode else 'semestre' if 'SEMESTRE' in self.periode else 'mensuel'
        result_centralized = calculer_moyenne_generale_eleve(
            self.eleve, 
            matieres, 
            self.periode, 
            system_type
        )
        
        # Extraire les données du calcul centralisé
        moyenne_generale = result_centralized['moyenne_generale']
        resultats_matieres = result_centralized['details_matieres']
        
        # Calculer le rang en utilisant la source centralisée des rangs
        # ET utiliser la moyenne de cette même source pour garantir la cohérence
        # entre bulletin, classement et satisfécit.
        rang = None
        from .utils_rangs import calculer_rangs_classe_periode
        rangs_dict = calculer_rangs_classe_periode(self.classe_note, self.periode, use_cache=False)
        rang_info = rangs_dict.get(self.eleve.id)
        if rang_info:
            rang = rang_info['rang']
            # Utiliser la moyenne de la source centralisée pour cohérence
            moyenne_source = float(rang_info['moyenne'])
            if moyenne_source is not None:
                moyenne_generale = moyenne_source
                total_coefficients = result_centralized.get('total_coefficients') or 0
                result_centralized['total_points'] = round(moyenne_source * float(total_coefficients), 2) if total_coefficients else result_centralized.get('total_points')
        
        return {
            'eleve': f"{self.eleve.prenom} {self.eleve.nom}",
            'prenom': self.eleve.prenom or '',
            'nom': self.eleve.nom or '',
            'classe': self.classe_note.nom,
            'periode': self.periode,
            'niveau': self.niveau,
            'serie': self.serie,
            'section': self.section,
            'matieres': resultats_matieres,
            'moyenne_generale': moyenne_generale,
            'total_points': result_centralized.get('total_points'),
            'total_coefficients': result_centralized.get('total_coefficients'),
            'mention': obtenir_mention_intelligente(moyenne_generale, self.niveau) if moyenne_generale else None,
            'appreciation': obtenir_appreciation_intelligente(moyenne_generale, self.eleve.prenom, self.niveau) if moyenne_generale else None,
            'rang': rang,
            'total_eleves': Eleve.objects.filter(classe=self.eleve.classe, statut='ACTIF').count()
        }
    
    def _generer_bulletin_maternelle(self, matieres):
        """Génère le bulletin pour la maternelle avec appréciations"""
        from .models import AppreciationMaternelle
        from .utils_rangs import calculer_rangs_classe_periode
        
        # Récupérer les appréciations pour chaque matière
        resultats_matieres = []
        for matiere in matieres:
            appreciation_data = {
                'matiere': matiere.nom,
                'matiere_obj': matiere,
                'coefficient': 1,
                'appreciation': None,
                'appreciation_display': '-',
                'commentaire': '-',
                'absent': False
            }
            
            try:
                app_obj = AppreciationMaternelle.objects.get(
                    eleve=self.eleve,
                    matiere=matiere,
                    trimestre=self.periode,
                    annee_scolaire=self.classe_note.annee_scolaire
                )
                appreciation_data['appreciation'] = app_obj.appreciation
                appreciation_data['appreciation_display'] = app_obj.get_appreciation_display()
                appreciation_data['commentaire'] = app_obj.commentaire or self._get_observation_auto(app_obj.appreciation)
                appreciation_data['absent'] = app_obj.absent
            except AppreciationMaternelle.DoesNotExist:
                pass
            
            resultats_matieres.append(appreciation_data)
        
        # Calculer le rang et taux d'acquisition
        rangs_dict = calculer_rangs_classe_periode(self.classe_note, self.periode, use_cache=False)
        rang_info = rangs_dict.get(self.eleve.id)
        
        taux_acquisition = None
        rang = None
        mention = None
        appreciation = None
        
        if rang_info:
            taux_acquisition = float(rang_info['moyenne'])
            rang = f"{rang_info['rang']}/{rang_info['total_eleves']}"
            
            # Mention basée sur le taux
            if taux_acquisition >= 90:
                mention = 'Excellent'
            elif taux_acquisition >= 75:
                mention = 'Très Bien'
            elif taux_acquisition >= 60:
                mention = 'Bien'
            elif taux_acquisition >= 50:
                mention = 'Assez Bien'
            else:
                mention = 'À encourager'
            
            appreciation = f"Bon trimestre {self.eleve.prenom}. Continue ainsi !"
        
        return {
            'eleve': f"{self.eleve.prenom} {self.eleve.nom}",
            'prenom': self.eleve.prenom or '',
            'nom': self.eleve.nom or '',
            'classe': self.classe_note.nom,
            'periode': self.periode,
            'niveau': 'MATERNELLE',
            'serie': None,
            'section': None,
            'matieres': resultats_matieres,
            'moyenne_generale': taux_acquisition,
            'mention': mention,
            'appreciation': appreciation,
            'rang': rang,
            'total_eleves': Eleve.objects.filter(classe=self.eleve.classe, statut='ACTIF').count(),
            'est_maternelle': True
        }
    
    def _get_observation_auto(self, appreciation):
        """Retourne une observation automatique basée sur l'appréciation"""
        observations = {
            'TRES_BIEN_ACQUIS': 'Excellent travail, continue ainsi !',
            'BIEN_ACQUIS': 'Bon niveau, quelques perfectionnements possibles',
            'EN_COURS': 'Des progrès à faire, persévère !',
            'NON_ACQUIS': 'Nécessite un accompagnement renforcé'
        }
        return observations.get(appreciation, '-')
    
    def _calculer_rang_eleve(self, moyenne_eleve):
        """
        Calcule le rang de l'élève dans la classe.
        
        OPTIMISATION: Utilise le cache centralisé des rangs pour éviter
        de recalculer les moyennes de tous les élèves à chaque appel.
        """
        if not moyenne_eleve:
            return None
        
        # OPTIMISATION: Utiliser le cache centralisé des rangs
        from .utils_rangs import get_rang_eleve
        
        rang_info = get_rang_eleve(self.classe_note, self.periode, self.eleve.id)
        if rang_info:
            return rang_info['rang']
        
        return None


def generer_pdf_avec_filigrane(bulletin_data, logo_path=None, ecole=None):
    """Génère un PDF avec le même design exact que le modèle imprimé"""
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Couleurs du design (exactement comme le modèle)
    BLEU_HEADER = colors.HexColor('#4a90d9')  # Bleu de l'en-tête du tableau
    BLEU_CLAIR = colors.HexColor('#e8f4fd')   # Fond des infos élève
    BLEU_APPRECIATION = colors.HexColor('#d6eaf8')  # Fond appréciation
    VERT_MOY = colors.HexColor('#c8e6c9')     # Colonne MOY
    ROUGE_PTS = colors.HexColor('#ffcdd2')    # Colonne PTS
    ROUGE_DRAPEAU = colors.HexColor('#CE1126')
    JAUNE_DRAPEAU = colors.HexColor('#FCD116')
    VERT_DRAPEAU = colors.HexColor('#009460')
    GRIS_TOTAL = colors.HexColor('#37474f')   # Ligne total
    
    # Couleurs des mentions
    MENTION_COLORS = {
        'EXCELLENT': colors.HexColor('#1b5e20'),
        'TRÈS BIEN': colors.HexColor('#2e7d32'),
        'BIEN': colors.HexColor('#0277bd'),
        'ASSEZ BIEN': colors.HexColor('#f9a825'),
        'PASSABLE': colors.HexColor('#ef6c00'),
        'INSUFFISANT': colors.HexColor('#c62828'),
        'FAIBLE': colors.HexColor('#b71c1c'),
    }
    
    # ===== FILIGRANE (logo en transparence au centre) =====
    if logo_path:
        try:
            img_reader = ImageReader(logo_path)
            c.saveState()
            c.setFillAlpha(0.08)  # Transparence 8%
            filigrane_size = 14 * cm
            x = (width - filigrane_size) / 2
            y = (height - filigrane_size) / 2
            c.drawImage(img_reader, x, y, width=filigrane_size, height=filigrane_size, 
                       preserveAspectRatio=True, mask='auto')
            c.restoreState()
        except Exception as e:
            print(f"Erreur filigrane: {e}")
    
    # ===== EN-TÊTE =====
    # Logo en haut à gauche
    if logo_path:
        try:
            img_reader = ImageReader(logo_path)
            c.drawImage(img_reader, 1.2*cm, height - 3.2*cm, width=2.5*cm, height=2.5*cm, 
                       preserveAspectRatio=True, mask='auto')
        except:
            pass
    
    # Photo de l'élève en haut à droite (si disponible), sinon drapeau guinéen
    photo_path = bulletin_data.get('photo_path')
    _photo_affichee = False
    if photo_path:
        try:
            import os
            if os.path.exists(photo_path):
                photo_reader = ImageReader(photo_path)
                photo_w = 2.2*cm
                photo_h = 2.5*cm
                photo_x = width - 1.2*cm - photo_w
                photo_y = height - 3.2*cm
                c.drawImage(photo_reader, photo_x, photo_y, width=photo_w, height=photo_h,
                           preserveAspectRatio=True, mask='auto')
                # Bordure fine autour de la photo
                c.setStrokeColor(colors.HexColor('#b3d4fc'))
                c.setLineWidth(0.5)
                c.rect(photo_x, photo_y, photo_w, photo_h, fill=0, stroke=1)
                _photo_affichee = True
        except Exception:
            pass

    if not _photo_affichee:
        # Drapeau guinéen par défaut
        drapeau_x = width - 2.5*cm
        drapeau_y = height - 2.2*cm
        drapeau_w = 1.2*cm
        drapeau_h = 0.8*cm
        c.setFillColor(ROUGE_DRAPEAU)
        c.rect(drapeau_x, drapeau_y, drapeau_w/3, drapeau_h, fill=1, stroke=0)
        c.setFillColor(JAUNE_DRAPEAU)
        c.rect(drapeau_x + drapeau_w/3, drapeau_y, drapeau_w/3, drapeau_h, fill=1, stroke=0)
        c.setFillColor(VERT_DRAPEAU)
        c.rect(drapeau_x + 2*drapeau_w/3, drapeau_y, drapeau_w/3, drapeau_h, fill=1, stroke=0)
        c.setStrokeColor(colors.HexColor('#cccccc'))
        c.rect(drapeau_x, drapeau_y, drapeau_w, drapeau_h, fill=0, stroke=1)
    
    # Textes de l'en-tête
    y_header = height - 1.2*cm
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(width/2, y_header, "RÉPUBLIQUE DE GUINÉE")
    
    # Devise avec couleurs
    y_header -= 0.5*cm
    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(ROUGE_DRAPEAU)
    c.drawCentredString(width/2 - 1.8*cm, y_header, "Travail")
    c.setFillColor(JAUNE_DRAPEAU)
    c.drawCentredString(width/2, y_header, "Justice")
    c.setFillColor(VERT_DRAPEAU)
    c.drawCentredString(width/2 + 1.8*cm, y_header, "Solidarité")
    # Tirets
    c.setFillColor(colors.black)
    c.drawCentredString(width/2 - 0.9*cm, y_header, "-")
    c.drawCentredString(width/2 + 0.9*cm, y_header, "-")
    
    # MEPU-A
    y_header -= 0.45*cm
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 6)
    c.drawCentredString(width/2, y_header, "Ministère de l'Enseignement Pré-Universitaire et de l'Alphabétisation (MEPU-A)")
    
    # Nom de l'école
    y_header -= 0.5*cm
    c.setFont("Helvetica-Bold", 10)
    nom_ecole = ecole.nom.upper() if ecole else "GROUPE SCOLAIRE HADJA KANFING DIANÉ-SONFONIA"
    c.drawCentredString(width/2, y_header, nom_ecole)
    
    # Titre du bulletin
    y_header -= 0.6*cm
    c.setFont("Helvetica-Bold", 12)
    periode_libelle = _formater_periode_libelle(bulletin_data.get('periode', ''))
    c.drawCentredString(width/2, y_header, f"BULLETIN DE NOTES - {periode_libelle}")
    
    # Année scolaire dynamique
    annee_scolaire = ecole.annee_scolaire if hasattr(ecole, 'annee_scolaire') and ecole.annee_scolaire else "2025-2026"
    c.setFont("Helvetica", 7)
    c.setFillColor(colors.HexColor('#666666'))
    c.drawRightString(width - 1.5*cm, y_header, f"Année Scolaire {annee_scolaire}")
    
    # ===== LIGNE DE SÉPARATION NOIRE =====
    y_header -= 0.4*cm
    c.setStrokeColor(colors.black)
    c.setLineWidth(1.5)
    c.line(1.2*cm, y_header, width - 1.2*cm, y_header)
    c.setLineWidth(1)
    
    # ===== INFORMATIONS ÉLÈVE (grille 2x3) =====
    y = y_header - 0.6*cm
    
    # Extraire nom et prénom
    # Utiliser les champs prenom/nom séparés (fiables même pour un prénom
    # composé). Repli sur l'ancien découpage si absents.
    prenom = bulletin_data.get('prenom')
    nom = bulletin_data.get('nom')
    if prenom is None and nom is None:
        eleve_nom_complet = bulletin_data.get('eleve', '')
        parties = eleve_nom_complet.split(' ', 1) if eleve_nom_complet else ['', '']
        prenom = parties[0] if len(parties) > 0 else ''
        nom = parties[1] if len(parties) > 1 else ''
    prenom = prenom or ''
    nom = nom or ''
    
    # Première ligne: PRÉNOM, NOM, MATRICULE (Prénom avant Nom)
    # Deuxième ligne: CLASSE, PÉRIODE, EFFECTIF
    # Formater la période pour l'affichage
    periode_info_indiv = _formater_periode_libelle(bulletin_data.get('periode', '-'))
    
    info_data = [
        [('PRÉNOM', prenom.title()), ('NOM', nom.upper()), ('MATRICULE', bulletin_data.get('matricule', '-'))],
        [('CLASSE', bulletin_data.get('classe', '-')), ('PÉRIODE', periode_info_indiv), ('EFFECTIF', f"{bulletin_data.get('total_eleves', '-')} élèves")],
    ]
    
    # Largeur totale alignée avec la ligne (de 1.2cm à width-1.2cm)
    info_total_width = width - 2.4*cm
    box_width = (info_total_width - 0.4*cm) / 3  # 3 colonnes avec espacement
    box_height = 0.9*cm
    info_margin_left = 1.2*cm
    
    for row_idx, row in enumerate(info_data):
        for col_idx, (label, value) in enumerate(row):
            x = info_margin_left + col_idx * (box_width + 0.2*cm)
            box_y = y - row_idx * (box_height + 0.15*cm)
            
            # Fond bleu clair
            c.setFillColor(BLEU_CLAIR)
            c.rect(x, box_y - box_height, box_width, box_height, fill=1, stroke=0)
            # Bordure
            c.setStrokeColor(colors.HexColor('#b3d4fc'))
            c.rect(x, box_y - box_height, box_width, box_height, fill=0, stroke=1)
            
            # Label en bleu
            c.setFillColor(colors.HexColor('#1976d2'))
            c.setFont("Helvetica-Bold", 7)
            c.drawString(x + 0.15*cm, box_y - 0.3*cm, label)
            
            # Valeur en noir — taille réduite si trop longue pour la case
            c.setFillColor(colors.black)
            val_str = str(value)
            val_fs = 9
            max_val_w = box_width - 0.3*cm
            from reportlab.pdfbase.pdfmetrics import stringWidth as _sw
            while val_fs > 6 and _sw(val_str, "Helvetica-Bold", val_fs) > max_val_w:
                val_fs -= 0.5
            c.setFont("Helvetica-Bold", val_fs)
            c.drawString(x + 0.15*cm, box_y - 0.7*cm, val_str)

    # ===== TABLEAU DES NOTES =====
    y -= 2.3*cm
    
    # Largeur totale du tableau = même que la ligne de séparation (de 1.2cm à width-1.2cm)
    table_total_width = width - 2.4*cm
    margin_left = 1.2*cm
    
    # Déterminer le type de système pour le bulletin individuel
    system_type_indiv = bulletin_data.get('system_type', 'mensuel')
    periode = bulletin_data.get('periode', '')
    
    # Détecter le niveau scolaire
    from notes.calculs_moyennes import detecter_niveau_scolaire
    classe_nom = bulletin_data.get('classe', '')
    niveau_scolaire = detecter_niveau_scolaire(classe_nom)
    est_primaire = (niveau_scolaire == 'PRIMAIRE')
    est_maternelle = (niveau_scolaire == 'MATERNELLE')
    
    # MATERNELLE/GARDERIE: Afficher les appréciations au lieu des notes
    if est_maternelle:
        _dessiner_bulletin_maternelle(c, bulletin_data, width, height, y, ecole)
        # Finaliser le PDF pour la maternelle
        c.showPage()
        c.save()
        buffer.seek(0)
        return buffer
    
    # Déterminer les mois selon la période pour trimestre/semestre
    # SYSTÈME GUINÉEN: Le dernier mois de chaque période = Composition
    # Trimestres: T1(Oct,Nov,Déc), T2(Jan,Fév,Mars), T3(Avr,Mai,Juin)
    # Semestres: S1(Oct,Nov,Déc,Jan,Fév), S2(Mars,Avr,Mai,Juin)
    mois_labels = []
    periodes_labels = []  # Pour bulletin annuel
    
    if system_type_indiv == 'trimestriel':
        if 'TRIMESTRE_1' in periode:
            mois_labels = ['Oct.', 'Nov.']  # Déc. = Compo
        elif 'TRIMESTRE_2' in periode:
            mois_labels = ['Jan.', 'Fév.']  # Mars = Compo
        elif 'TRIMESTRE_3' in periode:
            mois_labels = ['Avr.', 'Mai']   # Juin = Compo
    elif system_type_indiv == 'semestriel':
        if 'SEMESTRE_1' in periode:
            mois_labels = ['Oct.', 'Nov.', 'Déc.', 'Jan.']  # Fév. = Compo
        elif 'SEMESTRE_2' in periode:
            mois_labels = ['Mars', 'Avr.', 'Mai']   # Juin = Compo
    elif system_type_indiv == 'annuel_trimestriel':
        # Bulletin annuel basé sur les trimestres
        periodes_labels = ['1er Trim.', '2ème Trim.', '3ème Trim.']
    elif system_type_indiv == 'annuel_semestriel':
        # Bulletin annuel basé sur les semestres
        periodes_labels = ['1er Sem.', '2ème Sem.']
    
    # Détecter si des notes mensuelles existent (pour masquer les colonnes si seulement compositions)
    has_notes_mensuelles = bulletin_data.get('has_notes_mensuelles', True)  # Par défaut True
    
    # Adapter les colonnes selon le système ET le niveau (primaire = sans COEF ni PTS)
    # Structure: MATIÈRE | [COEF] | Mois1 | Mois2 | [Mois...] | Moy.C | Compo | MOY | [PTS]
    # Si pas de notes mensuelles: MATIÈRE | [COEF] | Compo | MOY | [PTS]
    if system_type_indiv in ['annuel_trimestriel', 'annuel_semestriel'] and periodes_labels:
        # BULLETIN ANNUEL: MATIÈRE | [COEF] | T1/S1 | T2/S2 | [T3] | Moy. Ann. | [PTS]
        if est_primaire:
            header = ['MATIÈRE'] + periodes_labels + ['Moy. Ann.']
        else:
            header = ['MATIÈRE', 'COEF'] + periodes_labels + ['Moy. Ann.', 'PTS']
        data = [header]
        nb_cols = len(header)
    elif system_type_indiv == 'trimestriel':
        if has_notes_mensuelles and mois_labels:
            # Avec notes mensuelles: toutes les colonnes
            if est_primaire:
                header = ['MATIÈRE'] + mois_labels + ['Moy.C', 'Compo', 'MOY']
            else:
                header = ['MATIÈRE', 'COEF'] + mois_labels + ['Moy.C', 'Compo', 'MOY', 'PTS']
        else:
            # Sans notes mensuelles: colonnes simplifiées
            if est_primaire:
                header = ['MATIÈRE', 'Compo', 'MOY']
            else:
                header = ['MATIÈRE', 'COEF', 'Compo', 'MOY', 'PTS']
        data = [header]
        nb_cols = len(header)
    elif system_type_indiv == 'semestriel':
        if has_notes_mensuelles and mois_labels:
            # Avec notes mensuelles: toutes les colonnes
            if est_primaire:
                header = ['MATIÈRE'] + mois_labels + ['Moy.C', 'Compo', 'MOY']
            else:
                header = ['MATIÈRE', 'COEF'] + mois_labels + ['Moy.C', 'Compo', 'MOY', 'PTS']
        else:
            # Sans notes mensuelles: colonnes simplifiées
            if est_primaire:
                header = ['MATIÈRE', 'Compo', 'MOY']
            else:
                header = ['MATIÈRE', 'COEF', 'Compo', 'MOY', 'PTS']
        data = [header]
        nb_cols = len(header)
    else:
        if est_primaire:
            # Primaire Mensuel: MATIÈRE | NOTE | MOY (sans COEF ni PTS)
            data = [['MATIÈRE', 'NOTE', 'MOY']]
            nb_cols = 3
        else:
            # Système mensuel: MATIÈRE | COEF | NOTE | MOY | PTS
            data = [['MATIÈRE', 'COEF', 'NOTE', 'MOY', 'PTS']]
            nb_cols = 5

    total_coef = 0
    total_moy = 0
    total_points = 0
    nb_matieres_avec_moy = 0
    total_moy_continue = 0
    total_note_compo = 0
    nb_moy_continue = 0
    nb_note_compo = 0
    total_note_cours = 0
    nb_note_cours = 0
    # Accumulateurs pour colonnes intermédiaires (mois/périodes)
    totaux_cols_inter = {}
    counts_cols_inter = {}

    for matiere in bulletin_data['matieres']:
        nom_matiere = str(matiere.get('matiere', '-'))
        if hasattr(matiere.get('matiere'), 'nom'):
            nom_matiere = matiere['matiere'].nom
        
        # PRIMAIRE: coefficient = 1 (pas de pondération)
        # SECONDAIRE (Collège/Lycée): coefficient réel de la matière
        coef_brut = float(matiere.get('coefficient', 1))
        coef = 1.0 if est_primaire else coef_brut
        
        moy_continue = matiere.get('moyenne_continue')
        note_compo = matiere.get('note_composition')
        moyenne = matiere.get('moyenne') or matiere.get('moyenne_calculee')
        points = matiere.get('points')
        
        # Recalculer les points avec le bon coefficient
        if points is None and moyenne:
            points = moyenne * coef
        elif est_primaire and moyenne:
            # Forcer le recalcul pour primaire (coef=1)
            points = moyenne * 1.0
        
        # Récupérer les moyennes mensuelles si disponibles
        moyennes_mensuelles = matiere.get('moyennes_mensuelles', [])
        
        total_coef += coef
        if moyenne:
            total_moy += moyenne
            nb_matieres_avec_moy += 1
        if points:
            total_points += points
        if moy_continue:
            total_moy_continue += moy_continue
            nb_moy_continue += 1
        if note_compo:
            total_note_compo += note_compo
            nb_note_compo += 1
        if moy_continue and system_type_indiv == 'mensuel':
            total_note_cours += moy_continue
            nb_note_cours += 1

        if system_type_indiv in ['annuel_trimestriel', 'annuel_semestriel'] and periodes_labels:
            # BULLETIN ANNUEL: Construire la ligne avec les moyennes par période
            if est_primaire:
                row = [nom_matiere]
            else:
                row = [nom_matiere, f"{coef:.0f}"]

            # Ajouter les moyennes par période (T1, T2, T3 ou S1, S2)
            for i, periode_label in enumerate(periodes_labels):
                note_periode = '-'
                val_num = None
                if moyennes_mensuelles and i < len(moyennes_mensuelles):
                    moy_per = moyennes_mensuelles[i]
                    if isinstance(moy_per, dict):
                        val_num = moy_per.get('moyenne')
                        if val_num is not None:
                            note_periode = f"{val_num:.2f}"
                    elif moy_per is not None:
                        val_num = float(moy_per)
                        note_periode = f"{val_num:.2f}"
                if val_num is not None:
                    totaux_cols_inter[i] = totaux_cols_inter.get(i, 0) + val_num
                    counts_cols_inter[i] = counts_cols_inter.get(i, 0) + 1
                row.append(note_periode)

            # Ajouter Moy. Annuelle (et PTS seulement pour collège/lycée)
            row.append(f"{moyenne:.2f}" if moyenne else '-')
            if not est_primaire:
                row.append(f"{points:.2f}" if points else '-')
            data.append(row)
        elif system_type_indiv in ['trimestriel', 'semestriel']:
            # Construire la ligne selon le mode de saisie
            if est_primaire:
                row = [nom_matiere]  # Sans COEF pour primaire
            else:
                row = [nom_matiere, f"{coef:.0f}"]
            
            if has_notes_mensuelles and mois_labels:
                # Ajouter les notes mensuelles
                for i, mois_label in enumerate(mois_labels):
                    note_mois = '-'
                    val_num = None
                    if moyennes_mensuelles and i < len(moyennes_mensuelles):
                        moy_mens = moyennes_mensuelles[i]
                        if isinstance(moy_mens, dict):
                            val_num = moy_mens.get('moyenne')
                            if val_num is not None:
                                note_mois = f"{val_num:.2f}"
                        elif moy_mens is not None:
                            val_num = float(moy_mens)
                            note_mois = f"{val_num:.2f}"
                    if val_num is not None:
                        totaux_cols_inter[i] = totaux_cols_inter.get(i, 0) + val_num
                        counts_cols_inter[i] = counts_cols_inter.get(i, 0) + 1
                    row.append(note_mois)

                # Ajouter Moy.C (seulement si notes mensuelles)
                row.append(f"{moy_continue:.2f}" if moy_continue else '-')
            
            # Compo et MOY sont toujours affichés (PTS seulement pour collège/lycée)
            row.append(f"{note_compo:.2f}" if note_compo else '-')
            row.append(f"{moyenne:.2f}" if moyenne else '-')
            if not est_primaire:
                row.append(f"{points:.2f}" if points else '-')
            data.append(row)
        else:
            # Mensuel
            cours_str = f"{moy_continue:.2f}" if moy_continue else '-'
            moyenne_str = f"{moyenne:.2f}" if moyenne else '-'
            points_str = f"{points:.2f}" if points else '-'
            if est_primaire:
                # Primaire: sans COEF ni PTS
                data.append([nom_matiere, cours_str, moyenne_str])
            else:
                data.append([nom_matiere, f"{coef:.0f}", cours_str, moyenne_str, points_str])
    
    moyenne_officielle = bulletin_data.get('moyenne_generale')
    total_coefficients_officiel = bulletin_data.get('total_coefficients') or total_coef
    if moyenne_officielle is not None and total_coefficients_officiel:
        total_points = round(float(moyenne_officielle) * float(total_coefficients_officiel), 2)
        total_coef = float(total_coefficients_officiel)

    # Ligne TOTAL — avec totaux de toutes les colonnes
    if system_type_indiv in ['annuel_trimestriel', 'annuel_semestriel'] and periodes_labels:
        # BULLETIN ANNUEL: Ligne total
        if est_primaire:
            total_row = ['TOTAL']
        else:
            total_row = ['TOTAL', f"{total_coef:.0f}"]
        # Totaux par période (T1, T2, T3 ou S1, S2)
        for i in range(len(periodes_labels)):
            if i in totaux_cols_inter and counts_cols_inter.get(i, 0) > 0:
                total_row.append(f"{totaux_cols_inter[i]:.2f}")
            else:
                total_row.append('-')
        # Moy. Annuelle
        total_row.append(f"{total_moy:.2f}" if nb_matieres_avec_moy else '-')
        if not est_primaire:
            total_row.append(f"{total_points:.2f}")
        data.append(total_row)
    elif system_type_indiv in ['trimestriel', 'semestriel']:
        if est_primaire:
            total_row = ['TOTAL']
        else:
            total_row = ['TOTAL', f"{total_coef:.0f}"]

        if has_notes_mensuelles and mois_labels:
            # Totaux par mois
            for i in range(len(mois_labels)):
                if i in totaux_cols_inter and counts_cols_inter.get(i, 0) > 0:
                    total_row.append(f"{totaux_cols_inter[i]:.2f}")
                else:
                    total_row.append('-')
            # Total Moy.C
            total_row.append(f"{total_moy_continue:.2f}" if nb_moy_continue else '-')

        # Total Compo
        total_row.append(f"{total_note_compo:.2f}" if nb_note_compo else '-')
        # Total MOY
        total_row.append(f"{total_moy:.0f}" if nb_matieres_avec_moy else '-')
        if not est_primaire:
            total_row.append(f"{total_points:.2f}")
        data.append(total_row)
    else:
        # Mensuel
        note_cours_total = f"{total_note_cours:.2f}" if nb_note_cours else '-'
        if est_primaire:
            data.append(['TOTAL', note_cours_total, f"{total_moy:.0f}" if nb_matieres_avec_moy else '-'])
        else:
            data.append(['TOTAL', f"{total_coef:.0f}", note_cours_total, f"{total_moy:.0f}" if nb_matieres_avec_moy else '-', f"{total_points:.2f}"])
    
    # Calculer les largeurs de colonnes
    if system_type_indiv in ['annuel_trimestriel', 'annuel_semestriel'] and periodes_labels:
        # Bulletin annuel: MATIÈRE 25%, reste divisé équitablement
        col_matiere = table_total_width * 0.22
        nb_autres_cols = nb_cols - 1
        col_autres = (table_total_width - col_matiere) / nb_autres_cols
        col_widths = [col_matiere] + [col_autres] * nb_autres_cols
    elif system_type_indiv in ['trimestriel', 'semestriel']:
        # Adapter la largeur de la colonne MATIÈRE selon le nombre de colonnes
        if has_notes_mensuelles and mois_labels:
            col_matiere = table_total_width * 0.22
        else:
            col_matiere = table_total_width * 0.35  # Plus large si moins de colonnes
        nb_autres_cols = nb_cols - 1
        col_autres = (table_total_width - col_matiere) / nb_autres_cols
        col_widths = [col_matiere] + [col_autres] * nb_autres_cols
    else:
        if est_primaire:
            # Primaire: 4 colonnes (sans COEF)
            col_matiere = table_total_width * 0.40
            col_autres = (table_total_width - col_matiere) / 3
            col_widths = [col_matiere, col_autres, col_autres, col_autres]
        else:
            col_matiere = table_total_width * 0.35
            col_autres = (table_total_width - col_matiere) / 4
            col_widths = [col_matiere, col_autres, col_autres, col_autres, col_autres]
    
    table = Table(data, colWidths=col_widths)
    
    # Adapter la taille de police selon le nombre de matières pour éviter le chevauchement
    nb_matieres = len(data) - 2  # -2 pour en-tête et ligne TOTAL
    if nb_matieres > 15:
        # Beaucoup de matières: police réduite et padding réduit
        font_header = 11
        font_body = 10
        font_total = 11
        padding_top = 3
        padding_bottom = 3
    elif nb_matieres > 12:
        # Nombreuses matières: police moyenne
        font_header = 12
        font_body = 11
        font_total = 12
        padding_top = 4
        padding_bottom = 4
    else:
        # Normal
        font_header = 13
        font_body = 12
        font_total = 13
        padding_top = 5
        padding_bottom = 5
    
    # Style du tableau - base
    style = [
        # En-tête bleu
        ('BACKGROUND', (0, 0), (-1, 0), BLEU_HEADER),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), font_header),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), padding_bottom + 2),
        ('TOPPADDING', (0, 0), (-1, 0), padding_top + 2),
        
        # Corps
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), font_body),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('LEFTPADDING', (0, 1), (0, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), padding_bottom),
        ('TOPPADDING', (0, 1), (-1, -1), padding_top),
        
        # Grille
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        
        # Alternance de couleurs
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')]),
        
        # Ligne TOTAL
        ('BACKGROUND', (0, -1), (-1, -1), GRIS_TOTAL),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), font_total),
    ]
    
    # Colonnes MOY et PTS avec couleurs (position varie selon le système et le niveau)
    if system_type_indiv in ['trimestriel', 'semestriel'] and mois_labels:
        # Colonnes dynamiques: MOY = avant-dernière colonne, PTS = dernière colonne
        col_moy = nb_cols - 2
        col_pts = nb_cols - 1
        style.extend([
            ('BACKGROUND', (col_moy, 1), (col_moy, -2), VERT_MOY),
            ('FONTNAME', (col_moy, 1), (col_moy, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (col_pts, 1), (col_pts, -2), ROUGE_PTS),
            ('FONTNAME', (col_pts, 1), (col_pts, -1), 'Helvetica-Bold'),
            # Taille de police pour les colonnes nombreuses
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ])
    else:
        if est_primaire:
            # Primaire: 4 colonnes: MATIÈRE | NOTE | MOY | PTS (sans COEF)
            style.extend([
                ('BACKGROUND', (2, 1), (2, -2), VERT_MOY),    # MOY = colonne 2
                ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (3, 1), (3, -2), ROUGE_PTS),   # PTS = colonne 3
                ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Bold'),
            ])
        else:
            # 5 colonnes: MATIÈRE | COEF | NOTE | MOY | PTS
            style.extend([
                ('BACKGROUND', (3, 1), (3, -2), VERT_MOY),    # MOY = colonne 3
                ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (4, 1), (4, -2), ROUGE_PTS),   # PTS = colonne 4
                ('FONTNAME', (4, 1), (4, -1), 'Helvetica-Bold'),
            ])
    
    table.setStyle(TableStyle(style))
    
    # Calculer la hauteur du tableau et le dessiner aligné avec la ligne
    table_w, table_h = table.wrap(width, height)
    table.drawOn(c, margin_left, y - table_h)
    
    # ===== SECTION RÉSULTATS =====
    y = y - table_h - 0.8*cm
    
    # Trois colonnes alignées avec le tableau (de 1.2cm à width-1.2cm)
    result_total_width = width - 2.4*cm
    col_width = (result_total_width - 0.4*cm) / 3  # 3 colonnes avec espacement
    col_height = 1.4*cm
    start_x = 1.2*cm
    
    # MOYENNE GÉNÉRALE
    c.setFillColor(colors.white)
    c.rect(start_x, y - col_height, col_width, col_height, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(start_x + col_width/2, y - 0.35*cm, "MOYENNE GÉNÉRALE")
    c.setFont("Helvetica-Bold", 16)
    moy_gen = bulletin_data.get('moyenne_generale')
    # Primaire = notation sur 10, Secondaire = notation sur 20
    base_notation = 10 if est_primaire else 20
    moy_str = f"{moy_gen:.2f}/{base_notation}" if moy_gen else '-'
    c.drawCentredString(start_x + col_width/2, y - 1*cm, moy_str)
    
    # RANG
    rang_x = start_x + col_width + 0.2*cm
    c.setFillColor(colors.white)
    c.rect(rang_x, y - col_height, col_width, col_height, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(rang_x + col_width/2, y - 0.35*cm, "RANG")
    c.setFont("Helvetica-Bold", 16)
    rang = bulletin_data.get('rang')
    # Afficher "-" si rang est None ou vide
    rang_str = str(rang) if rang is not None else '-'
    c.drawCentredString(rang_x + col_width/2, y - 1*cm, rang_str)
    
    # MENTION (avec badge coloré)
    mention_x = start_x + 2*(col_width + 0.2*cm)
    c.setFillColor(colors.white)
    c.rect(mention_x, y - col_height, col_width, col_height, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(mention_x + col_width/2, y - 0.35*cm, "MENTION")
    
    # Badge de mention
    mention = bulletin_data.get('mention', '-')
    mention_upper = mention.upper() if mention else '-'
    mention_color = MENTION_COLORS.get(mention_upper, colors.HexColor('#666666'))
    
    # Dessiner le badge
    badge_width = min(4*cm, col_width - 0.4*cm)
    badge_height = 0.6*cm
    badge_x = mention_x + (col_width - badge_width) / 2
    badge_y = y - 1.15*cm
    
    c.setFillColor(mention_color)
    c.roundRect(badge_x, badge_y, badge_width, badge_height, 3, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(mention_x + col_width/2, badge_y + 0.15*cm, mention_upper)
    
    # ===== APPRÉCIATION =====
    y -= col_height + 0.6*cm
    
    appreciation_height = 1.2*cm
    c.setFillColor(BLEU_APPRECIATION)
    c.rect(1.2*cm, y - appreciation_height, width - 2.4*cm, appreciation_height, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor('#90caf9'))
    c.rect(1.2*cm, y - appreciation_height, width - 2.4*cm, appreciation_height, fill=0, stroke=1)
    
    c.setFillColor(colors.HexColor('#1565c0'))
    c.setFont("Helvetica-Bold", 9)
    c.drawString(1.4*cm, y - 0.35*cm, "APPRÉCIATION DU CONSEIL DE CLASSE")
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Oblique", 8)
    # Utiliser l'appréciation intelligente si non fournie
    appreciation = bulletin_data.get('appreciation')
    if not appreciation:
        moyenne_gen = bulletin_data.get('moyenne_generale')
        niveau = bulletin_data.get('niveau_scolaire', 'SECONDAIRE')
        eleve_nom = bulletin_data.get('eleve', '').split()[0] if bulletin_data.get('eleve') else 'L\'élève'
        appreciation = obtenir_appreciation_intelligente(moyenne_gen, eleve_nom, niveau) if moyenne_gen else 'Aucune note disponible pour cette période.'
    if appreciation and len(appreciation) > 100:
        appreciation = appreciation[:97] + '...'
    c.drawString(1.4*cm, y - 0.85*cm, appreciation)
    
    # ===== SIGNATURES =====
    y -= appreciation_height + 0.8*cm
    
    sig_width = 6*cm
    
    # Censeur / Directeur du primaire (aligné à gauche avec 1.2cm)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 9)
    titre_signataire = "Directeur du primaire" if est_primaire else "Censeur de l'établissement"
    c.drawCentredString(1.2*cm + sig_width/2, y, titre_signataire)
    c.setStrokeColor(colors.black)
    c.line(1.2*cm + 0.5*cm, y - 1.5*cm, 1.2*cm + sig_width - 0.5*cm, y - 1.5*cm)
    c.setFont("Helvetica", 8)
    c.drawCentredString(1.2*cm + sig_width/2, y - 1.8*cm, "Signature")
    
    # Directeur Général (aligné à droite avec 1.2cm)
    dir_x = width - 1.2*cm - sig_width
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(dir_x + sig_width/2, y, "Directeur Général")
    c.line(dir_x + 0.5*cm, y - 1.5*cm, dir_x + sig_width - 0.5*cm, y - 1.5*cm)
    c.setFont("Helvetica", 8)
    c.drawCentredString(dir_x + sig_width/2, y - 1.8*cm, "Signature")
    
    # ===== MÉTHODE DE CALCUL =====
    y -= 2.5*cm
    
    c.setFillColor(colors.HexColor('#f5f5f5'))
    c.rect(1.2*cm, y - 1.8*cm, width - 2.4*cm, 1.8*cm, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor('#1976d2'))
    c.setLineWidth(2)
    c.line(1.2*cm, y, 1.2*cm, y - 1.8*cm)
    c.setLineWidth(1)
    
    c.setFillColor(colors.HexColor('#1976d2'))
    c.setFont("Helvetica-Bold", 8)
    
    # Adapter le texte selon le système (pour bulletin individuel)
    system_type_indiv = bulletin_data.get('system_type', 'mensuel')
    
    # Texte pour les points (différent pour primaire)
    points_text = "• Points : Moyenne (pas de coefficient pour le primaire)" if est_primaire else "• Points : Moyenne × Coefficient"
    
    if system_type_indiv == 'annuel_trimestriel':
        titre_systeme = "MÉTHODE DE CALCUL - BULLETIN ANNUEL (TRIMESTRES)" if not est_primaire else "MÉTHODE DE CALCUL - BULLETIN ANNUEL (PRIMAIRE)"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• T1, T2, T3 : Moyennes des 1er, 2ème et 3ème trimestres")
        c.drawString(1.4*cm, y - 0.9*cm, "• Moy. Ann. : (T1 + T2 + T3) / 3 (moyenne des 3 trimestres)")
        c.drawString(1.4*cm, y - 1.2*cm, points_text)
    elif system_type_indiv == 'annuel_semestriel':
        titre_systeme = "MÉTHODE DE CALCUL - BULLETIN ANNUEL (SEMESTRES)" if not est_primaire else "MÉTHODE DE CALCUL - BULLETIN ANNUEL (PRIMAIRE)"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• S1, S2 : Moyennes des 1er et 2ème semestres")
        c.drawString(1.4*cm, y - 0.9*cm, "• Moy. Ann. : (S1 + S2) / 2 (moyenne des 2 semestres)")
        c.drawString(1.4*cm, y - 1.2*cm, points_text)
    elif system_type_indiv == 'trimestriel':
        titre_systeme = "MÉTHODE DE CALCUL DES NOTES - SYSTÈME TRIMESTRIEL (PRIMAIRE)" if est_primaire else "MÉTHODE DE CALCUL DES NOTES - SYSTÈME TRIMESTRIEL"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• Cours : Moyenne des notes d'évaluations continues du trimestre")
        c.drawString(1.4*cm, y - 0.9*cm, "• Compo : Note de composition du trimestre")
        c.drawString(1.4*cm, y - 1.2*cm, f"• {'Moyenne : composition' if est_primaire else 'Moyenne : 40% cours + 60% compo'} | {points_text}")
    elif system_type_indiv == 'semestriel':
        titre_systeme = "MÉTHODE DE CALCUL DES NOTES - SYSTÈME SEMESTRIEL (PRIMAIRE)" if est_primaire else "MÉTHODE DE CALCUL DES NOTES - SYSTÈME SEMESTRIEL"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• Cours : Moyenne des notes d'évaluations continues du semestre")
        c.drawString(1.4*cm, y - 0.9*cm, "• Compo : Note de composition (examen) du semestre")
        c.drawString(1.4*cm, y - 1.2*cm, f"• {'Moyenne : composition' if est_primaire else 'Moyenne : 40% cours + 60% compo'} | {points_text}")
    else:
        titre_systeme = "MÉTHODE DE CALCUL DES NOTES - SYSTÈME MENSUEL (PRIMAIRE)" if est_primaire else "MÉTHODE DE CALCUL DES NOTES - SYSTÈME MENSUEL"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• Note : Moyenne des notes d'évaluations continues du mois")
        c.drawString(1.4*cm, y - 0.9*cm, "• Moyenne : Égale à la note du cours pour le système mensuel")
        c.drawString(1.4*cm, y - 1.2*cm, points_text)
    
    # Adapter les mentions selon le niveau scolaire
    if est_primaire:
        c.drawString(1.4*cm, y - 1.5*cm, "• Mentions : Excellent ≥9 | Très bien ≥8 | Bien ≥7 | Assez bien ≥6 | Passable ≥5 | Insuffisant ≥4 | Faible ≥3 | Très faible <3")
    else:
        c.drawString(1.4*cm, y - 1.5*cm, "• Mentions : Excellent ≥18 | Très bien ≥16 | Bien ≥14 | Assez bien ≥12 | Passable ≥10 | Insuffisant ≥8 | Faible ≥6 | Très faible <6")
    
    # ===== PIED DE PAGE avec infos dynamiques de l'école =====
    c.setFillColor(colors.HexColor('#999999'))
    c.setFont("Helvetica", 6)
    # Construire le footer avec les infos de l'école
    footer_parts = [f"© 2025 {ecole.nom if ecole else 'Myschool'}. Tous droits réservés."]
    if ecole and ecole.telephone:
        footer_parts.append(f"Tél: {ecole.tous_telephones}")
    if ecole and ecole.email:
        footer_parts.append(ecole.email)
    c.drawCentredString(width/2, 0.8*cm, " | ".join(footer_parts))
    c.drawCentredString(width/2, 0.5*cm, 
                       f"Bulletin généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    
    c.showPage()
    c.save()
    
    buffer.seek(0)
    return buffer


def _dessiner_bulletin_maternelle(c, bulletin_data, width, height, y, ecole):
    """Dessine un bulletin spécifique pour la maternelle/garderie avec appréciations"""
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import Table, TableStyle
    
    # Couleurs
    BLEU_HEADER = colors.HexColor('#4a90d9')
    BLEU_CLAIR = colors.HexColor('#e8f4fd')
    VERT_ACQUIS = colors.HexColor('#c8e6c9')
    JAUNE_ENCOURS = colors.HexColor('#fff9c4')
    ROUGE_NONACQUIS = colors.HexColor('#ffcdd2')
    GRIS_TOTAL = colors.HexColor('#37474f')
    
    table_total_width = width - 2.4*cm
    margin_left = 1.2*cm
    
    # Récupérer les matières/appréciations depuis bulletin_data
    matieres = bulletin_data.get('matieres', [])
    
    # En-tête du tableau
    header = ['DOMAINE D\'ACTIVITÉ', 'APPRÉCIATION', 'COMMENTAIRE']
    data = [header]
    
    # Mapping des appréciations
    APPRECIATION_DISPLAY = {
        'TRES_BIEN_ACQUIS': 'Très Bien Acquis',
        'BIEN_ACQUIS': 'Bien Acquis',
        'EN_COURS': 'En Cours d\'Acquisition',
        'NON_ACQUIS': 'Non Acquis',
    }
    
    # Observations automatiques
    OBSERVATIONS_AUTO = {
        'TRES_BIEN_ACQUIS': 'Excellent travail !',
        'BIEN_ACQUIS': 'Bon niveau',
        'EN_COURS': 'Persévère !',
        'NON_ACQUIS': 'À renforcer',
    }
    
    # Afficher les matières avec leurs appréciations
    if matieres:
        for mat in matieres:
            # Récupérer le nom de la matière
            matiere_nom = mat.get('matiere', '-')
            if isinstance(matiere_nom, str):
                nom = matiere_nom
            elif hasattr(matiere_nom, 'nom'):
                nom = matiere_nom.nom
            else:
                nom = str(matiere_nom)
            
            # Récupérer l'appréciation
            appreciation = mat.get('appreciation')
            appreciation_display_val = mat.get('appreciation_display', '-')
            commentaire = mat.get('commentaire', '')
            absent = mat.get('absent', False)
            
            if absent:
                appreciation_display = 'Absent'
                commentaire = 'Absent(e) lors de l\'évaluation'
            elif appreciation:
                appreciation_display = APPRECIATION_DISPLAY.get(appreciation, appreciation_display_val)
                if not commentaire or commentaire == '-':
                    commentaire = OBSERVATIONS_AUTO.get(appreciation, '-')
            else:
                appreciation_display = '-'
                commentaire = '-'
            
            data.append([nom, appreciation_display, commentaire])
    else:
        data.append(['Aucune activité', '-', '-'])
    
    # Largeurs de colonnes
    col_widths = [table_total_width * 0.35, table_total_width * 0.30, table_total_width * 0.35]
    
    table = Table(data, colWidths=col_widths)
    
    # Styles du tableau
    style = [
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), BLEU_HEADER),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Corps - Matière et Appréciation
        ('FONTNAME', (0, 1), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (1, -1), 9),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('LEFTPADDING', (0, 1), (0, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        
        # Commentaires - Police plus petite pour éviter débordement
        ('FONTNAME', (2, 1), (2, -1), 'Helvetica'),
        ('FONTSIZE', (2, 1), (2, -1), 7),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('LEFTPADDING', (2, 1), (2, -1), 4),
        
        # Grille
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        
        # Alternance de couleurs
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ]
    
    # Colorer les appréciations selon leur valeur
    for i, row in enumerate(data[1:], start=1):
        appreciation = row[1]
        if 'Très Bien' in appreciation:
            style.append(('BACKGROUND', (1, i), (1, i), VERT_ACQUIS))
        elif 'Bien Acquis' in appreciation:
            style.append(('BACKGROUND', (1, i), (1, i), colors.HexColor('#a5d6a7')))
        elif 'En Cours' in appreciation:
            style.append(('BACKGROUND', (1, i), (1, i), JAUNE_ENCOURS))
        elif 'Non Acquis' in appreciation:
            style.append(('BACKGROUND', (1, i), (1, i), ROUGE_NONACQUIS))
    
    table.setStyle(TableStyle(style))
    
    # Dessiner le tableau
    table_w, table_h = table.wrap(width, height)
    table.drawOn(c, margin_left, y - table_h)
    
    # ===== SECTION RÉSULTATS (Taux, Rang, Appréciation) =====
    y = y - table_h - 0.8*cm
    
    # Récupérer les données
    taux_acquisition = bulletin_data.get('moyenne_generale')
    rang = bulletin_data.get('rang')
    mention = bulletin_data.get('mention') or 'Suivi continu'
    
    # Couleurs des cartes
    VERT_TAUX = colors.HexColor('#27ae60')
    BLEU_RANG = colors.HexColor('#3498db')
    VIOLET_APPRECIATION = colors.HexColor('#9b59b6')
    
    # Largeur de chaque carte (3 cartes)
    card_width = (table_total_width - 0.4*cm) / 3
    card_height = 1.8*cm
    
    # === Carte 1: TAUX D'ACQUISITION ===
    card_x = margin_left
    c.setFillColor(VERT_TAUX)
    c.roundRect(card_x, y - card_height, card_width, card_height, 5, fill=1, stroke=0)
    
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(card_x + card_width/2, y - 0.4*cm, "TAUX D'ACQUISITION")
    
    c.setFont("Helvetica-Bold", 16)
    taux_str = f"{taux_acquisition:.1f}%" if taux_acquisition else "-"
    c.drawCentredString(card_x + card_width/2, y - 1.2*cm, taux_str)
    
    # === Carte 2: RANG ===
    card_x = margin_left + card_width + 0.2*cm
    c.setFillColor(BLEU_RANG)
    c.roundRect(card_x, y - card_height, card_width, card_height, 5, fill=1, stroke=0)
    
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(card_x + card_width/2, y - 0.4*cm, "RANG")
    
    c.setFont("Helvetica-Bold", 16)
    rang_str = str(rang) if rang else "-"
    c.drawCentredString(card_x + card_width/2, y - 1.2*cm, rang_str)
    
    # === Carte 3: APPRÉCIATION ===
    card_x = margin_left + 2*(card_width + 0.2*cm)
    c.setFillColor(VIOLET_APPRECIATION)
    c.roundRect(card_x, y - card_height, card_width, card_height, 5, fill=1, stroke=0)
    
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(card_x + card_width/2, y - 0.4*cm, "APPRÉCIATION")
    
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(card_x + card_width/2, y - 1.2*cm, str(mention))
    
    # Ajuster y pour la suite
    y = y - card_height - 0.3*cm
    
    # ===== SIGNATURES =====
    y -= 1.5*cm
    
    sig_width = 6*cm
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(1.2*cm + sig_width/2, y, "L'Éducateur(trice)")
    c.setStrokeColor(colors.black)
    c.line(1.2*cm + 0.5*cm, y - 1.5*cm, 1.2*cm + sig_width - 0.5*cm, y - 1.5*cm)
    c.setFont("Helvetica", 8)
    c.drawCentredString(1.2*cm + sig_width/2, y - 1.8*cm, "Signature")
    
    dir_x = width - 1.2*cm - sig_width
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(dir_x + sig_width/2, y, "Le Directeur / La Directrice")
    c.line(dir_x + 0.5*cm, y - 1.5*cm, dir_x + sig_width - 0.5*cm, y - 1.5*cm)
    c.setFont("Helvetica", 8)
    c.drawCentredString(dir_x + sig_width/2, y - 1.8*cm, "Signature")
    
    # ===== MÉTHODE D'ÉVALUATION =====
    y -= 2.5*cm
    
    c.setFillColor(colors.HexColor('#f5f5f5'))
    c.rect(1.2*cm, y - 1.5*cm, table_total_width, 1.5*cm, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor('#1976d2'))
    c.setLineWidth(2)
    c.line(1.2*cm, y, 1.2*cm, y - 1.5*cm)
    c.setLineWidth(1)
    
    c.setFillColor(colors.HexColor('#1976d2'))
    c.setFont("Helvetica-Bold", 8)
    c.drawString(1.4*cm, y - 0.3*cm, "MÉTHODE D'ÉVALUATION - MATERNELLE / GARDERIE")
    
    c.setFillColor(colors.HexColor('#333333'))
    c.setFont("Helvetica", 7)
    c.drawString(1.4*cm, y - 0.6*cm, "• Évaluation qualitative par compétences (pas de notes numériques)")
    c.drawString(1.4*cm, y - 0.9*cm, "• Appréciations : Très Bien Acquis | Bien Acquis | En Cours d'Acquisition | Non Acquis")
    c.drawString(1.4*cm, y - 1.2*cm, "• Suivi individualisé du développement de l'enfant")
    
    # ===== PIED DE PAGE =====
    c.setFillColor(colors.HexColor('#999999'))
    c.setFont("Helvetica", 6)
    footer_parts = [f"© 2025 {ecole.nom if ecole else 'Myschool'}. Tous droits réservés."]
    if ecole and ecole.telephone:
        footer_parts.append(f"Tél: {ecole.tous_telephones}")
    if ecole and ecole.email:
        footer_parts.append(ecole.email)
    c.drawCentredString(width/2, 0.8*cm, " | ".join(footer_parts))


def generer_excel(bulletin_data):
    """Génère un fichier Excel du bulletin"""
    if not EXCEL_AVAILABLE:
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulletin"
    
    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-tête
    ws.merge_cells('A1:F1')
    ws['A1'] = "BULLETIN DE NOTES"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Informations
    ws['A3'] = f"Élève: {bulletin_data['eleve']}"
    ws['A3'].font = subtitle_font
    
    ws['A4'] = f"Classe: {bulletin_data['classe']}"
    ws['D4'] = f"Période: {bulletin_data['periode']}"
    
    # En-têtes du tableau
    row = 6
    headers = ['Matière', 'Coefficient', 'Moy. Cours', 'Composition', 'Moyenne', 'Points']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Données
    row += 1
    for matiere in bulletin_data['matieres']:
        # Moyenne cours
        moy_cours = ''
        if matiere['notes_mensuelles']:
            notes_list = []
            for notes in matiere['notes_mensuelles'].values():
                notes_list.extend(notes)
            if notes_list:
                moy_cours = float(sum(notes_list) / len(notes_list))
        
        composition = float(matiere['composition']) if matiere['composition'] else ''
        moyenne = float(matiere['moyenne']) if matiere['moyenne'] else ''
        points = float(matiere['moyenne'] * matiere['coefficient']) if matiere['moyenne'] else ''
        
        ws.cell(row=row, column=1, value=matiere['matiere']).border = border
        ws.cell(row=row, column=2, value=matiere['coefficient']).border = border
        ws.cell(row=row, column=3, value=moy_cours).border = border
        ws.cell(row=row, column=4, value=composition).border = border
        ws.cell(row=row, column=5, value=moyenne).border = border
        ws.cell(row=row, column=6, value=points).border = border
        
        row += 1
    
    # Résultats
    row += 2
    if bulletin_data['moyenne_generale']:
        ws.cell(row=row, column=1, value="Moyenne Générale:").font = subtitle_font
        ws.cell(row=row, column=2, value=float(bulletin_data['moyenne_generale']))
        
        row += 1
        ws.cell(row=row, column=1, value="Rang:").font = subtitle_font
        rang_affiche = bulletin_data['rang'] if bulletin_data['rang'] else "-"
        ws.cell(row=row, column=2, value=rang_affiche)
        
        row += 1
        ws.cell(row=row, column=1, value="Mention:").font = subtitle_font
        ws.cell(row=row, column=2, value=bulletin_data['mention'])
        
        row += 1
        ws.cell(row=row, column=1, value="Appréciation:").font = subtitle_font
        ws.merge_cells(f'B{row}:F{row}')
        ws.cell(row=row, column=2, value=bulletin_data['appreciation'])
    
    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # Sauvegarder dans un buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_intelligent_view(request, eleve_id, classe_note_id, periode):
    """Vue pour afficher le bulletin intelligent"""
    eleve = get_object_or_404(Eleve, pk=eleve_id)
    classe_note = get_object_or_404(ClasseNote, pk=classe_note_id)
    
    # Déterminer le système
    systeme = 'SEMESTRE' if 'SEMESTRE' in periode else 'TRIMESTRE'
    
    # Détecter le niveau scolaire (primaire ou secondaire)
    niveau_scolaire = detecter_niveau_scolaire(classe_note.nom)
    est_primaire = (niveau_scolaire == 'PRIMAIRE')
    base_notation = 10 if est_primaire else 20
    
    # Calculer le bulletin
    calculateur = CalculateurBulletinIntelligent(eleve, classe_note, periode, systeme)
    bulletin_data = calculateur.generer_bulletin()
    
    return render(request, 'notes/bulletin_intelligent.html', {
        'bulletin': bulletin_data,
        'eleve': eleve,
        'classe_note': classe_note,
        'periode': periode,
        'est_primaire': est_primaire,
        'base_notation': base_notation
    })


@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_intelligent_pdf(request, eleve_id, classe_note_id, periode):
    """Génère le bulletin en PDF avec filigrane"""
    eleve = get_object_or_404(Eleve, pk=eleve_id)
    classe_note = get_object_or_404(ClasseNote, pk=classe_note_id)
    
    # Déterminer le système et le type de système pour l'affichage
    systeme = 'SEMESTRE' if 'SEMESTRE' in periode else 'TRIMESTRE'
    system_type = 'mensuel'
    
    # Détecter le type de bulletin (mensuel, trimestriel, semestriel, annuel)
    if 'ANNUEL_TRIM' in periode:
        system_type = 'annuel_trimestriel'
        systeme = 'TRIMESTRE'
    elif 'ANNUEL_SEM' in periode:
        system_type = 'annuel_semestriel'
        systeme = 'SEMESTRE'
    elif 'TRIMESTRE' in periode:
        system_type = 'trimestriel'
    elif 'SEMESTRE' in periode:
        system_type = 'semestriel'
    
    # Calculer le bulletin
    calculateur = CalculateurBulletinIntelligent(eleve, classe_note, periode, systeme)
    bulletin_data = calculateur.generer_bulletin()
    
    # Ajouter le system_type et la période
    bulletin_data['system_type'] = system_type
    bulletin_data['periode'] = periode
    
    # Détecter si la classe a des notes mensuelles ou seulement des compositions
    from notes.calculs_moyennes import calculer_bulletin_intelligent, detecter_notes_mensuelles_classe
    detection_notes = detecter_notes_mensuelles_classe(classe_note, periode)
    bulletin_data['has_notes_mensuelles'] = detection_notes['has_notes_mensuelles']
    
    # Enrichir avec les données détaillées selon le type de système
    
    matieres_enrichies = []
    for mat_data in bulletin_data.get('matieres', []):
        mat_obj = mat_data.get('matiere')
        if mat_obj:
            try:
                # Utiliser la fonction centralisée intelligente
                result = calculer_bulletin_intelligent(eleve, mat_obj, periode, system_type)
                mat_data['moyennes_mensuelles'] = result.get('moyennes_mensuelles', [])
                mat_data['note_composition'] = result.get('note_composition')
                mat_data['moyenne_continue'] = result.get('moyenne_continue')
                mat_data['moyenne'] = result.get('moyenne')
                mat_data['points'] = result.get('points')
            except:
                pass
        matieres_enrichies.append(mat_data)
    
    bulletin_data['matieres'] = matieres_enrichies
    
    # ===== RECALCULER LA MOYENNE GÉNÉRALE ET LES TOTAUX POUR BULLETINS ANNUELS =====
    # IMPORTANT: Utiliser UNE SEULE SOURCE pour garantir la cohérence moyenne/rang
    if system_type in ['annuel_trimestriel', 'annuel_semestriel']:
        from notes.calculs_moyennes import calculer_classement_classe, detecter_niveau_scolaire
        from eleves.models import Classe as ClasseEleve
        
        # Récupérer les matières de la classe
        matieres = MatiereNote.objects.filter(classe=classe_note)
        
        # Récupérer tous les élèves de la classe
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe_note.nom,
            annee_scolaire=classe_note.annee_scolaire,
            ecole=classe_note.ecole
        ).first()
        
        if classe_eleve:
            eleves_classe = list(Eleve.objects.filter(classe=classe_eleve, statut='ACTIF'))
            total_eleves = len(eleves_classe)
            bulletin_data['total_eleves'] = total_eleves
            
            # CALCUL UNIQUE: Le classement calcule les moyennes ET les rangs avec la même source
            classement_result = calculer_classement_classe(eleves_classe, matieres, periode, system_type)
            rang_map = classement_result.get('rang_map', {})
            details_map = classement_result.get('details_par_eleve', {})
            
            # Récupérer les données de l'élève depuis le classement (MÊME SOURCE)
            if eleve.id in details_map:
                details_eleve = details_map[eleve.id]
                bulletin_data['moyenne_generale'] = details_eleve.get('moyenne_generale')
                bulletin_data['total_points'] = details_eleve.get('total_points', 0)
                bulletin_data['total_coefficients'] = details_eleve.get('total_coefficients', 0)
            
            # Récupérer le rang de l'élève (MÊME SOURCE que la moyenne)
            rang_brut = rang_map.get(eleve.id, '-')
            
            # Formater le rang avec accord grammatical
            if rang_brut and rang_brut != '-':
                sexe = getattr(eleve, 'sexe', 'M')
                rang_formate = formater_rang_intelligent(rang_brut, sexe, total_eleves)
                bulletin_data['rang'] = rang_formate
            else:
                bulletin_data['rang'] = '-'
        
        # Calculer la mention basée sur la moyenne (cohérente)
        niveau = detecter_niveau_scolaire(classe_note.nom)
        if bulletin_data.get('moyenne_generale'):
            bulletin_data['mention'] = obtenir_mention_intelligente(bulletin_data['moyenne_generale'], niveau)
    
    # Chemin du logo
    ecole = eleve.classe.ecole
    logo_path = None
    if hasattr(ecole, 'logo') and ecole.logo:
        logo_path = ecole.logo.path
    
    # Ajouter le matricule et la photo aux données du bulletin
    bulletin_data['matricule'] = eleve.matricule
    if hasattr(eleve, 'photo') and eleve.photo:
        try:
            bulletin_data['photo_path'] = eleve.photo.path
        except Exception:
            bulletin_data['photo_path'] = None
    else:
        bulletin_data['photo_path'] = None

    # Générer le PDF avec l'école
    pdf_buffer = generer_pdf_avec_filigrane(bulletin_data, logo_path, ecole)
    
    # Réponse HTTP
    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    filename = f"bulletin_{eleve.nom}_{eleve.prenom}_{periode}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_intelligent_excel(request, eleve_id, classe_note_id, periode):
    """Génère le bulletin en Excel"""
    if not EXCEL_AVAILABLE:
        return HttpResponse("Excel export n'est pas disponible", status=500)
    
    eleve = get_object_or_404(Eleve, pk=eleve_id)
    classe_note = get_object_or_404(ClasseNote, pk=classe_note_id)
    
    # Déterminer le système
    systeme = 'SEMESTRE' if 'SEMESTRE' in periode else 'TRIMESTRE'
    
    # Calculer le bulletin
    calculateur = CalculateurBulletinIntelligent(eleve, classe_note, periode, systeme)
    bulletin_data = calculateur.generer_bulletin()
    
    # Générer l'Excel
    excel_buffer = generer_excel(bulletin_data)
    
    if not excel_buffer:
        return HttpResponse("Erreur lors de la génération Excel", status=500)
    
    # Réponse HTTP
    response = HttpResponse(
        excel_buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"bulletin_{eleve.nom}_{eleve.prenom}_{periode}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def bulletins_classe_pdf(request, classe_note_id, periode):
    """Génère tous les bulletins d'une classe en un seul PDF - VERSION OPTIMISÉE"""
    import re
    
    classe_note = get_object_or_404(ClasseNote, pk=classe_note_id)
    
    # Récupérer tous les élèves de la classe
    classe_eleve = Classe.objects.filter(
        nom=classe_note.nom,
        annee_scolaire=classe_note.annee_scolaire,
        ecole=classe_note.ecole
    ).first()
    
    if not classe_eleve:
        return HttpResponse("Classe non trouvée", status=404)
    
    eleves = list(Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('nom', 'prenom'))
    
    if not eleves:
        return HttpResponse("Aucun élève dans cette classe", status=404)
    
    # Déterminer le système et le type de système
    systeme = 'SEMESTRE' if 'SEMESTRE' in periode else 'TRIMESTRE'
    system_type = 'mensuel'
    
    # Détecter le type de bulletin (mensuel, trimestriel, semestriel, annuel)
    if 'ANNUEL_TRIM' in periode:
        system_type = 'annuel_trimestriel'
        systeme = 'TRIMESTRE'
    elif 'ANNUEL_SEM' in periode:
        system_type = 'annuel_semestriel'
        systeme = 'SEMESTRE'
    elif 'TRIMESTRE' in periode:
        system_type = 'trimestriel'
    elif 'SEMESTRE' in periode:
        system_type = 'semestriel'
    
    # Chemin du logo
    ecole = classe_note.ecole
    logo_path = None
    if hasattr(ecole, 'logo') and ecole.logo:
        logo_path = ecole.logo.path
    
    # ===== OPTIMISATION: Pré-calculer le classement pour tous les élèves =====
    from notes.calculs_moyennes import calculer_classement_classe, detecter_notes_mensuelles_classe
    from notes.utils_rangs import calculer_rangs_classe_periode
    
    # Récupérer les matières de la classe (MatiereNote.classe est une FK vers ClasseNote)
    matieres = MatiereNote.objects.filter(classe=classe_note)
    
    # Détecter si la classe a des notes mensuelles ou seulement des compositions
    detection_notes = detecter_notes_mensuelles_classe(classe_note, periode)
    has_notes_mensuelles = detection_notes['has_notes_mensuelles']
    
    # Calculer le classement officiel une seule fois pour toute la classe.
    # C'est la même source que les exports notes_completes et resultats.
    rangs_officiels = calculer_rangs_classe_periode(classe_note, periode, use_cache=False)

    # ORDRE DE MÉRITE: trier les bulletins par rang (du 1er au dernier).
    def _rang_key(e):
        r = rangs_officiels.get(e.id, {}).get('rang_num')
        return r if r is not None else 9999
    eleves.sort(key=_rang_key)

    # Garder ce calcul pour obtenir les détails par matière du bulletin.
    classement_result = calculer_classement_classe(eleves, matieres, periode, system_type)
    
    # Extraire les données pré-calculées
    moyennes_map = classement_result.get('moyennes_par_eleve', {})
    details_map = classement_result.get('details_par_eleve', {})
    
    total_eleves = len(eleves)
    
    # Pré-charger le logo une seule fois
    logo_reader = None
    if logo_path:
        try:
            logo_reader = ImageReader(logo_path)
        except:
            pass
    
    # ===== Générer un seul PDF multi-pages =====
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    
    # Import pour les détails mensuels et la fonction centralisée
    from notes.calculs_moyennes import detecter_niveau_scolaire, calculer_bulletin_intelligent
    
    # Détecter le niveau scolaire
    niveau_scolaire = detecter_niveau_scolaire(classe_note.nom)
    est_maternelle = (niveau_scolaire == 'MATERNELLE')
    
    for idx, eleve in enumerate(eleves):
        try:
            # Pour la maternelle, toujours utiliser le calculateur spécifique
            if est_maternelle:
                calculateur = CalculateurBulletinIntelligent(eleve, classe_note, periode, systeme)
                bulletin_data = calculateur.generer_bulletin()
                bulletin_data['matricule'] = eleve.matricule
                bulletin_data['total_eleves'] = total_eleves
                bulletin_data['system_type'] = system_type
            # Utiliser les données pré-calculées si disponibles
            elif eleve.id in details_map:
                details = details_map[eleve.id]
                matieres_data = details.get('details_matieres', [])
                
                # Enrichir avec les données détaillées (fonctionne pour tous les types)
                if matieres:
                    matieres_enrichies = []
                    for mat_data in matieres_data:
                        mat_obj = mat_data.get('matiere')
                        if mat_obj:
                            # Utiliser la fonction centralisée intelligente
                            try:
                                result = calculer_bulletin_intelligent(eleve, mat_obj, periode, system_type)
                                mat_data['moyennes_mensuelles'] = result.get('moyennes_mensuelles', [])
                                mat_data['note_composition'] = result.get('note_composition')
                                mat_data['moyenne_continue'] = result.get('moyenne_continue')
                                mat_data['moyenne'] = result.get('moyenne')
                                mat_data['points'] = result.get('points')
                            except:
                                pass
                        matieres_enrichies.append(mat_data)
                    matieres_data = matieres_enrichies
                
                # Utiliser le rang officiel pour rester aligné avec les exports.
                rang_info = rangs_officiels.get(eleve.id)
                if rang_info:
                    rang_formate = rang_info['rang']
                else:
                    rang_formate = '-'
                
                moyenne_gen = float(rang_info['moyenne']) if rang_info else details.get('moyenne_generale')
                total_coefficients = details.get('total_coefficients')
                total_points = (
                    round(float(moyenne_gen) * float(total_coefficients), 2)
                    if moyenne_gen is not None and total_coefficients
                    else details.get('total_points')
                )
                bulletin_data = {
                    'eleve': f"{eleve.prenom} {eleve.nom}",
                    'prenom': eleve.prenom or '',
                    'nom': eleve.nom or '',
                    'classe': classe_note.nom,
                    'periode': periode,
                    'system_type': system_type,
                    'matieres': matieres_data,
                    'moyenne_generale': moyenne_gen,
                    'total_points': total_points,
                    'total_coefficients': total_coefficients,
                    'rang': rang_formate,
                    'mention': obtenir_mention_intelligente(moyenne_gen, niveau_scolaire),
                    'appreciation': obtenir_appreciation_intelligente(moyenne_gen, eleve.prenom, niveau_scolaire) if moyenne_gen else None,
                    'matricule': eleve.matricule,
                    'total_eleves': total_eleves,
                    'niveau_scolaire': niveau_scolaire,
                    'has_notes_mensuelles': has_notes_mensuelles,  # Pour masquer colonnes mensuelles si seulement compositions
                }
            else:
                # Fallback: calculer si pas dans le cache
                calculateur = CalculateurBulletinIntelligent(eleve, classe_note, periode, systeme)
                bulletin_data = calculateur.generer_bulletin()
                
                rang_info = rangs_officiels.get(eleve.id)
                if rang_info:
                    moyenne_gen = float(rang_info['moyenne'])
                    bulletin_data['moyenne_generale'] = moyenne_gen
                    bulletin_data['rang'] = rang_info['rang']
                    total_coefficients = bulletin_data.get('total_coefficients')
                    if total_coefficients:
                        bulletin_data['total_points'] = round(moyenne_gen * float(total_coefficients), 2)
                else:
                    bulletin_data['rang'] = '-'
                
                bulletin_data['matricule'] = eleve.matricule
                bulletin_data['total_eleves'] = total_eleves
                bulletin_data['system_type'] = system_type
                bulletin_data['has_notes_mensuelles'] = has_notes_mensuelles
            
            # Dessiner le bulletin sur la page courante (passer le logo pré-chargé)
            _dessiner_bulletin_page(c, bulletin_data, logo_path, ecole, logo_reader)
            
            # Nouvelle page sauf pour le dernier élève
            if idx < len(eleves) - 1:
                c.showPage()
                
        except Exception as e:
            print(f"Erreur pour {eleve.nom} {eleve.prenom}: {str(e)}")
            # Ajouter une page vide avec message d'erreur
            c.setFont("Helvetica", 12)
            c.drawString(2*cm, A4[1]/2, f"Erreur pour {eleve.nom} {eleve.prenom}")
            if idx < len(eleves) - 1:
                c.showPage()
            continue
    
    c.save()
    buffer.seek(0)
    
    # Nettoyer le nom de fichier
    nom_classe_clean = re.sub(r'[^\w\s-]', '', classe_note.nom).replace(' ', '_')
    filename = f"bulletins_{nom_classe_clean}_{periode}.pdf"
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def _formater_periode_libelle(periode):
    """Convertit le code période en libellé lisible"""
    mapping = {
        'TRIMESTRE_1': '1ER TRIMESTRE',
        'TRIMESTRE_2': '2ÈME TRIMESTRE', 
        'TRIMESTRE_3': '3ÈME TRIMESTRE',
        'SEMESTRE_1': '1ER SEMESTRE',
        'SEMESTRE_2': '2ÈME SEMESTRE',
        'ANNUEL_TRIM': 'ANNUEL',
        'ANNUEL_SEM': 'ANNUEL',
    }
    return mapping.get(periode, periode.upper().replace('_', ' '))


def _dessiner_bulletin_page(c, bulletin_data, logo_path, ecole, logo_reader=None):
    """Dessine un bulletin sur la page courante du canvas - fonction interne optimisée"""
    width, height = A4
    
    # Couleurs du design (constantes)
    BLEU_HEADER = colors.HexColor('#4a90d9')
    BLEU_CLAIR = colors.HexColor('#e8f4fd')
    BLEU_APPRECIATION = colors.HexColor('#d6eaf8')
    VERT_MOY = colors.HexColor('#c8e6c9')
    ROUGE_PTS = colors.HexColor('#ffcdd2')
    ROUGE_DRAPEAU = colors.HexColor('#CE1126')
    JAUNE_DRAPEAU = colors.HexColor('#FCD116')
    VERT_DRAPEAU = colors.HexColor('#009460')
    GRIS_TOTAL = colors.HexColor('#37474f')
    
    MENTION_COLORS = {
        'EXCELLENT': colors.HexColor('#1b5e20'),
        'TRÈS BIEN': colors.HexColor('#2e7d32'),
        'BIEN': colors.HexColor('#0277bd'),
        'ASSEZ BIEN': colors.HexColor('#f9a825'),
        'PASSABLE': colors.HexColor('#ef6c00'),
        'INSUFFISANT': colors.HexColor('#c62828'),
        'FAIBLE': colors.HexColor('#b71c1c'),
    }
    
    # Utiliser le logo pré-chargé ou le charger si nécessaire
    img_reader = logo_reader
    if img_reader is None and logo_path:
        try:
            img_reader = ImageReader(logo_path)
        except:
            pass
    
    # ===== FILIGRANE =====
    if img_reader:
        try:
            c.saveState()
            c.setFillAlpha(0.08)
            filigrane_size = 14 * cm
            x = (width - filigrane_size) / 2
            y = (height - filigrane_size) / 2
            c.drawImage(img_reader, x, y, width=filigrane_size, height=filigrane_size, 
                       preserveAspectRatio=True, mask='auto')
            c.restoreState()
        except:
            pass
    
    # ===== EN-TÊTE =====
    if img_reader:
        try:
            c.drawImage(img_reader, 1.2*cm, height - 3.2*cm, width=2.5*cm, height=2.5*cm, 
                       preserveAspectRatio=True, mask='auto')
        except:
            pass
    
    # Drapeau guinéen
    drapeau_x = width - 2.5*cm
    drapeau_y = height - 2.2*cm
    drapeau_w = 1.2*cm
    drapeau_h = 0.8*cm
    c.setFillColor(ROUGE_DRAPEAU)
    c.rect(drapeau_x, drapeau_y, drapeau_w/3, drapeau_h, fill=1, stroke=0)
    c.setFillColor(JAUNE_DRAPEAU)
    c.rect(drapeau_x + drapeau_w/3, drapeau_y, drapeau_w/3, drapeau_h, fill=1, stroke=0)
    c.setFillColor(VERT_DRAPEAU)
    c.rect(drapeau_x + 2*drapeau_w/3, drapeau_y, drapeau_w/3, drapeau_h, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor('#cccccc'))
    c.rect(drapeau_x, drapeau_y, drapeau_w, drapeau_h, fill=0, stroke=1)
    
    # Textes en-tête
    y_header = height - 1.2*cm
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(width/2, y_header, "RÉPUBLIQUE DE GUINÉE")
    
    y_header -= 0.5*cm
    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(ROUGE_DRAPEAU)
    c.drawCentredString(width/2 - 1.8*cm, y_header, "Travail")
    c.setFillColor(JAUNE_DRAPEAU)
    c.drawCentredString(width/2, y_header, "Justice")
    c.setFillColor(VERT_DRAPEAU)
    c.drawCentredString(width/2 + 1.8*cm, y_header, "Solidarité")
    c.setFillColor(colors.black)
    c.drawCentredString(width/2 - 0.9*cm, y_header, "-")
    c.drawCentredString(width/2 + 0.9*cm, y_header, "-")
    
    y_header -= 0.45*cm
    c.setFont("Helvetica", 6)
    c.drawCentredString(width/2, y_header, "Ministère de l'Enseignement Pré-Universitaire et de l'Alphabétisation (MEPU-A)")
    
    y_header -= 0.5*cm
    c.setFont("Helvetica-Bold", 10)
    nom_ecole = ecole.nom.upper() if ecole else "GROUPE SCOLAIRE"
    c.drawCentredString(width/2, y_header, nom_ecole)
    
    y_header -= 0.6*cm
    c.setFont("Helvetica-Bold", 12)
    periode_libelle = _formater_periode_libelle(bulletin_data.get('periode', ''))
    c.drawCentredString(width/2, y_header, f"BULLETIN DE NOTES - {periode_libelle}")
    
    # Récupérer l'année scolaire depuis l'école ou utiliser l'année courante
    annee_scolaire = ecole.annee_scolaire if hasattr(ecole, 'annee_scolaire') and ecole.annee_scolaire else "2025-2026"
    c.setFont("Helvetica", 7)
    c.setFillColor(colors.HexColor('#666666'))
    c.drawRightString(width - 1.2*cm, y_header, f"Année Scolaire {annee_scolaire}")
    
    # Ligne de séparation
    y_header -= 0.4*cm
    c.setStrokeColor(colors.black)
    c.setLineWidth(1.5)
    c.line(1.2*cm, y_header, width - 1.2*cm, y_header)
    c.setLineWidth(1)
    
    # ===== INFORMATIONS ÉLÈVE =====
    y = y_header - 0.6*cm
    
    # Utiliser les champs prenom/nom séparés (fiables même pour un prénom
    # composé). Repli sur l'ancien découpage si absents.
    prenom = bulletin_data.get('prenom')
    nom = bulletin_data.get('nom')
    if prenom is None and nom is None:
        eleve_nom_complet = bulletin_data.get('eleve', '')
        parties = eleve_nom_complet.split(' ', 1) if eleve_nom_complet else ['', '']
        prenom = parties[0] if len(parties) > 0 else ''
        nom = parties[1] if len(parties) > 1 else ''
    prenom = prenom or ''
    nom = nom or ''
    
    # Formater la période pour l'affichage dans les infos élève
    periode_info = _formater_periode_libelle(bulletin_data.get('periode', '-'))
    
    info_data = [
        [('PRÉNOM', prenom.title()), ('NOM', nom.upper()), ('MATRICULE', bulletin_data.get('matricule', '-'))],
        [('CLASSE', bulletin_data.get('classe', '-')), ('PÉRIODE', periode_info), ('EFFECTIF', f"{bulletin_data.get('total_eleves', '-')} élèves")],
    ]
    
    info_total_width = width - 2.4*cm
    box_width = (info_total_width - 0.4*cm) / 3
    box_height = 0.9*cm
    
    for row_idx, row in enumerate(info_data):
        for col_idx, (label, value) in enumerate(row):
            x = 1.2*cm + col_idx * (box_width + 0.2*cm)
            box_y = y - row_idx * (box_height + 0.15*cm)
            
            c.setFillColor(BLEU_CLAIR)
            c.rect(x, box_y - box_height, box_width, box_height, fill=1, stroke=0)
            c.setStrokeColor(colors.HexColor('#b3d4fc'))
            c.rect(x, box_y - box_height, box_width, box_height, fill=0, stroke=1)
            
            c.setFillColor(colors.HexColor('#1976d2'))
            c.setFont("Helvetica-Bold", 7)
            c.drawString(x + 0.15*cm, box_y - 0.3*cm, label)
            
            c.setFillColor(colors.black)
            val_str = str(value)
            val_fs = 9
            max_val_w = box_width - 0.3*cm
            from reportlab.pdfbase.pdfmetrics import stringWidth as _sw
            while val_fs > 6 and _sw(val_str, "Helvetica-Bold", val_fs) > max_val_w:
                val_fs -= 0.5
            c.setFont("Helvetica-Bold", val_fs)
            c.drawString(x + 0.15*cm, box_y - 0.7*cm, val_str)

    # ===== TABLEAU DES NOTES =====
    y -= 2.3*cm
    
    table_total_width = width - 2.4*cm
    margin_left = 1.2*cm
    
    # Déterminer le type de système
    system_type = bulletin_data.get('system_type', 'mensuel')
    periode = bulletin_data.get('periode', '')
    
    # Détecter le niveau scolaire
    from notes.calculs_moyennes import detecter_niveau_scolaire
    classe_nom = bulletin_data.get('classe', '')
    niveau_scolaire = detecter_niveau_scolaire(classe_nom)
    est_primaire = (niveau_scolaire == 'PRIMAIRE')
    est_maternelle = (niveau_scolaire == 'MATERNELLE')
    
    # MATERNELLE/GARDERIE: Afficher les appréciations au lieu des notes
    if est_maternelle:
        _dessiner_bulletin_maternelle(c, bulletin_data, width, height, y, ecole)
        # Ne pas faire c.save() ici - c'est géré par la fonction appelante
        return
    
    # Déterminer les mois selon la période pour trimestre/semestre
    # SYSTÈME GUINÉEN: Le dernier mois de chaque période = Composition
    # Trimestres: T1(Oct,Nov,Déc), T2(Jan,Fév,Mars), T3(Avr,Mai,Juin)
    # Semestres: S1(Oct,Nov,Déc,Jan,Fév), S2(Mars,Avr,Mai,Juin)
    mois_labels = []
    periodes_labels = []  # Pour bulletin annuel
    
    if system_type == 'trimestriel':
        if 'TRIMESTRE_1' in periode:
            mois_labels = ['Oct.', 'Nov.']  # Déc. = Compo
        elif 'TRIMESTRE_2' in periode:
            mois_labels = ['Jan.', 'Fév.']  # Mars = Compo
        elif 'TRIMESTRE_3' in periode:
            mois_labels = ['Avr.', 'Mai']   # Juin = Compo
    elif system_type == 'semestriel':
        if 'SEMESTRE_1' in periode:
            mois_labels = ['Oct.', 'Nov.', 'Déc.', 'Jan.']  # Fév. = Compo
        elif 'SEMESTRE_2' in periode:
            mois_labels = ['Mars', 'Avr.', 'Mai']   # Juin = Compo
    elif system_type == 'annuel_trimestriel':
        # Bulletin annuel basé sur les trimestres
        periodes_labels = ['1er Trim.', '2ème Trim.', '3ème Trim.']
    elif system_type == 'annuel_semestriel':
        # Bulletin annuel basé sur les semestres
        periodes_labels = ['1er Sem.', '2ème Sem.']
    
    # Détecter si des notes mensuelles existent (pour masquer les colonnes si seulement compositions)
    has_notes_mensuelles = bulletin_data.get('has_notes_mensuelles', True)  # Par défaut True
    
    # Adapter les colonnes selon le système ET le niveau (primaire = sans COEF ni PTS)
    # Structure: MATIÈRE | [COEF] | Mois1 | Mois2 | [Mois...] | Moy.C | Compo | MOY | [PTS]
    # Si pas de notes mensuelles: MATIÈRE | [COEF] | Compo | MOY | [PTS]
    if system_type in ['annuel_trimestriel', 'annuel_semestriel'] and periodes_labels:
        # BULLETIN ANNUEL: MATIÈRE | [COEF] | T1/S1 | T2/S2 | [T3] | Moy. Ann. | [PTS]
        if est_primaire:
            header = ['MATIÈRE'] + periodes_labels + ['Moy. Ann.']
        else:
            header = ['MATIÈRE', 'COEF'] + periodes_labels + ['Moy. Ann.', 'PTS']
        data = [header]
        nb_cols = len(header)
    elif system_type == 'trimestriel':
        if has_notes_mensuelles and mois_labels:
            # Avec notes mensuelles: afficher toutes les colonnes
            if est_primaire:
                header = ['MATIÈRE'] + mois_labels + ['Moy.C', 'Compo', 'MOY']
            else:
                header = ['MATIÈRE', 'COEF'] + mois_labels + ['Moy.C', 'Compo', 'MOY', 'PTS']
        else:
            # Sans notes mensuelles: afficher seulement Compo et MOY
            if est_primaire:
                header = ['MATIÈRE', 'Compo', 'MOY']
            else:
                header = ['MATIÈRE', 'COEF', 'Compo', 'MOY', 'PTS']
        data = [header]
        nb_cols = len(header)
    elif system_type == 'semestriel':
        if has_notes_mensuelles and mois_labels:
            # Avec notes mensuelles: afficher toutes les colonnes
            if est_primaire:
                header = ['MATIÈRE'] + mois_labels + ['Moy.C', 'Compo', 'MOY']
            else:
                header = ['MATIÈRE', 'COEF'] + mois_labels + ['Moy.C', 'Compo', 'MOY', 'PTS']
        else:
            # Sans notes mensuelles: afficher seulement Compo et MOY
            if est_primaire:
                header = ['MATIÈRE', 'Compo', 'MOY']
            else:
                header = ['MATIÈRE', 'COEF', 'Compo', 'MOY', 'PTS']
        data = [header]
        nb_cols = len(header)
    else:
        if est_primaire:
            data = [['MATIÈRE', 'NOTE', 'MOY']]
            nb_cols = 3
        else:
            data = [['MATIÈRE', 'COEF', 'NOTE', 'MOY', 'PTS']]
            nb_cols = 5
    
    total_coef = 0
    total_moy = 0
    total_points = 0
    nb_matieres_avec_moy = 0
    
    for matiere in bulletin_data['matieres']:
        nom_matiere = str(matiere.get('matiere', '-'))
        if hasattr(matiere.get('matiere'), 'nom'):
            nom_matiere = matiere['matiere'].nom
        
        # PRIMAIRE: coefficient = 1 (pas de pondération)
        # SECONDAIRE (Collège/Lycée): coefficient réel de la matière
        coef_brut = float(matiere.get('coefficient', 1))
        coef = 1.0 if est_primaire else coef_brut
        
        moy_continue = matiere.get('moyenne_continue')
        note_compo = matiere.get('note_composition')
        moyenne = matiere.get('moyenne') or matiere.get('moyenne_calculee')
        points = matiere.get('points')
        
        # Recalculer les points avec le bon coefficient
        if points is None and moyenne:
            points = moyenne * coef
        elif est_primaire and moyenne:
            # Forcer le recalcul pour primaire (coef=1)
            points = moyenne * 1.0
        
        # Récupérer les moyennes mensuelles si disponibles
        moyennes_mensuelles = matiere.get('moyennes_mensuelles', [])
        
        total_coef += coef
        if moyenne:
            total_moy += moyenne
            nb_matieres_avec_moy += 1
        if points:
            total_points += points
        
        if system_type in ['annuel_trimestriel', 'annuel_semestriel'] and periodes_labels:
            # BULLETIN ANNUEL: Construire la ligne avec les moyennes par période
            if est_primaire or est_maternelle:
                row = [nom_matiere]  # Sans COEF pour primaire/maternelle
            else:
                row = [nom_matiere, f"{coef:.0f}"]
            
            # Ajouter les moyennes par période (T1, T2, T3 ou S1, S2)
            for i, periode_label in enumerate(periodes_labels):
                note_periode = '-'
                if moyennes_mensuelles and i < len(moyennes_mensuelles):
                    moy_per = moyennes_mensuelles[i]
                    if isinstance(moy_per, dict):
                        val = moy_per.get('moyenne')
                        if val is not None:
                            note_periode = f"{val:.2f}"
                    elif moy_per is not None:
                        note_periode = f"{moy_per:.2f}"
                row.append(note_periode)
            
            # Ajouter Moy. Annuelle (et PTS seulement pour collège/lycée)
            row.append(f"{moyenne:.2f}" if moyenne else '-')
            if not (est_primaire or est_maternelle):
                row.append(f"{points:.2f}" if points else '-')
            data.append(row)
        elif system_type in ['trimestriel', 'semestriel']:
            # Construire la ligne selon le mode de saisie
            if est_primaire or est_maternelle:
                row = [nom_matiere]  # Sans COEF pour primaire/maternelle
            else:
                row = [nom_matiere, f"{coef:.0f}"]
            
            if has_notes_mensuelles and mois_labels:
                # Ajouter les notes mensuelles
                for i, mois_label in enumerate(mois_labels):
                    note_mois = '-'
                    if moyennes_mensuelles and i < len(moyennes_mensuelles):
                        moy_mens = moyennes_mensuelles[i]
                        if isinstance(moy_mens, dict):
                            val = moy_mens.get('moyenne')
                            if val is not None:
                                note_mois = f"{val:.2f}"
                        elif moy_mens is not None:
                            note_mois = f"{moy_mens:.2f}"
                    row.append(note_mois)
                
                # Ajouter Moy.C, Compo, MOY (et PTS seulement pour collège/lycée)
                row.append(f"{moy_continue:.2f}" if moy_continue else '-')
            
            # Compo et MOY sont toujours affichés
            row.append(f"{note_compo:.2f}" if note_compo else '-')
            row.append(f"{moyenne:.2f}" if moyenne else '-')
            if not (est_primaire or est_maternelle):
                row.append(f"{points:.2f}" if points else '-')
            data.append(row)
        else:
            # Mensuel
            cours_str = f"{moy_continue:.2f}" if moy_continue else '-'
            moyenne_str = f"{moyenne:.2f}" if moyenne else '-'
            points_str = f"{points:.2f}" if points else '-'
            if est_primaire or est_maternelle:
                data.append([nom_matiere, cours_str, moyenne_str])
            else:
                data.append([nom_matiere, f"{coef:.0f}", cours_str, moyenne_str, points_str])
    
    # Ligne TOTAL
    if system_type in ['annuel_trimestriel', 'annuel_semestriel'] and periodes_labels:
        # BULLETIN ANNUEL: Ligne total
        if est_primaire or est_maternelle:
            total_row = ['TOTAL']
        else:
            total_row = ['TOTAL', f"{total_coef:.0f}"]
        total_row += ['-'] * len(periodes_labels)  # Colonnes périodes vides (T1, T2, T3 ou S1, S2)
        total_row.append('-')  # Colonne Moy. Ann. = tiret (pas de somme des moyennes)
        if not (est_primaire or est_maternelle):
            total_row.append(f"{total_points:.2f}")  # Colonne PTS = total des points
        data.append(total_row)
    elif system_type in ['trimestriel', 'semestriel']:
        if est_primaire or est_maternelle:
            total_row = ['TOTAL']
        else:
            total_row = ['TOTAL', f"{total_coef:.0f}"]
        
        if has_notes_mensuelles and mois_labels:
            total_row += ['-'] * len(mois_labels)  # Colonnes mois vides
            total_row.append('-')  # Moy.C
        
        total_row.append('-')  # Compo
        total_row.append(f"{total_moy:.0f}" if nb_matieres_avec_moy else '-')
        if not (est_primaire or est_maternelle):
            total_row.append(f"{total_points:.2f}")
        data.append(total_row)
    else:
        if est_primaire or est_maternelle:
            data.append(['TOTAL', '-', f"{total_moy:.0f}" if nb_matieres_avec_moy else '-'])
        else:
            data.append(['TOTAL', f"{total_coef:.0f}", '-', f"{total_moy:.0f}" if nb_matieres_avec_moy else '-', f"{total_points:.2f}"])
    
    # Calculer les largeurs de colonnes
    if system_type in ['annuel_trimestriel', 'annuel_semestriel'] and periodes_labels:
        # Bulletin annuel: MATIÈRE 25%, reste divisé équitablement
        col_matiere = table_total_width * 0.22
        nb_autres_cols = nb_cols - 1
        col_autres = (table_total_width - col_matiere) / nb_autres_cols
        col_widths = [col_matiere] + [col_autres] * nb_autres_cols
    elif system_type in ['trimestriel', 'semestriel']:
        # Répartition: MATIÈRE 25-35%, reste divisé équitablement
        if has_notes_mensuelles and mois_labels:
            col_matiere = table_total_width * 0.22
        else:
            col_matiere = table_total_width * 0.35  # Plus large si moins de colonnes
        nb_autres_cols = nb_cols - 1
        col_autres = (table_total_width - col_matiere) / nb_autres_cols
        col_widths = [col_matiere] + [col_autres] * nb_autres_cols
    else:
        if est_primaire or est_maternelle:
            # Primaire/Maternelle: 3 colonnes (MATIÈRE, NOTE, MOY)
            col_matiere = table_total_width * 0.50
            col_autres = (table_total_width - col_matiere) / 2
            col_widths = [col_matiere, col_autres, col_autres]
        else:
            col_matiere = table_total_width * 0.35
            col_autres = (table_total_width - col_matiere) / 4
            col_widths = [col_matiere, col_autres, col_autres, col_autres, col_autres]
    
    table = Table(data, colWidths=col_widths)
    
    # Styles de base
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), BLEU_HEADER),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 13),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 12),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('LEFTPADDING', (0, 1), (0, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')]),
        ('BACKGROUND', (0, -1), (-1, -1), GRIS_TOTAL),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
    ]
    
    # Colonnes MOY et PTS avec couleurs (position varie selon le système et le niveau)
    if system_type in ['trimestriel', 'semestriel'] and mois_labels:
        # Colonnes dynamiques: MOY = avant-dernière colonne, PTS = dernière colonne
        col_moy = nb_cols - 2
        col_pts = nb_cols - 1
        style.extend([
            ('BACKGROUND', (col_moy, 1), (col_moy, -2), VERT_MOY),
            ('FONTNAME', (col_moy, 1), (col_moy, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (col_pts, 1), (col_pts, -2), ROUGE_PTS),
            ('FONTNAME', (col_pts, 1), (col_pts, -1), 'Helvetica-Bold'),
            # Taille de police pour les colonnes nombreuses
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ])
    else:
        if est_primaire:
            # Primaire: 4 colonnes: MATIÈRE | NOTE | MOY | PTS (sans COEF)
            style.extend([
                ('BACKGROUND', (2, 1), (2, -2), VERT_MOY),    # MOY = colonne 2
                ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (3, 1), (3, -2), ROUGE_PTS),   # PTS = colonne 3
                ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Bold'),
            ])
        else:
            # 5 colonnes: MATIÈRE | COEF | NOTE | MOY | PTS
            style.extend([
                ('BACKGROUND', (3, 1), (3, -2), VERT_MOY),    # MOY = colonne 3
                ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (4, 1), (4, -2), ROUGE_PTS),   # PTS = colonne 4
                ('FONTNAME', (4, 1), (4, -1), 'Helvetica-Bold'),
            ])
    
    table.setStyle(TableStyle(style))
    table_w, table_h = table.wrap(width, height)
    table.drawOn(c, margin_left, y - table_h)
    
    # ===== SECTION RÉSULTATS =====
    y = y - table_h - 0.8*cm
    
    result_total_width = width - 2.4*cm
    col_width = (result_total_width - 0.4*cm) / 3
    col_height = 1.4*cm
    start_x = 1.2*cm
    
    # MOYENNE GÉNÉRALE
    c.setFillColor(colors.white)
    c.rect(start_x, y - col_height, col_width, col_height, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(start_x + col_width/2, y - 0.35*cm, "MOYENNE GÉNÉRALE")
    c.setFont("Helvetica-Bold", 16)
    moy_gen = bulletin_data.get('moyenne_generale')
    # Primaire = notation sur 10, Secondaire = notation sur 20
    base_notation = 10 if est_primaire else 20
    moy_str = f"{moy_gen:.2f}/{base_notation}" if moy_gen else '-'
    c.drawCentredString(start_x + col_width/2, y - 1*cm, moy_str)
    
    # RANG
    rang_x = start_x + col_width + 0.2*cm
    c.setFillColor(colors.white)
    c.rect(rang_x, y - col_height, col_width, col_height, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(rang_x + col_width/2, y - 0.35*cm, "RANG")
    c.setFont("Helvetica-Bold", 16)
    rang = bulletin_data.get('rang', '-')
    c.drawCentredString(rang_x + col_width/2, y - 1*cm, str(rang))
    
    # MENTION
    mention_x = start_x + 2*(col_width + 0.2*cm)
    c.setFillColor(colors.white)
    c.rect(mention_x, y - col_height, col_width, col_height, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(mention_x + col_width/2, y - 0.35*cm, "MENTION")
    
    mention = bulletin_data.get('mention', '-')
    mention_upper = mention.upper() if mention else '-'
    mention_color = MENTION_COLORS.get(mention_upper, colors.HexColor('#666666'))
    
    badge_width = min(4*cm, col_width - 0.4*cm)
    badge_height = 0.6*cm
    badge_x = mention_x + (col_width - badge_width) / 2
    badge_y = y - 1.15*cm
    
    c.setFillColor(mention_color)
    c.roundRect(badge_x, badge_y, badge_width, badge_height, 3, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(mention_x + col_width/2, badge_y + 0.15*cm, mention_upper)
    
    # ===== APPRÉCIATION =====
    y -= col_height + 0.6*cm
    
    appreciation_height = 1.2*cm
    c.setFillColor(BLEU_APPRECIATION)
    c.rect(1.2*cm, y - appreciation_height, width - 2.4*cm, appreciation_height, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor('#90caf9'))
    c.rect(1.2*cm, y - appreciation_height, width - 2.4*cm, appreciation_height, fill=0, stroke=1)
    
    c.setFillColor(colors.HexColor('#1565c0'))
    c.setFont("Helvetica-Bold", 9)
    c.drawString(1.4*cm, y - 0.35*cm, "APPRÉCIATION DU CONSEIL DE CLASSE")
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Oblique", 8)
    # Utiliser l'appréciation intelligente si non fournie
    appreciation = bulletin_data.get('appreciation')
    if not appreciation:
        moyenne_gen = bulletin_data.get('moyenne_generale')
        niveau = bulletin_data.get('niveau_scolaire', 'SECONDAIRE')
        eleve_nom = bulletin_data.get('eleve', '').split()[0] if bulletin_data.get('eleve') else 'L\'élève'
        appreciation = obtenir_appreciation_intelligente(moyenne_gen, eleve_nom, niveau) if moyenne_gen else 'Aucune note disponible pour cette période.'
    if appreciation and len(appreciation) > 100:
        appreciation = appreciation[:97] + '...'
    c.drawString(1.4*cm, y - 0.85*cm, appreciation)
    
    # ===== SIGNATURES =====
    y -= appreciation_height + 0.8*cm
    
    sig_width = 6*cm
    
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 9)
    titre_signataire = "Directeur du primaire" if est_primaire else "Censeur de l'établissement"
    c.drawCentredString(1.2*cm + sig_width/2, y, titre_signataire)
    c.setStrokeColor(colors.black)
    c.line(1.2*cm + 0.5*cm, y - 1.5*cm, 1.2*cm + sig_width - 0.5*cm, y - 1.5*cm)
    c.setFont("Helvetica", 8)
    c.drawCentredString(1.2*cm + sig_width/2, y - 1.8*cm, "Signature")
    
    dir_x = width - 1.2*cm - sig_width
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(dir_x + sig_width/2, y, "Directeur Général")
    c.line(dir_x + 0.5*cm, y - 1.5*cm, dir_x + sig_width - 0.5*cm, y - 1.5*cm)
    c.setFont("Helvetica", 8)
    c.drawCentredString(dir_x + sig_width/2, y - 1.8*cm, "Signature")
    
    # ===== MÉTHODE DE CALCUL =====
    y -= 2.5*cm
    
    c.setFillColor(colors.HexColor('#f5f5f5'))
    c.rect(1.2*cm, y - 1.8*cm, width - 2.4*cm, 1.8*cm, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor('#1976d2'))
    c.setLineWidth(2)
    c.line(1.2*cm, y, 1.2*cm, y - 1.8*cm)
    c.setLineWidth(1)
    
    c.setFillColor(colors.HexColor('#1976d2'))
    c.setFont("Helvetica-Bold", 8)
    
    # Texte pour les points (différent pour primaire)
    points_text = "• Points : Moyenne (pas de coefficient pour le primaire)" if est_primaire else "• Points : Moyenne × Coefficient"
    
    # Adapter le texte selon le système
    if system_type == 'annuel_trimestriel':
        titre_systeme = "MÉTHODE DE CALCUL - BULLETIN ANNUEL (TRIMESTRES)" if not est_primaire else "MÉTHODE DE CALCUL - BULLETIN ANNUEL (PRIMAIRE)"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• T1, T2, T3 : Moyennes des 1er, 2ème et 3ème trimestres")
        c.drawString(1.4*cm, y - 0.9*cm, "• Moy. Ann. : (T1 + T2 + T3) / 3 (moyenne des 3 trimestres)")
        c.drawString(1.4*cm, y - 1.2*cm, points_text)
    elif system_type == 'annuel_semestriel':
        titre_systeme = "MÉTHODE DE CALCUL - BULLETIN ANNUEL (SEMESTRES)" if not est_primaire else "MÉTHODE DE CALCUL - BULLETIN ANNUEL (PRIMAIRE)"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• S1, S2 : Moyennes des 1er et 2ème semestres")
        c.drawString(1.4*cm, y - 0.9*cm, "• Moy. Ann. : (S1 + S2) / 2 (moyenne des 2 semestres)")
        c.drawString(1.4*cm, y - 1.2*cm, points_text)
    elif system_type == 'trimestriel':
        titre_systeme = "MÉTHODE DE CALCUL DES NOTES - SYSTÈME TRIMESTRIEL (PRIMAIRE)" if est_primaire else "MÉTHODE DE CALCUL DES NOTES - SYSTÈME TRIMESTRIEL"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• Cours : Moyenne des notes d'évaluations continues du trimestre")
        c.drawString(1.4*cm, y - 0.9*cm, "• Compo : Note de composition du trimestre")
        c.drawString(1.4*cm, y - 1.2*cm, f"• {'Moyenne : composition' if est_primaire else 'Moyenne : 40% cours + 60% compo'} | {points_text}")
    elif system_type == 'semestriel':
        titre_systeme = "MÉTHODE DE CALCUL DES NOTES - SYSTÈME SEMESTRIEL (PRIMAIRE)" if est_primaire else "MÉTHODE DE CALCUL DES NOTES - SYSTÈME SEMESTRIEL"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• Cours : Moyenne des notes d'évaluations continues du semestre")
        c.drawString(1.4*cm, y - 0.9*cm, "• Compo : Note de composition (examen) du semestre")
        c.drawString(1.4*cm, y - 1.2*cm, f"• {'Moyenne : composition' if est_primaire else 'Moyenne : 40% cours + 60% compo'} | {points_text}")
    else:
        titre_systeme = "MÉTHODE DE CALCUL DES NOTES - SYSTÈME MENSUEL (PRIMAIRE)" if est_primaire else "MÉTHODE DE CALCUL DES NOTES - SYSTÈME MENSUEL"
        c.drawString(1.4*cm, y - 0.3*cm, titre_systeme)
        c.setFillColor(colors.HexColor('#333333'))
        c.setFont("Helvetica", 7)
        c.drawString(1.4*cm, y - 0.6*cm, "• Note : Moyenne des notes d'évaluations continues du mois")
        c.drawString(1.4*cm, y - 0.9*cm, "• Moyenne : Égale à la note du cours pour le système mensuel")
        c.drawString(1.4*cm, y - 1.2*cm, points_text)
    
    # Adapter les mentions selon le niveau scolaire
    if est_primaire:
        c.drawString(1.4*cm, y - 1.5*cm, "• Mentions : Excellent ≥9 | Très bien ≥8 | Bien ≥7 | Assez bien ≥6 | Passable ≥5 | Insuffisant ≥4 | Faible ≥3 | Très faible <3")
    else:
        c.drawString(1.4*cm, y - 1.5*cm, "• Mentions : Excellent ≥18 | Très bien ≥16 | Bien ≥14 | Assez bien ≥12 | Passable ≥10 | Insuffisant ≥8 | Faible ≥6 | Très faible <6")
    
    # ===== PIED DE PAGE avec infos dynamiques de l'école =====
    c.setFillColor(colors.HexColor('#999999'))
    c.setFont("Helvetica", 6)
    # Construire le footer avec les infos de l'école
    footer_parts = [f"© 2025 {ecole.nom if ecole else 'Myschool'}. Tous droits réservés."]
    if ecole and ecole.telephone:
        footer_parts.append(f"Tél: {ecole.tous_telephones}")
    if ecole and ecole.email:
        footer_parts.append(ecole.email)
    c.drawCentredString(width/2, 0.8*cm, " | ".join(footer_parts))
    c.drawCentredString(width/2, 0.5*cm, f"Bulletin généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
