"""
Vues pour l'importation d'élèves depuis Excel/CSV
"""
import os
import tempfile
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction

from eleves.models import Classe, Eleve


@login_required
@require_http_methods(["GET", "POST"])
def importer_eleves(request):
    """
    Vue principale pour importer des élèves
    """
    # Vérifier les permissions
    peut_importer = (
        request.user.is_staff or 
        request.user.is_superuser or
        request.user.groups.filter(name__in=['Administrateurs', 'Directeurs', 'Comptables']).exists() or
        (hasattr(request.user, 'profil') and request.user.profil.peut_importer_eleves) or
        (hasattr(request.user, 'profil') and request.user.profil.role == 'COMPTABLE')
    )
    
    if not peut_importer:
        messages.error(request, "Vous n'avez pas la permission d'importer des élèves.")
        return redirect('eleves:liste_eleves')
    
    # Déterminer l'année scolaire à utiliser : on prend la plus récente existante
    annee_courante = Classe.objects.order_by('-annee_scolaire').values_list('annee_scolaire', flat=True).first()

    # Récupérer les classes pour cette année scolaire
    classes = Classe.objects.all()
    if annee_courante:
        classes = classes.filter(annee_scolaire=annee_courante)
    classes = classes.order_by('nom')
    
    # Filtrage par école si nécessaire
    if not request.user.is_superuser:
        from utilisateurs.utils import user_school
        ecole = user_school(request.user)
        if ecole:
            classes = classes.filter(ecole=ecole)
    
    context = {
        'classes': classes,
        'annee_courante': annee_courante,
    }
    
    if request.method == 'POST':
        return _traiter_import_eleves(request)
    
    return render(request, 'eleves/importer_eleves.html', context)


def _traiter_import_eleves(request):
    """
    Traite l'importation d'élèves
    """
    from eleves.import_eleves import (
        ImportElevesError,
        ImportElevesValidator,
        ImportElevesProcessor,
        lire_fichier_eleves,
    )

    try:
        # Récupérer les paramètres
        classe_id = request.POST.get('classe_id')
        generer_matricules = request.POST.get('generer_matricules') == 'on'
        fichier = request.FILES.get('fichier')
        
        if not fichier:
            messages.error(request, "Veuillez sélectionner un fichier.")
            return redirect('eleves:importer_eleves')
        
        # Sauvegarder temporairement le fichier
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(fichier.name)[1]) as tmp_file:
            for chunk in fichier.chunks():
                tmp_file.write(chunk)
            tmp_path = tmp_file.name
        
        try:
            # Lire le fichier
            df = lire_fichier_eleves(tmp_path)
            
            groupes = []
            if classe_id:
                classe = get_object_or_404(Classe, pk=classe_id)
                if not request.user.is_superuser:
                    from utilisateurs.utils import user_school
                    ecole_user = user_school(request.user)
                    if not ecole_user or classe.ecole_id != ecole_user.id:
                        raise ImportElevesError("Classe de destination non autorisée.")
                groupes.append((classe, df))
            elif 'Classe' in df.columns:
                # Un export complet peut être réimporté sans choisir une
                # classe : chaque ligne est dirigée vers sa classe d'origine.
                group_columns = ['Classe']
                if 'Année scolaire' in df.columns:
                    group_columns.append('Année scolaire')
                if request.user.is_superuser and 'École' in df.columns:
                    group_columns.append('École')

                for keys, groupe_df in df.groupby(group_columns, dropna=False, sort=False):
                    if not isinstance(keys, tuple):
                        keys = (keys,)
                    criteres = dict(zip(group_columns, keys))
                    classe_nom = str(criteres['Classe']).strip()
                    classes_qs = Classe.objects.filter(nom__iexact=classe_nom)
                    annee = criteres.get('Année scolaire')
                    if annee is not None and str(annee).strip() not in ('', 'nan'):
                        classes_qs = classes_qs.filter(annee_scolaire=str(annee).strip())
                    if request.user.is_superuser:
                        ecole_nom = criteres.get('École')
                        if ecole_nom is not None and str(ecole_nom).strip() not in ('', 'nan'):
                            classes_qs = classes_qs.filter(ecole__nom__iexact=str(ecole_nom).strip())
                    else:
                        from utilisateurs.utils import user_school
                        ecole_user = user_school(request.user)
                        if not ecole_user:
                            raise ImportElevesError("Aucune école n'est associée à votre compte.")
                        classes_qs = classes_qs.filter(ecole=ecole_user)

                    if classes_qs.count() != 1:
                        raise ImportElevesError(
                            f"Classe introuvable ou ambiguë dans le poste de destination : "
                            f"{classe_nom} ({annee or 'année non indiquée'})."
                        )
                    groupes.append((classes_qs.first(), groupe_df.copy()))
            else:
                raise ImportElevesError(
                    "Sélectionnez une classe, ou utilisez un export complet contenant la colonne 'Classe'."
                )

            # Tout valider avant la première écriture.
            validations = []
            for classe, groupe_df in groupes:
                validator = ImportElevesValidator(groupe_df, classe.id)
                validations.append((classe, groupe_df, validator))
                if not validator.valider():
                    for erreur in validator.erreurs[:5]:
                        messages.error(request, f"{classe.nom} : {erreur}")
                    if len(validator.erreurs) > 5:
                        messages.error(request, f"{classe.nom} : ... et {len(validator.erreurs) - 5} autres erreurs")
                    return redirect('eleves:importer_eleves')

            stats = {'total': 0, 'crees': 0, 'modifies': 0, 'erreurs': 0, 'matricules_generes': 0}
            for classe, groupe_df, validator in validations:
                for avertissement in validator.avertissements[:3]:
                    messages.warning(request, f"{classe.nom} : {avertissement}")
                resultat = ImportElevesProcessor(
                    df=groupe_df,
                    classe_id=classe.id,
                    user=request.user,
                    generer_matricules=generer_matricules,
                ).importer()
                for key in stats:
                    stats[key] += resultat[key]

            messages.success(request, f"✅ Importation terminée pour {len(groupes)} classe(s).")
            
            if stats['crees'] > 0:
                messages.success(
                    request,
                    f"📝 {stats['crees']} élève(s) créé(s)"
                )
            
            if stats['modifies'] > 0:
                messages.info(
                    request,
                    f"✏️ {stats['modifies']} élève(s) mis à jour"
                )
            
            if stats['matricules_generes'] > 0:
                messages.info(
                    request,
                    f"🔢 {stats['matricules_generes']} matricule(s) généré(s) automatiquement"
                )
            
            if stats['erreurs'] > 0:
                messages.warning(
                    request,
                    f"⚠️ {stats['erreurs']} erreur(s) rencontrée(s)"
                )
            
            messages.info(
                request,
                f"📊 Total traité: {stats['total']} élève(s)"
            )
            
            return redirect('eleves:gestion_classes')
            
        finally:
            # Nettoyer le fichier temporaire
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    except ImportElevesError as e:
        messages.error(request, f"Erreur d'importation: {e}")
    except Exception as e:
        messages.error(request, f"Erreur inattendue: {e}")
        import traceback
        print(traceback.format_exc())
    
    return redirect('eleves:importer_eleves')


@login_required
def telecharger_template_eleves(request):
    import pandas as pd
    from eleves.import_eleves import generer_template_eleves

    """
    Télécharge un template Excel pour l'importation d'élèves
    """
    try:
        classe_id = request.GET.get('classe_id')
        
        # Générer le template
        df = generer_template_eleves(classe_id)
        
        # Créer la réponse Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nom du fichier
        if classe_id:
            classe = Classe.objects.get(id=classe_id)
            filename = f"template_eleves_{classe.nom.replace(' ', '_')}.xlsx"
        else:
            filename = f"template_eleves_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Écrire le DataFrame dans la réponse
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Élèves', index=False)
            
            # Obtenir la feuille pour la formater
            worksheet = writer.sheets['Élèves']
            
            # Ajuster la largeur des colonnes
            for idx, col in enumerate(df.columns, 1):
                column_letter = chr(64 + idx) if idx <= 26 else 'A' + chr(64 + idx - 26)
                if col in ['Prénom', 'Nom', 'Lieu de Naissance', 'Adresse']:
                    worksheet.column_dimensions[column_letter].width = 20
                elif col == 'Matricule':
                    worksheet.column_dimensions[column_letter].width = 15
                elif col in ['Date de Naissance', 'Téléphone Principal', 'Téléphone Secondaire']:
                    worksheet.column_dimensions[column_letter].width = 18
                else:
                    worksheet.column_dimensions[column_letter].width = 15
            
            # Ajouter un style à l'en-tête
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Ajouter des instructions dans une nouvelle feuille
            instructions_sheet = writer.book.create_sheet('Instructions')
            instructions = [
                ["INSTRUCTIONS POUR L'IMPORTATION DES ÉLÈVES"],
                [""],
                ["1. COLONNES OBLIGATOIRES:"],
                ["   - Prénom: Le prénom de l'élève"],
                ["   - Nom: Le nom de famille de l'élève"],
                ["   - Sexe: M (Masculin) ou F (Féminin)"],
                ["   - Date de Naissance: Format JJ/MM/AAAA (ex: 15/01/2010)"],
                ["   - Lieu de Naissance: Ville ou commune de naissance"],
                ["   - Nom du Père/Tuteur: Nom du responsable principal"],
                ["   - Prénom du Père/Tuteur: Prénom du responsable principal"],
                ["   - Téléphone Principal: Numéro de téléphone (8 chiffres minimum)"],
                ["   - Adresse: Adresse complète de la famille"],
                [""],
                ["2. COLONNES OPTIONNELLES:"],
                ["   - Matricule: Si vide, sera généré automatiquement"],
                ["   - Nom de la Mère: Nom du responsable secondaire"],
                ["   - Prénom de la Mère: Prénom du responsable secondaire"],
                ["   - Téléphone Secondaire: Numéro secondaire"],
                ["   - Email: Adresse email de contact"],
                [""],
                ["3. RÈGLES IMPORTANTES:"],
                ["   - Ne pas modifier les noms des colonnes"],
                ["   - Respecter le format de date JJ/MM/AAAA"],
                ["   - Le sexe doit être uniquement M ou F"],
                ["   - Les téléphones doivent contenir au moins 8 chiffres"],
                ["   - Supprimer les lignes d'exemple avant l'import"],
                [""],
                ["4. GÉNÉRATION AUTOMATIQUE DES MATRICULES:"],
                ["   - Format: [CODE_CLASSE]-[ANNÉE]-[NUMÉRO]"],
                ["   - Exemple: 6A-2024-001"],
                ["   - Cochez l'option lors de l'importation"],
                [""],
                ["5. EN CAS D'ERREUR:"],
                ["   - Vérifier que toutes les colonnes obligatoires sont remplies"],
                ["   - Vérifier le format des dates"],
                ["   - Vérifier que les téléphones sont valides"],
                ["   - S'assurer qu'il n'y a pas de doublons"]
            ]
            
            for row_idx, instruction in enumerate(instructions, 1):
                for col_idx, text in enumerate(instruction, 1):
                    instructions_sheet.cell(row=row_idx, column=col_idx, value=text)
            
            # Formater la feuille d'instructions
            instructions_sheet.column_dimensions['A'].width = 80
            title_cell = instructions_sheet['A1']
            title_cell.font = Font(bold=True, size=14, color="366092")
            
            for row in [3, 14, 21, 27, 32]:
                instructions_sheet.cell(row=row, column=1).font = Font(bold=True, color="366092")
        
        return response
    
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du template: {e}")
        return redirect('eleves:importer_eleves')


@login_required
def exporter_eleves_classe(request, classe_id):
    import pandas as pd
    from eleves.import_eleves import exporter_liste_eleves

    """
    Exporte la liste des élèves d'une classe
    """
    try:
        classe = get_object_or_404(Classe, id=classe_id)
        
        # Vérifier les permissions
        if not request.user.is_superuser:
            from utilisateurs.utils import user_school
            ecole = user_school(request.user)
            if ecole and classe.ecole != ecole:
                messages.error(request, "Vous n'avez pas accès à cette classe.")
                return redirect('eleves:liste_eleves')
        
        # Exporter les élèves
        df = exporter_liste_eleves(classe_id)
        
        # Créer la réponse Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"eleves_{classe.nom.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Écrire le DataFrame
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=f'Élèves {classe.nom}', index=False)
            
            # Formater
            worksheet = writer.sheets[f'Élèves {classe.nom}']
            
            # Largeur des colonnes
            for idx, col in enumerate(df.columns, 1):
                column_letter = chr(64 + idx) if idx <= 26 else 'A' + chr(64 + idx - 26)
                worksheet.column_dimensions[column_letter].width = 18
            
            # Style de l'en-tête
            from openpyxl.styles import Font, PatternFill, Alignment
            
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        return response
    
    except Exception as e:
        messages.error(request, f"Erreur lors de l'export: {e}")
        return redirect('eleves:gestion_classes')


@login_required
def exporter_tous_eleves_modele(request):
    """Exporte tous les élèves visibles dans un fichier directement réimportable."""
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
    from eleves.import_eleves import exporter_liste_eleves_queryset

    eleves = Eleve.objects.filter(est_dans_corbeille=False).select_related(
        'classe', 'classe__ecole', 'responsable_principal', 'responsable_secondaire',
    )
    if not request.user.is_superuser:
        from utilisateurs.utils import user_school
        ecole = user_school(request.user)
        if not ecole:
            messages.error(request, "Aucune école n'est associée à votre compte.")
            return redirect('eleves:liste_eleves')
        eleves = eleves.filter(classe__ecole=ecole)
    eleves = eleves.order_by('classe__ecole__nom', 'classe__nom', 'nom', 'prenom')

    df = exporter_liste_eleves_queryset(eleves, inclure_classe=True)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"eleves_complets_reimportables_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Élèves', index=False)
        ws = writer.sheets['Élèves']
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1F4E78')
            cell.alignment = Alignment(horizontal='center')
        for column_cells in ws.columns:
            width = min(35, max(12, max(len(str(cell.value or '')) for cell in column_cells) + 2))
            ws.column_dimensions[column_cells[0].column_letter].width = width

        instructions = writer.book.create_sheet('Instructions')
        lignes = [
            "EXPORT COMPLET RÉIMPORTABLE",
            "",
            "Ce fichier reprend exactement les colonnes du modèle d'importation.",
            "Sur l'autre poste : Élèves > Importer, laissez la classe de destination vide, puis choisissez ce fichier.",
            "Les colonnes École, Classe et Année scolaire permettent de répartir automatiquement les élèves.",
            "Les classes correspondantes doivent exister sur le poste de destination.",
            "Ne modifiez pas les noms des colonnes.",
        ]
        for index, value in enumerate(lignes, 1):
            instructions.cell(index, 1, value)
        instructions['A1'].font = Font(bold=True, size=14, color='1F4E78')
        instructions.column_dimensions['A'].width = 110

    return response
