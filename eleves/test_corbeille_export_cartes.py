from datetime import date, timedelta
from io import BytesIO

from django.conf import settings
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook
from pypdf import PdfReader

from bus.models import AbonnementBus, AbonnementCantine
from paiements.models import ModePaiement, Paiement, TypePaiement

from .models import Classe, Ecole, Eleve


TEST_MIDDLEWARE = tuple(
    item for item in settings.MIDDLEWARE
    if item != 'ecole_moderne.licence_middleware.LicenceMiddleware'
)


@override_settings(MIDDLEWARE=TEST_MIDDLEWARE)
class CorbeilleExportCartesTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser('admin-fonctions', 'admin@test.local', 'secret')
        self.client.force_login(self.user)
        self.ecole = Ecole.objects.create(
            nom='École Fonctions', adresse='Conakry', telephone='+224622000101',
            directeur='Direction', etat='VALIDE',
        )
        self.classe = Classe.objects.create(
            ecole=self.ecole, nom='1ère primaire A', niveau='PRIMAIRE_1',
            code_matricule='T8', annee_scolaire='2026-2027',
        )
        self.eleve = Eleve.objects.create(
            matricule='T8-001', prenom='Aminata', nom='Diallo', sexe='F',
            classe=self.classe, statut='ACTIF',
        )

    def test_suppression_place_dans_corbeille_et_restaure_sans_perte(self):
        type_paiement = TypePaiement.objects.create(nom='Test corbeille')
        mode = ModePaiement.objects.create(nom='Espèces test')
        paiement = Paiement.objects.create(
            eleve=self.eleve, type_paiement=type_paiement, mode_paiement=mode,
            montant=100000, date_paiement=date.today(), statut='VALIDE',
        )

        response = self.client.post(reverse('eleves:supprimer_eleve', args=[self.eleve.id]))
        self.assertRedirects(response, reverse('eleves:liste_eleves'))
        self.eleve.refresh_from_db()
        self.assertTrue(self.eleve.est_dans_corbeille)
        self.assertTrue(Paiement.objects.filter(pk=paiement.pk).exists())
        self.assertNotContains(self.client.get(reverse('eleves:liste_eleves')), 'T8-001')
        self.assertContains(self.client.get(reverse('eleves:corbeille_eleves')), 'T8-001')

        response = self.client.post(reverse('eleves:restaurer_eleve', args=[self.eleve.id]))
        self.assertRedirects(response, reverse('eleves:corbeille_eleves'))
        self.eleve.refresh_from_db()
        self.assertFalse(self.eleve.est_dans_corbeille)
        self.assertEqual(self.eleve.statut, 'ACTIF')
        self.assertTrue(Paiement.objects.filter(pk=paiement.pk).exists())

    def test_export_complet_est_reimportable_sans_classe_forcee(self):
        response = self.client.get(reverse('eleves:exporter_tous_eleves_modele'))
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook['Élèves']
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(headers[:3], ['École', 'Classe', 'Année scolaire'])
        self.assertIn('Matricule', headers)
        self.assertIn('Prénom', headers)
        self.assertEqual(sheet.cell(2, headers.index('Classe') + 1).value, self.classe.nom)

        upload = SimpleUploadedFile(
            'eleves_reimportables.xlsx', response.content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        import_response = self.client.post(reverse('eleves:importer_eleves'), {
            'classe_id': '', 'generer_matricules': 'on', 'fichier': upload,
        })
        self.assertRedirects(import_response, reverse('eleves:gestion_classes'))
        self.assertEqual(Eleve.objects.filter(matricule='T8-001').count(), 1)

    def test_suppression_ecole_ne_cree_plus_de_contrainte_sync(self):
        ecole = Ecole.objects.create(
            nom='École à supprimer', adresse='Conakry', telephone='+224622000102',
            directeur='Direction', etat='BROUILLON',
        )
        Classe.objects.create(
            ecole=ecole, nom='Classe temporaire', niveau='COLLEGE_7',
            annee_scolaire='2026-2027',
        )
        ecole_id = ecole.id
        ecole.delete()
        self.assertFalse(Ecole.objects.filter(pk=ecole_id).exists())

    def test_cartes_retrait_bus_cantine_sont_huit_par_page(self):
        expiration = date.today() + timedelta(days=30)
        for index in range(9):
            eleve = self.eleve if index == 0 else Eleve.objects.create(
                matricule=f'T8-{index + 1:03d}', prenom=f'Élève{index}',
                nom='Test', sexe='M', classe=self.classe, statut='ACTIF',
            )
            AbonnementBus.objects.create(
                eleve=eleve, montant=10000, date_expiration=expiration,
                statut='ACTIF', zone='Zone A', point_arret='Arrêt A',
            )
            AbonnementCantine.objects.create(
                eleve=eleve, montant=10000, date_expiration=expiration,
                statut='ACTIF', type_repas='DEJEUNER', periodicite='MENSUEL',
            )

        urls = (
            reverse('eleves:tickets_retrait_classe_pdf', args=[self.classe.id]),
            reverse('eleves:tickets_bus_classe_pdf', args=[self.classe.id]),
            reverse('eleves:tickets_cantine_classe_pdf', args=[self.classe.id]),
            reverse('eleves:cartes_scolaires_classe_pdf', args=[self.classe.id]),
        )
        for url in urls:
            with self.subTest(url=url):
                response = self.client.get(url)
                self.assertEqual(response.status_code, 200)
                self.assertEqual(len(PdfReader(BytesIO(response.content)).pages), 2)
