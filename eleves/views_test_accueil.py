"""Pointage et exports du test d'accueil des élèves."""

import io
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas

from ecole_moderne.pdf_utils import draw_logo_watermark
from utilisateurs.utils import filter_by_user_school

from .models import Eleve


STATUTS = {
    'evalues': (True, 'Élèves évalués au test d’accueil'),
    'non-evalues': (False, 'Élèves non évalués au test d’accueil'),
}


def _statut(statut):
    try:
        return STATUTS[statut]
    except KeyError as exc:
        raise Http404("Statut de test d'accueil inconnu") from exc


def _queryset(request, statut):
    value, _label = _statut(statut)
    queryset = (
        Eleve.objects
        .select_related('classe', 'classe__ecole', 'responsable_principal')
        .filter(est_dans_corbeille=False, test_accueil_evalue=value)
    )
    queryset = filter_by_user_school(
        queryset, request.user, 'classe__ecole'
    )
    return queryset.order_by('-date_creation', '-id')


@login_required
@require_POST
def pointer_test_accueil(request, eleve_id):
    queryset = filter_by_user_school(
        Eleve.objects.filter(est_dans_corbeille=False),
        request.user,
        'classe__ecole',
    )
    eleve = get_object_or_404(queryset, pk=eleve_id)
    eleve.test_accueil_evalue = not eleve.test_accueil_evalue
    eleve.save(update_fields=['test_accueil_evalue', 'date_modification'])
    etat = 'évalué' if eleve.test_accueil_evalue else 'non évalué'
    messages.success(request, f"{eleve.nom_complet} est maintenant marqué {etat}.")
    retour = request.POST.get('next') or reverse('eleves:liste_eleves')
    if not url_has_allowed_host_and_scheme(
        retour, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        retour = reverse('eleves:liste_eleves')
    else:
        # Une liste chargée par AJAX contient ``partial=1`` dans son URL. Ne pas
        # rediriger le navigateur vers ce fragment seul après le pointage.
        parts = urlsplit(retour)
        query = urlencode([
            (key, value)
            for key, value in parse_qsl(parts.query, keep_blank_values=True)
            if key != 'partial'
        ])
        retour = urlunsplit((parts.scheme, parts.netloc, parts.path, query, parts.fragment))
    return redirect(retour)


@login_required
def export_test_accueil_excel(request, statut):
    _value, label = _statut(statut)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Test accueil'
    headers = [
        'Matricule', 'Prénom', 'Nom', 'Sexe', 'Date de naissance',
        'Classe', 'École', "Date d'inscription", 'Statut test accueil',
        'Responsable', 'Téléphone',
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='4472C4')
    for eleve in _queryset(request, statut):
        responsable = eleve.responsable_principal
        sheet.append([
            eleve.matricule or '', eleve.prenom, eleve.nom,
            eleve.get_sexe_display(),
            eleve.date_naissance.strftime('%d/%m/%Y') if eleve.date_naissance else '',
            eleve.classe.nom, eleve.classe.ecole.nom,
            eleve.date_inscription.strftime('%d/%m/%Y') if eleve.date_inscription else '',
            'Évalué' if eleve.test_accueil_evalue else 'Non évalué',
            responsable.nom_complet if responsable else '',
            responsable.telephone if responsable else '',
        ])
    widths = [18, 20, 20, 12, 18, 20, 30, 18, 20, 28, 18]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    stream = io.BytesIO()
    workbook.save(stream)
    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="test_accueil_{statut}.xlsx"'
    return response


@login_required
def export_test_accueil_pdf(request, statut):
    _value, label = _statut(statut)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="test_accueil_{statut}.pdf"'
    page_width, page_height = landscape(A4)
    pdf = canvas.Canvas(response, pagesize=(page_width, page_height))

    def header():
        draw_logo_watermark(pdf, page_width, page_height, opacity=0.035, scale=1.3)
        pdf.setFont('Helvetica-Bold', 16)
        pdf.drawCentredString(page_width / 2, page_height - 34, label.upper())
        pdf.setFont('Helvetica-Bold', 8)
        columns = [
            (28, 'N°'), (52, 'Matricule'), (142, 'Prénom et nom'),
            (295, 'Sexe'), (345, 'Classe'), (450, 'École'),
            (650, 'Inscription'), (725, 'Résultat'),
        ]
        for x, title in columns:
            pdf.drawString(x, page_height - 58, title)
        pdf.setStrokeColor(colors.HexColor('#4472C4'))
        pdf.line(28, page_height - 63, page_width - 25, page_height - 63)
        return page_height - 78

    y = header()
    pdf.setFont('Helvetica', 8)
    for index, eleve in enumerate(_queryset(request, statut), 1):
        if y < 35:
            pdf.showPage()
            y = header()
            pdf.setFont('Helvetica', 8)
        values = [
            (28, str(index)), (52, (eleve.matricule or '')[:16]),
            (142, eleve.nom_complet[:30]), (295, eleve.get_sexe_display()),
            (345, eleve.classe.nom[:18]), (450, eleve.classe.ecole.nom[:36]),
            (650, eleve.date_inscription.strftime('%d/%m/%Y') if eleve.date_inscription else '-'),
            (725, 'Évalué' if eleve.test_accueil_evalue else 'Non évalué'),
        ]
        for x, value in values:
            pdf.drawString(x, y, value)
        y -= 17
    pdf.save()
    return response
