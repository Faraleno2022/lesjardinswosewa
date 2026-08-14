"""Vues des modules de recouvrement.

Entrées, Cuisine, Documents et Versements partagent la même structure : une
liste avec saisie intégrée, un tableau de bord et deux exports. Ils sont donc
pilotés par une configuration commune (`MODULES`) plutôt que par autant de jeux
de vues identiques. L'informatique, plus riche, a ses propres vues.
"""
import io
from datetime import date, timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncMonth
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from eleves.models import Eleve
from utilisateurs.utils import filter_by_user_school, user_is_superadmin, user_school

from .forms_recouvrement import (
    AbonnementInformatiqueForm, DepenseCuisineForm, DepenseDocumentForm, EntreeForm,
    VersementForm,
)
from .models import Depense
from .models_recouvrement import (
    AbonnementInformatique, DepenseCuisine, DepenseDocument, Entree, Versement,
)
from .views_personnel import GROUPES as GROUPES_PERSONNEL, membres_du_personnel, totaux_personnel

# --------------------------------------------------------------------------
# Configuration des modules simples
#
# `colonnes_extra` : colonnes propres au module, ajoutées après l'intitulé dans
# les listes et les exports. Les autres modules n'en déclarent aucune.
# --------------------------------------------------------------------------

MODULES = {
    'entree': {
        'modele': Entree,
        'formulaire': EntreeForm,
        'titre': "Entrées de montants",
        'singulier': 'entrée',
        'icone': 'fas fa-arrow-down-to-line',
        'couleur': 'success',
        'champ_libelle': 'source',
        'entete_libelle': 'Provenance',
        'sens': 'entree',
        'colonnes_extra': [
            ('type_entree_libelle', "Type d'entrée"),
            ('mode_paiement_libelle', "Mode d'encaissement"),
            ('reference', 'Référence'),
        ],
        'champs_recherche': ['source', 'reference'],
        'filtre_choix': {
            'champ': 'type_entree',
            'label': "Type d'entrée",
            'choix': Entree.TYPE_CHOICES,
        },
    },
    'cuisine': {
        'modele': DepenseCuisine,
        'formulaire': DepenseCuisineForm,
        'titre': 'Dépenses de la cuisine',
        'singulier': 'dépense de cuisine',
        'icone': 'fas fa-utensils',
        'couleur': 'warning',
        'champ_libelle': 'designation',
        'entete_libelle': 'Désignation',
        'sens': 'sortie',
    },
    'document': {
        'modele': DepenseDocument,
        'formulaire': DepenseDocumentForm,
        'titre': 'Dépenses de documents',
        'singulier': 'dépense de document',
        'icone': 'fas fa-file-invoice',
        'couleur': 'info',
        'champ_libelle': 'designation',
        'entete_libelle': 'Désignation',
        'sens': 'sortie',
    },
    'versement': {
        'modele': Versement,
        'formulaire': VersementForm,
        'titre': 'Versements',
        'singulier': 'versement',
        'icone': 'fas fa-hand-holding-dollar',
        'couleur': 'success',
        'champ_libelle': 'lieu_versement',
        'entete_libelle': 'Lieu de versement',
        'sens': 'entree',
    },
}


def _config(cle):
    try:
        return MODULES[cle]
    except KeyError:
        raise Http404(f"Module de recouvrement inconnu : {cle}")


def _queryset(cle, user):
    """Opérations du module, restreintes à l'école de l'utilisateur."""
    modele = _config(cle)['modele']
    return filter_by_user_school(modele.objects.all(), user, 'ecole')


def _filtrer_periode(qs, request, cle):
    """Applique les filtres de recherche, de période et de type aux modules."""
    cfg = _config(cle)
    recherche = (request.GET.get('q') or '').strip()
    debut = (request.GET.get('date_debut') or '').strip()
    fin = (request.GET.get('date_fin') or '').strip()
    type_choisi = (request.GET.get('type') or '').strip()

    if recherche:
        champs = cfg.get('champs_recherche') or [cfg['champ_libelle']]
        condition = Q(observation__icontains=recherche)
        for champ in champs:
            condition |= Q(**{f'{champ}__icontains': recherche})
        qs = qs.filter(condition)
    if debut:
        qs = qs.filter(date__gte=debut)
    if fin:
        qs = qs.filter(date__lte=fin)

    filtre_choix = cfg.get('filtre_choix')
    if filtre_choix and type_choisi:
        qs = qs.filter(**{filtre_choix['champ']: type_choisi})

    return qs, {'q': recherche, 'date_debut': debut, 'date_fin': fin, 'type': type_choisi}


def _totaux(qs):
    agg = qs.aggregate(total=Sum('montant'), nombre=Count('id'))
    return int(agg['total'] or 0), agg['nombre'] or 0


# --------------------------------------------------------------------------
# Portail : tableau de bord général + cartes des modules
# --------------------------------------------------------------------------

@login_required
def hub_recouvrement(request):
    """Page d'entrée du menu Recouvrement.

    Présente un tableau de bord consolidé puis une carte par module, chacune
    affichant ses propres chiffres et menant à son tableau de bord détaillé.
    """
    aujourdhui = timezone.localdate()
    debut_mois = aujourdhui.replace(day=1)

    cartes = []
    total_sorties = total_entrees = 0
    total_sorties_mois = total_entrees_mois = 0

    for cle, cfg in MODULES.items():
        qs = _queryset(cle, request.user)
        total, nombre = _totaux(qs)
        total_mois, nombre_mois = _totaux(qs.filter(date__gte=debut_mois))
        if cfg['sens'] == 'sortie':
            total_sorties += total
            total_sorties_mois += total_mois
        else:
            total_entrees += total
            total_entrees_mois += total_mois
        cartes.append({
            'cle': cle, 'titre': cfg['titre'], 'icone': cfg['icone'],
            'couleur': cfg['couleur'], 'sens': cfg['sens'],
            'total': total, 'nombre': nombre,
            'total_mois': total_mois, 'nombre_mois': nombre_mois,
        })

    # Module informatique : compté à part, sa logique étant différente
    abonnements = filter_by_user_school(
        AbonnementInformatique.objects.all(), request.user, 'ecole')
    total_abo, nombre_abo = _totaux(abonnements)
    total_entrees += total_abo
    total_entrees_mois += _totaux(abonnements.filter(date__gte=debut_mois))[0]
    expirant = abonnements.filter(
        date_fin__gte=aujourdhui,
        date_fin__lte=aujourdhui + timedelta(days=AbonnementInformatique.SEUIL_ALERTE_JOURS),
    ).count()
    expires = abonnements.filter(date_fin__lt=aujourdhui).count()

    # Module salaires du personnel : registre autonome du recouvrement, compté
    # à part car ses montants sont saisis mois par mois et non à la date du jour.
    total_salaires, _ = totaux_personnel(request.user)
    total_salaires_mois, nombre_salaires_mois = totaux_personnel(
        request.user, annee=aujourdhui.year, mois=aujourdhui.month)
    membres = membres_du_personnel(request.user)
    effectif_personnel = membres.count()
    effectifs_par_groupe = {
        groupe['code']: membres.filter(groupe=groupe['code']).count()
        for groupe in GROUPES_PERSONNEL
    }

    # Dépenses générales (module Dépenses principal, factures fournisseurs) :
    # comptées à part, sa logique (validation, fournisseurs...) étant différente.
    depenses_qs = Depense.objects.all()
    if not user_is_superadmin(request.user):
        ecole = user_school(request.user)
        depenses_qs = depenses_qs.filter(cree_par__profil__ecole=ecole) if ecole else Depense.objects.none()
    total_depenses = depenses_qs.aggregate(total=Sum('montant_ttc'))['total'] or 0
    nombre_depenses = depenses_qs.count()
    depenses_mois_qs = depenses_qs.filter(date_facture__gte=debut_mois)
    total_depenses_mois = depenses_mois_qs.aggregate(total=Sum('montant_ttc'))['total'] or 0
    nombre_depenses_mois = depenses_mois_qs.count()
    total_sorties += total_depenses
    total_sorties_mois += total_depenses_mois

    # Versements et entrées : mis en avant à part dans le bandeau de synthèse
    total_versements = next((c['total'] for c in cartes if c['cle'] == 'versement'), 0)
    carte_entree = next((c for c in cartes if c['cle'] == 'entree'), None)

    contexte = {
        'titre_page': 'Recouvrement',
        'cartes': cartes,
        'carte_depenses': {
            'total': int(total_depenses), 'nombre': nombre_depenses,
            'total_mois': int(total_depenses_mois), 'nombre_mois': nombre_depenses_mois,
        },
        'total_versements': total_versements,
        'carte_entree': carte_entree,
        'carte_informatique': {
            'total': total_abo, 'nombre': nombre_abo,
            'expirant': expirant, 'expires': expires,
            'actifs': abonnements.filter(date_fin__gte=aujourdhui).count(),
        },
        'carte_salaires': {
            'total': total_salaires, 'effectif': effectif_personnel,
            'total_mois': total_salaires_mois, 'nombre_mois': nombre_salaires_mois,
            'effectifs': [
                {'libelle': groupe['libelle'], 'effectif': effectifs_par_groupe[groupe['code']]}
                for groupe in GROUPES_PERSONNEL
            ],
        },
        'total_sorties': total_sorties,
        'total_sorties_mois': total_sorties_mois,
        'total_entrees': total_entrees,
        'total_entrees_mois': total_entrees_mois,
        'solde': total_entrees - total_sorties,
        'mois_courant': debut_mois,
    }
    return render(request, 'depenses/recouvrement/hub.html', contexte)


# --------------------------------------------------------------------------
# Modules simples : liste, saisie, modification, suppression
# --------------------------------------------------------------------------

@login_required
def liste_operations(request, cle):
    """Liste des opérations du module, avec le formulaire de saisie intégré."""
    cfg = _config(cle)
    ecole = user_school(request.user)

    if request.method == 'POST':
        form = cfg['formulaire'](request.POST)
        if ecole is None:
            messages.error(request, "Aucun établissement n'est rattaché à votre compte.")
        elif form.is_valid():
            operation = form.save(commit=False)
            operation.ecole = ecole
            operation.cree_par = request.user
            operation.save()
            messages.success(request, f"{cfg['singulier'].capitalize()} enregistrée.")
            return redirect('depenses:recouvrement_liste', cle=cle)
        else:
            messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = cfg['formulaire']()

    qs, filtres = _filtrer_periode(_queryset(cle, request.user), request, cle)
    total, nombre = _totaux(qs)

    return render(request, 'depenses/recouvrement/liste.html', {
        'titre_page': cfg['titre'],
        'cle': cle, 'cfg': cfg, 'form': form,
        'operations': qs.select_related('cree_par')[:300],
        'total': total, 'nombre': nombre, 'filtres': filtres,
    })


@login_required
def modifier_operation(request, cle, pk):
    cfg = _config(cle)
    operation = get_object_or_404(_queryset(cle, request.user), pk=pk)

    if request.method == 'POST':
        form = cfg['formulaire'](request.POST, instance=operation)
        if form.is_valid():
            form.save()
            messages.success(request, "Modification enregistrée.")
            return redirect('depenses:recouvrement_liste', cle=cle)
        messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = cfg['formulaire'](instance=operation)

    return render(request, 'depenses/recouvrement/form.html', {
        'titre_page': f"Modifier — {cfg['titre']}",
        'cle': cle, 'cfg': cfg, 'form': form, 'operation': operation,
    })


@login_required
def supprimer_operation(request, cle, pk):
    cfg = _config(cle)
    operation = get_object_or_404(_queryset(cle, request.user), pk=pk)
    if request.method == 'POST':
        operation.delete()
        messages.success(request, f"{cfg['singulier'].capitalize()} supprimée.")
        return redirect('depenses:recouvrement_liste', cle=cle)
    return render(request, 'depenses/recouvrement/confirmer_suppression.html', {
        'titre_page': "Confirmer la suppression",
        'cle': cle, 'cfg': cfg, 'operation': operation,
    })


@login_required
def tableau_bord_module(request, cle):
    """Tableau de bord propre au module : totaux, évolution et derniers postes."""
    cfg = _config(cle)
    qs = _queryset(cle, request.user)
    aujourdhui = timezone.localdate()
    debut_mois = aujourdhui.replace(day=1)

    total, nombre = _totaux(qs)
    total_mois, nombre_mois = _totaux(qs.filter(date__gte=debut_mois))
    total_annee, _ = _totaux(qs.filter(date__year=aujourdhui.year))

    par_mois = list(
        qs.annotate(mois=TruncMonth('date'))
          .values('mois')
          .annotate(total=Sum('montant'), nombre=Count('id'))
          .order_by('-mois')[:12]
    )
    champ = cfg['champ_libelle']
    par_poste = list(
        qs.values(champ)
          .annotate(total=Sum('montant'), nombre=Count('id'))
          .order_by('-total')[:10]
    )
    for poste in par_poste:
        poste['libelle'] = poste.pop(champ)

    # Répartition par catégorie, pour les modules qui en déclarent une (Entrées)
    filtre_choix = cfg.get('filtre_choix')
    par_categorie = []
    if filtre_choix:
        libelles = dict(filtre_choix['choix'])
        par_categorie = [
            {
                'libelle': libelles.get(ligne[filtre_choix['champ']], ligne[filtre_choix['champ']]),
                'total': ligne['total'], 'nombre': ligne['nombre'],
            }
            for ligne in qs.values(filtre_choix['champ'])
                           .annotate(total=Sum('montant'), nombre=Count('id'))
                           .order_by('-total')
        ]

    return render(request, 'depenses/recouvrement/tableau_bord.html', {
        'titre_page': f"Tableau de bord — {cfg['titre']}",
        'cle': cle, 'cfg': cfg,
        'total': total, 'nombre': nombre,
        'total_mois': total_mois, 'nombre_mois': nombre_mois,
        'total_annee': total_annee,
        'moyenne': int(total / nombre) if nombre else 0,
        'par_mois': par_mois, 'par_poste': par_poste,
        'par_categorie': par_categorie,
        'titre_categorie': filtre_choix['label'] if filtre_choix else '',
        'dernieres': qs.select_related('cree_par')[:10],
    })


# --------------------------------------------------------------------------
# Exports des modules simples
# --------------------------------------------------------------------------

def _lignes_export(cle, request):
    """Entêtes et lignes de l'export.

    Le montant occupe toujours l'avant-dernière colonne et l'observation la
    dernière : les colonnes propres au module s'insèrent entre l'intitulé et le
    montant.
    """
    cfg = _config(cle)
    qs, filtres = _filtrer_periode(_queryset(cle, request.user), request, cle)
    champ = cfg['champ_libelle']
    extras = cfg.get('colonnes_extra', [])

    entetes = (
        ['Date', cfg['entete_libelle']]
        + [entete for _, entete in extras]
        + ['Montant (GNF)', 'Observation']
    )
    lignes = [
        [
            op.date.strftime('%d/%m/%Y'),
            getattr(op, champ),
            *[getattr(op, attribut) or '' for attribut, _ in extras],
            int(op.montant),
            op.observation or '',
        ]
        for op in qs
    ]
    return cfg, entetes, lignes, filtres


@login_required
def export_excel_module(request, cle):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    cfg, entetes, lignes, _ = _lignes_export(cle, request)
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = cfg['titre'][:31]

    feuille.append(entetes)
    for cellule in feuille[1]:
        cellule.font = Font(bold=True, color='FFFFFF')
        cellule.fill = PatternFill('solid', fgColor='1657A8')
        cellule.alignment = Alignment(horizontal='center')

    for ligne in lignes:
        feuille.append(ligne)

    total = sum(l[-2] for l in lignes)
    feuille.append([])
    feuille.append([''] * (len(entetes) - 3) + ['TOTAL', total, ''])
    for colonne in (len(entetes) - 2, len(entetes) - 1):
        feuille.cell(row=feuille.max_row, column=colonne).font = Font(bold=True)

    largeurs = [14, 42] + [20] * (len(entetes) - 4) + [18, 40]
    for i, largeur in enumerate(largeurs, start=1):
        feuille.column_dimensions[get_column_letter(i)].width = largeur

    flux = io.BytesIO()
    classeur.save(flux)
    reponse = HttpResponse(
        flux.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    nom = f"{cle}_{date.today():%Y%m%d}.xlsx"
    reponse['Content-Disposition'] = f'attachment; filename="{nom}"'
    return reponse


@login_required
def export_pdf_module(request, cle):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    cfg, entetes, lignes, filtres = _lignes_export(cle, request)
    ecole = user_school(request.user)
    nb_extras = len(entetes) - 4
    paysage = nb_extras > 0  # les modules à colonnes supplémentaires respirent mieux à l'horizontale

    flux = io.BytesIO()
    format_page = landscape(A4) if paysage else A4
    document = SimpleDocTemplate(
        flux, pagesize=format_page,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"<b>{cfg['titre']}</b>", styles['Title']),
        Paragraph(getattr(ecole, 'nom', '') or '', styles['Normal']),
    ]
    if filtres['date_debut'] or filtres['date_fin']:
        elements.append(Paragraph(
            f"Période : {filtres['date_debut'] or '—'} au {filtres['date_fin'] or '—'}",
            styles['Normal']))
    if filtres.get('type') and cfg.get('filtre_choix'):
        libelles = dict(cfg['filtre_choix']['choix'])
        elements.append(Paragraph(
            f"{cfg['filtre_choix']['label']} : {libelles.get(filtres['type'], filtres['type'])}",
            styles['Normal']))
    elements.append(Spacer(1, 0.6 * cm))

    donnees = [entetes]
    for ligne in lignes:
        donnees.append([
            ligne[0],
            Paragraph(str(ligne[1]), styles['BodyText']),
            *[str(valeur) for valeur in ligne[2:-2]],
            f"{ligne[-2]:,}".replace(',', ' '),
            Paragraph(str(ligne[-1]), styles['BodyText']),
        ])
    total = sum(l[-2] for l in lignes)
    donnees.append([''] * (len(entetes) - 3) + ['TOTAL', f"{total:,}".replace(',', ' '), ''])

    # Largeurs proportionnelles, ajustées à la largeur utile de la page
    poids = [1.6, 4.0] + [2.2] * nb_extras + [2.2, 4.0]
    largeur_utile = format_page[0] - 3 * cm
    colonne_montant = len(entetes) - 2

    tableau = Table(
        donnees, repeatRows=1,
        colWidths=[p / sum(poids) * largeur_utile for p in poids],
    )
    tableau.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1657A8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EAF1FB')),
        ('ALIGN', (colonne_montant, 1), (colonne_montant, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#B9C7D9')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    elements.append(tableau)
    document.build(elements)

    reponse = HttpResponse(flux.getvalue(), content_type='application/pdf')
    nom = f"{cle}_{date.today():%Y%m%d}.pdf"
    reponse['Content-Disposition'] = f'attachment; filename="{nom}"'
    return reponse


# --------------------------------------------------------------------------
# Module informatique
# --------------------------------------------------------------------------

def _abonnements(user):
    return filter_by_user_school(
        AbonnementInformatique.objects.select_related('eleve', 'eleve__classe'),
        user, 'ecole')


@login_required
def informatique_liste(request):
    """Élèves et abonnements, avec recherche par matricule ou par nom."""
    recherche = (request.GET.get('q') or '').strip()
    statut = (request.GET.get('statut') or '').strip()
    aujourdhui = timezone.localdate()

    qs = _abonnements(request.user)
    if recherche:
        qs = qs.filter(
            Q(eleve__matricule__icontains=recherche)
            | Q(eleve__nom__icontains=recherche)
            | Q(eleve__prenom__icontains=recherche)
        )
    if statut == 'ACTIF':
        qs = qs.filter(date_fin__gte=aujourdhui)
    elif statut == 'EXPIRE':
        qs = qs.filter(date_fin__lt=aujourdhui)
    elif statut == 'BIENTOT':
        qs = qs.filter(
            date_fin__gte=aujourdhui,
            date_fin__lte=aujourdhui + timedelta(days=AbonnementInformatique.SEUIL_ALERTE_JOURS),
        )

    # Élèves sans abonnement en cours : la liste depuis laquelle on en crée un
    eleves = filter_by_user_school(
        Eleve.objects.filter(statut='ACTIF').select_related('classe'),
        request.user, 'classe__ecole')
    if recherche:
        eleves = eleves.filter(
            Q(matricule__icontains=recherche)
            | Q(nom__icontains=recherche)
            | Q(prenom__icontains=recherche)
        )
    eleves = eleves.exclude(
        abonnements_informatique__date_fin__gte=aujourdhui
    ).order_by('nom', 'prenom')[:100]

    return render(request, 'depenses/recouvrement/informatique_liste.html', {
        'titre_page': 'Abonnements informatique',
        'abonnements': qs[:300],
        'eleves_sans_abonnement': eleves,
        'q': recherche, 'statut': statut,
        'alertes': _abonnements(request.user).filter(
            date_fin__gte=aujourdhui,
            date_fin__lte=aujourdhui + timedelta(days=AbonnementInformatique.SEUIL_ALERTE_JOURS),
        ),
    })


@login_required
def informatique_nouveau(request, eleve_id=None):
    ecole = user_school(request.user)
    initial = {}
    if eleve_id:
        eleves = filter_by_user_school(Eleve.objects.all(), request.user, 'classe__ecole')
        initial['eleve'] = get_object_or_404(eleves, pk=eleve_id)

    if request.method == 'POST':
        form = AbonnementInformatiqueForm(request.POST, user=request.user)
        if ecole is None:
            messages.error(request, "Aucun établissement n'est rattaché à votre compte.")
        elif form.is_valid():
            abonnement = form.save(commit=False)
            abonnement.ecole = ecole
            abonnement.cree_par = request.user
            abonnement.save()
            messages.success(request, "Abonnement enregistré.")
            return redirect('depenses:informatique_carte', pk=abonnement.pk)
        else:
            messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = AbonnementInformatiqueForm(user=request.user, initial=initial)

    return render(request, 'depenses/recouvrement/informatique_form.html', {
        'titre_page': "Nouvel abonnement informatique",
        'form': form,
    })


@login_required
def informatique_carte(request, pk):
    """Carte d'abonnement imprimable."""
    abonnement = get_object_or_404(_abonnements(request.user), pk=pk)
    return render(request, 'depenses/recouvrement/informatique_carte.html', {
        'titre_page': "Carte d'abonnement",
        'abonnement': abonnement,
        'ecole': abonnement.ecole,
    })


@login_required
def informatique_tableau_bord(request):
    qs = _abonnements(request.user)
    aujourdhui = timezone.localdate()
    seuil = aujourdhui + timedelta(days=AbonnementInformatique.SEUIL_ALERTE_JOURS)

    total, nombre = _totaux(qs)
    actifs = qs.filter(date_fin__gte=aujourdhui)
    par_mois = list(
        qs.annotate(mois=TruncMonth('date'))
          .values('mois').annotate(total=Sum('montant'), nombre=Count('id'))
          .order_by('-mois')[:12]
    )
    return render(request, 'depenses/recouvrement/informatique_tableau_bord.html', {
        'titre_page': "Tableau de bord — Informatique",
        'total': total, 'nombre': nombre,
        'nb_actifs': actifs.count(),
        'nb_expires': qs.filter(date_fin__lt=aujourdhui).count(),
        'nb_bientot': qs.filter(date_fin__gte=aujourdhui, date_fin__lte=seuil).count(),
        'moyenne': int(total / nombre) if nombre else 0,
        'par_mois': par_mois,
        'expirations': qs.filter(date_fin__gte=aujourdhui, date_fin__lte=seuil)[:20],
        'derniers': qs[:10],
    })


def _lignes_informatique(request):
    qs = _abonnements(request.user)
    recherche = (request.GET.get('q') or '').strip()
    if recherche:
        qs = qs.filter(
            Q(eleve__matricule__icontains=recherche)
            | Q(eleve__nom__icontains=recherche)
            | Q(eleve__prenom__icontains=recherche)
        )
    return [
        [
            a.eleve.matricule or '',
            f"{a.eleve.prenom} {a.eleve.nom}",
            getattr(a.eleve.classe, 'nom', '') or '',
            a.date_debut.strftime('%d/%m/%Y'),
            a.date_fin.strftime('%d/%m/%Y'),
            int(a.montant),
            a.statut_libelle,
        ]
        for a in qs
    ]


@login_required
def informatique_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    lignes = _lignes_informatique(request)
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = 'Abonnements'
    feuille.append(['Matricule', 'Élève', 'Classe', 'Début', 'Fin', 'Montant (GNF)', 'Statut'])
    for cellule in feuille[1]:
        cellule.font = Font(bold=True, color='FFFFFF')
        cellule.fill = PatternFill('solid', fgColor='1657A8')
        cellule.alignment = Alignment(horizontal='center')
    for ligne in lignes:
        feuille.append(ligne)
    feuille.append([])
    feuille.append(['', '', '', '', 'TOTAL', sum(l[5] for l in lignes), ''])
    for colonne, largeur in zip('ABCDEFG', (16, 30, 18, 13, 13, 16, 16)):
        feuille.column_dimensions[colonne].width = largeur

    flux = io.BytesIO()
    classeur.save(flux)
    reponse = HttpResponse(
        flux.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    reponse['Content-Disposition'] = (
        f'attachment; filename="abonnements_informatique_{date.today():%Y%m%d}.xlsx"')
    return reponse


@login_required
def informatique_export_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    lignes = _lignes_informatique(request)
    ecole = user_school(request.user)

    flux = io.BytesIO()
    document = SimpleDocTemplate(flux, pagesize=landscape(A4),
                                 leftMargin=1.2 * cm, rightMargin=1.2 * cm,
                                 topMargin=1.2 * cm, bottomMargin=1.2 * cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("<b>Abonnements informatique</b>", styles['Title']),
        Paragraph(getattr(ecole, 'nom', '') or '', styles['Normal']),
        Spacer(1, 0.5 * cm),
    ]

    donnees = [['Matricule', 'Élève', 'Classe', 'Début', 'Fin', 'Montant', 'Statut']]
    for l in lignes:
        donnees.append([l[0], Paragraph(l[1], styles['BodyText']), l[2], l[3], l[4],
                        f"{l[5]:,}".replace(',', ' '), l[6]])
    donnees.append(['', '', '', '', 'TOTAL',
                    f"{sum(l[5] for l in lignes):,}".replace(',', ' '), ''])

    tableau = Table(donnees, repeatRows=1,
                    colWidths=[3 * cm, 6 * cm, 4 * cm, 2.8 * cm, 2.8 * cm, 3 * cm, 3 * cm])
    tableau.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1657A8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#B9C7D9')),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    elements.append(tableau)
    document.build(elements)

    reponse = HttpResponse(flux.getvalue(), content_type='application/pdf')
    reponse['Content-Disposition'] = (
        f'attachment; filename="abonnements_informatique_{date.today():%Y%m%d}.pdf"')
    return reponse
