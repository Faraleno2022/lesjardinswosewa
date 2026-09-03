"""Module « Salaires du personnel » du recouvrement.

Registre autonome, sans lien avec l'app `salaires` : on ajoute directement un
membre du personnel (prénom, nom, fonction, groupe) et l'on saisit ses montants
mois par mois, à même le tableau.

Trois groupes sont suivis séparément — maternelle et primaire, collège, équipe
de direction — puis consolidés dans une statistique globale des salaires.
"""
import io
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from utilisateurs.utils import filter_by_user_school, user_school

from .forms_recouvrement import MembrePersonnelForm
from .models_recouvrement import MembrePersonnel, SalaireMensuelPersonnel

# Mois de l'année : numéro, libellé complet et abréviation pour les entêtes.
MOIS = [
    (1, 'Janvier', 'Jan'), (2, 'Février', 'Fév'), (3, 'Mars', 'Mar'),
    (4, 'Avril', 'Avr'), (5, 'Mai', 'Mai'), (6, 'Juin', 'Juin'),
    (7, 'Juillet', 'Juil'), (8, 'Août', 'Août'), (9, 'Septembre', 'Sep'),
    (10, 'Octobre', 'Oct'), (11, 'Novembre', 'Nov'), (12, 'Décembre', 'Déc'),
]

MOIS_ENTETES = [{'numero': n, 'nom': nom, 'court': court} for n, nom, court in MOIS]

MOIS_NOMS = {n: nom for n, nom, _ in MOIS}

# Présentation des trois groupes suivis, dans l'ordre d'affichage voulu.
GROUPES = [
    {
        'code': MembrePersonnel.GROUPE_MATERNELLE_PRIMAIRE,
        'libelle': 'Maternelle et primaire',
        'icone': 'fas fa-children',
        'couleur': 'success',
    },
    {
        'code': MembrePersonnel.GROUPE_COLLEGE,
        'libelle': 'Collège',
        'icone': 'fas fa-school',
        'couleur': 'info',
    },
    {
        'code': MembrePersonnel.GROUPE_DIRECTION,
        'libelle': 'Équipe de direction',
        'icone': 'fas fa-user-tie',
        'couleur': 'primary',
    },
]

# Préfixe des champs de saisie des montants : montant_<membre>_<mois>.
PREFIXE_MONTANT = 'montant'


def membres_du_personnel(user):
    """Membres du personnel visibles par l'utilisateur (cloisonnement école)."""
    return filter_by_user_school(MembrePersonnel.objects.all(), user, 'ecole')


def salaires_du_personnel(user):
    """Montants mensuels saisis, restreints à l'école de l'utilisateur."""
    return filter_by_user_school(
        SalaireMensuelPersonnel.objects.all(), user, 'membre__ecole')


def totaux_personnel(user, annee=None, mois=None):
    """Total versé et nombre de montants saisis, éventuellement bornés."""
    qs = salaires_du_personnel(user)
    if annee is not None:
        qs = qs.filter(annee=annee)
    if mois is not None:
        qs = qs.filter(mois=mois)
    agg = qs.aggregate(total=Sum('montant'))
    return int(agg['total'] or 0), qs.count()


def _annee_demandee(request, annees_disponibles, annee_courante):
    """Année choisie dans le sélecteur, avec repli sur l'année en cours."""
    defaut = annees_disponibles[0] if annees_disponibles else annee_courante
    try:
        return int(request.GET.get('annee') or request.POST.get('annee') or defaut)
    except (TypeError, ValueError):
        return defaut


def _annees_disponibles(user, annee_courante):
    annees = set(
        salaires_du_personnel(user).values_list('annee', flat=True).distinct()
    )
    annees.add(annee_courante)
    return sorted(annees, reverse=True)


def _construire_groupes(user, annee):
    """Construit les trois tableaux (un par groupe) pour l'année demandée.

    Chaque ligne porte le membre, ses douze cases mensuelles et son total ;
    chaque groupe porte ses totaux par mois, son total et son effectif.
    """
    membres = list(membres_du_personnel(user))
    montants = {
        (salaire.membre_id, salaire.mois): int(salaire.montant)
        for salaire in salaires_du_personnel(user).filter(annee=annee)
    }

    groupes = []
    for definition in GROUPES:
        lignes = []
        for membre in membres:
            if membre.groupe != definition['code']:
                continue
            cellules = [
                {'mois': entete['numero'], 'montant': montants.get((membre.id, entete['numero']), 0)}
                for entete in MOIS_ENTETES
            ]
            lignes.append({
                'membre': membre,
                'cellules': cellules,
                'total': sum(cellule['montant'] for cellule in cellules),
            })

        totaux_mois = [
            {
                'mois': entete['numero'],
                'total': sum(ligne['cellules'][index]['montant'] for ligne in lignes),
            }
            for index, entete in enumerate(MOIS_ENTETES)
        ]
        total_groupe = sum(ligne['total'] for ligne in lignes)
        effectif = len(lignes)
        groupes.append({
            **definition,
            'lignes': lignes,
            'totaux_mois': totaux_mois,
            'total': total_groupe,
            'effectif': effectif,
            'moyenne': int(total_groupe / effectif) if effectif else 0,
        })
    return groupes


def _statistique_globale(groupes, annee, mois_courant):
    """Synthèse consolidée des trois groupes : totaux, parts et repères."""
    total_general = sum(groupe['total'] for groupe in groupes)
    effectif_total = sum(groupe['effectif'] for groupe in groupes)

    totaux_mois = [
        {
            'mois': entete['numero'],
            'nom': entete['nom'],
            'court': entete['court'],
            'total': sum(groupe['totaux_mois'][index]['total'] for groupe in groupes),
        }
        for index, entete in enumerate(MOIS_ENTETES)
    ]
    mois_renseignes = [ligne for ligne in totaux_mois if ligne['total']]
    maximum = max((ligne['total'] for ligne in totaux_mois), default=0) or 1
    for ligne in totaux_mois:
        ligne['part'] = round(ligne['total'] * 100 / maximum)

    for groupe in groupes:
        groupe['part'] = round(groupe['total'] * 100 / total_general) if total_general else 0

    mois_le_plus_lourd = max(mois_renseignes, key=lambda l: l['total'], default=None)
    total_mois_courant = next(
        (ligne['total'] for ligne in totaux_mois if ligne['mois'] == mois_courant), 0
    )

    return {
        'annee': annee,
        'total_general': total_general,
        'effectif_total': effectif_total,
        'totaux_mois': totaux_mois,
        'nombre_mois_renseignes': len(mois_renseignes),
        'moyenne_mensuelle': int(total_general / len(mois_renseignes)) if mois_renseignes else 0,
        'moyenne_par_membre': int(total_general / effectif_total) if effectif_total else 0,
        'mois_le_plus_lourd': mois_le_plus_lourd,
        'total_mois_courant': total_mois_courant,
    }


def _nettoyer_montant(valeur):
    """Convertit une saisie en entier, en tolérant espaces et séparateurs.

    Retourne `None` si la case est vide, et lève `ValueError` si la saisie
    n'est pas un montant exploitable.
    """
    texte = (valeur or '').strip()
    if not texte:
        return None
    texte = texte.replace(' ', '').replace(' ', '').replace(',', '.')
    try:
        montant = Decimal(texte)
    except (InvalidOperation, ArithmeticError):
        raise ValueError(f"Montant invalide : {valeur}")
    if montant < 0:
        raise ValueError("Un salaire ne peut pas être négatif.")
    return int(montant)


def _enregistrer_montants(request, annee):
    """Enregistre les montants mensuels saisis dans le tableau.

    Une case vidée supprime le montant du mois : le tableau reflète exactement
    ce que l'utilisateur voit.
    """
    membres = {membre.id: membre for membre in membres_du_personnel(request.user)}
    existants = {
        (salaire.membre_id, salaire.mois): salaire
        for salaire in salaires_du_personnel(request.user).filter(annee=annee)
    }

    modifies = supprimes = 0
    erreurs = []
    a_creer = []
    for cle, valeur in request.POST.items():
        morceaux = cle.split('_')
        if len(morceaux) != 3 or morceaux[0] != PREFIXE_MONTANT:
            continue
        try:
            membre_id, mois = int(morceaux[1]), int(morceaux[2])
        except ValueError:
            continue
        if membre_id not in membres or not 1 <= mois <= 12:
            continue

        try:
            montant = _nettoyer_montant(valeur)
        except ValueError:
            erreurs.append(f"{membres[membre_id].nom_complet} ({MOIS_NOMS[mois]})")
            continue

        existant = existants.get((membre_id, mois))
        if montant is None or montant == 0:
            if existant is not None:
                existant.delete()
                supprimes += 1
            continue
        if existant is None:
            a_creer.append(SalaireMensuelPersonnel(
                membre=membres[membre_id], annee=annee, mois=mois, montant=montant))
            modifies += 1
        elif int(existant.montant) != montant:
            existant.montant = montant
            existant.save(update_fields=['montant', 'date_modification'])
            modifies += 1

    if a_creer:
        SalaireMensuelPersonnel.objects.bulk_create(a_creer)

    if erreurs:
        messages.error(
            request,
            "Montants ignorés car invalides : " + ', '.join(erreurs[:10])
            + ('…' if len(erreurs) > 10 else ''),
        )
    if modifies or supprimes:
        messages.success(
            request,
            f"Salaires enregistrés : {modifies} montant{'s' if modifies > 1 else ''} "
            f"mis à jour, {supprimes} effacé{'s' if supprimes > 1 else ''}.",
        )
    elif not erreurs:
        messages.info(request, "Aucun changement à enregistrer.")


@login_required
def personnel_dashboard(request):
    """Page unique du module : ajout d'un membre, tableaux et statistique globale."""
    aujourdhui = timezone.localdate()
    annees_disponibles = _annees_disponibles(request.user, aujourdhui.year)
    annee = _annee_demandee(request, annees_disponibles, aujourdhui.year)
    form = MembrePersonnelForm()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'montants':
            _enregistrer_montants(request, annee)
            return redirect(f"{request.path}?annee={annee}")
        if action == 'ajouter':
            ecole = user_school(request.user)
            form = MembrePersonnelForm(request.POST)
            if ecole is None:
                messages.error(request, "Aucun établissement n'est rattaché à votre compte.")
            elif form.is_valid():
                membre = form.save(commit=False)
                membre.ecole = ecole
                membre.cree_par = request.user
                membre.save()
                messages.success(
                    request, f"{membre.nom_complet} ajouté à « {membre.groupe_libelle} ».")
                return redirect(f"{request.path}?annee={annee}")
            else:
                messages.error(request, "Veuillez corriger les erreurs du formulaire.")

    if annee not in annees_disponibles:
        annees_disponibles = sorted(set(annees_disponibles) | {annee}, reverse=True)

    groupes = _construire_groupes(request.user, annee)
    statistique = _statistique_globale(groupes, annee, aujourdhui.month)

    return render(request, 'depenses/recouvrement/personnel_salaires.html', {
        'titre_page': 'Recouvrement — Salaires du personnel',
        'annee': annee,
        'annees_disponibles': annees_disponibles,
        'mois_entetes': MOIS_ENTETES,
        'groupes': groupes,
        'statistique': statistique,
        'form': form,
        'mois_courant_nom': MOIS_NOMS[aujourdhui.month],
        'prefixe_montant': PREFIXE_MONTANT,
    })


@login_required
def personnel_modifier(request, pk):
    """Modification de l'identité, de la fonction ou du groupe d'un membre."""
    membre = get_object_or_404(membres_du_personnel(request.user), pk=pk)
    annee = request.GET.get('annee') or request.POST.get('annee') or timezone.localdate().year

    if request.method == 'POST':
        form = MembrePersonnelForm(request.POST, instance=membre)
        if form.is_valid():
            form.save()
            messages.success(request, "Modification enregistrée.")
            return redirect(f"{reverse('depenses:personnel_dashboard')}?annee={annee}")
        messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = MembrePersonnelForm(instance=membre)

    return render(request, 'depenses/recouvrement/personnel_form.html', {
        'titre_page': f"Modifier — {membre.nom_complet}",
        'form': form, 'membre': membre, 'annee': annee,
    })


@login_required
def personnel_supprimer(request, pk):
    """Retrait d'un membre et de tous ses montants mensuels."""
    membre = get_object_or_404(membres_du_personnel(request.user), pk=pk)
    annee = request.GET.get('annee') or timezone.localdate().year

    if request.method == 'POST':
        nom = membre.nom_complet
        membre.delete()
        messages.success(request, f"{nom} a été retiré du registre des salaires.")
        return redirect(f"{reverse('depenses:personnel_dashboard')}?annee={annee}")

    return render(request, 'depenses/recouvrement/personnel_confirmer_suppression.html', {
        'titre_page': "Confirmer le retrait",
        'membre': membre, 'annee': annee,
        'nombre_montants': membre.salaires_mensuels.count(),
    })


@login_required
def personnel_export_excel(request):
    """Export Excel : un bloc par groupe, puis la statistique globale."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    aujourdhui = timezone.localdate()
    annee = _annee_demandee(request, _annees_disponibles(request.user, aujourdhui.year),
                            aujourdhui.year)
    groupes = _construire_groupes(request.user, annee)
    statistique = _statistique_globale(groupes, annee, aujourdhui.month)

    classeur = Workbook()
    feuille = classeur.active
    feuille.title = f'Salaires {annee}'

    gras_blanc = Font(bold=True, color='FFFFFF')
    fond_bleu = PatternFill('solid', fgColor='1657A8')
    fond_gris = PatternFill('solid', fgColor='E9EDF3')
    centre = Alignment(horizontal='center')

    feuille.append([f"Salaires du personnel — année {annee}"])
    feuille['A1'].font = Font(bold=True, size=14)
    feuille.append([])

    entetes = ['Prénom', 'Nom', 'Fonction'] + [nom for _, nom, _ in MOIS] + ['Total (GNF)']

    for groupe in groupes:
        feuille.append([groupe['libelle']])
        feuille.cell(row=feuille.max_row, column=1).font = Font(bold=True, size=12)

        feuille.append(entetes)
        for cellule in feuille[feuille.max_row]:
            cellule.font = gras_blanc
            cellule.fill = fond_bleu
            cellule.alignment = centre

        for ligne in groupe['lignes']:
            feuille.append([
                ligne['membre'].prenom,
                ligne['membre'].nom,
                ligne['membre'].fonction,
                *[cellule['montant'] for cellule in ligne['cellules']],
                ligne['total'],
            ])

        feuille.append(
            ['', '', f"Sous-total {groupe['libelle']}"]
            + [mois['total'] for mois in groupe['totaux_mois']]
            + [groupe['total']]
        )
        for colonne in range(1, len(entetes) + 1):
            cellule = feuille.cell(row=feuille.max_row, column=colonne)
            cellule.font = Font(bold=True)
            cellule.fill = fond_gris
        feuille.append([])

    feuille.append(['', '', 'TOTAL GÉNÉRAL']
                   + [mois['total'] for mois in statistique['totaux_mois']]
                   + [statistique['total_general']])
    for colonne in range(1, len(entetes) + 1):
        cellule = feuille.cell(row=feuille.max_row, column=colonne)
        cellule.font = gras_blanc
        cellule.fill = fond_bleu

    feuille.append([])
    feuille.append(['Statistique globale des salaires'])
    feuille.cell(row=feuille.max_row, column=1).font = Font(bold=True, size=12)
    synthese = [
        ('Total versé sur l\'année', statistique['total_general']),
        ('Effectif suivi', statistique['effectif_total']),
        ('Moyenne par mois renseigné', statistique['moyenne_mensuelle']),
        ('Moyenne par membre', statistique['moyenne_par_membre']),
        ('Mois le plus lourd',
         f"{statistique['mois_le_plus_lourd']['nom']} "
         f"({statistique['mois_le_plus_lourd']['total']} GNF)"
         if statistique['mois_le_plus_lourd'] else '—'),
    ]
    for libelle, valeur in synthese:
        feuille.append([libelle, valeur])
        feuille.cell(row=feuille.max_row, column=1).font = Font(bold=True)
    for groupe in groupes:
        feuille.append([
            f"{groupe['libelle']} ({groupe['effectif']} personne"
            f"{'s' if groupe['effectif'] > 1 else ''})",
            groupe['total'], f"{groupe['part']} %",
        ])

    largeurs = [20, 20, 30] + [13] * 12 + [16]
    for index, largeur in enumerate(largeurs, start=1):
        feuille.column_dimensions[get_column_letter(index)].width = largeur

    flux = io.BytesIO()
    classeur.save(flux)
    reponse = HttpResponse(
        flux.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    reponse['Content-Disposition'] = (
        f'attachment; filename="salaires_personnel_{annee}.xlsx"')
    return reponse
