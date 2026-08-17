"""Rapport professionnel des encaissements scolaires, en PDF et Excel."""

from collections import defaultdict
from copy import deepcopy
from datetime import date
from decimal import Decimal
from io import BytesIO
import re
import uuid
from xml.sax.saxutils import escape

from django.http import HttpResponse
from django.db.models import Count, Sum
from django.utils import timezone
from django.utils.dateparse import parse_date

from eleves.models import Classe
from eleves.utils_annee import get_annee_active
from rapports.utils import _get_logo_path
from utilisateurs.permissions import can_view_reports
from utilisateurs.utils import filter_by_user_school, user_school

from .allocation import allocate_amount_sequentially, due_balances
from .calculs import est_type_scolarite, filtre_types_scolarite
from .models import EcheancierPaiement, Paiement, PaiementRemise
from .services import calculer_situations_echeanciers


ZERO = Decimal("0")
BLUE = "174A6E"
LIGHT_BLUE = "DCEAF3"
GREEN = "207A54"
ORANGE = "C2761C"
RED = "B53A3A"
GREY = "5D6973"

COMPONENTS = (
    ("inscription", "Inscription"),
    ("reinscription", "Réinscription"),
    ("tranche_1", "Tranche 1"),
    ("tranche_2", "Tranche 2"),
    ("tranche_3", "Tranche 3"),
    ("autres", "Autres services"),
    ("non_affecte", "Non affecté / à contrôler"),
)


def _money(value):
    return f"{int(value or 0):,}".replace(",", " ")


def _percentage(amount, base):
    """Retourne un taux effectif borné, exprimé de 0 à 100."""
    amount = Decimal(str(amount or 0))
    base = Decimal(str(base or 0))
    if amount <= 0 or base <= 0:
        return ZERO
    return min(Decimal("100"), amount * Decimal("100") / base)


def _settlement_label(total_due, coverage, discount):
    total_due = Decimal(str(total_due or 0))
    coverage = Decimal(str(coverage or 0))
    discount = Decimal(str(discount or 0))
    if total_due <= 0:
        return "Échéancier vide"
    if coverage >= total_due:
        return "Soldé - remise appliquée" if discount else "Soldé"
    if coverage > 0:
        return "Partiel - remise appliquée" if discount else "Partiel"
    return "À payer"


def _safe_filename(value):
    return re.sub(r"[^\w-]+", "_", value or "").strip("_") or "etablissement"


def _display_user(user):
    if not user:
        return "Système"
    full_name = (user.get_full_name() or "").strip()
    return full_name or user.get_username() or "Système"


def _excel_safe(value):
    """Empêche une donnée saisie de devenir une formule dans le classeur."""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _requires_reference(mode_name):
    label = (mode_name or "").casefold()
    return not any(token in label for token in ("espèce", "espece", "cash", "caisse"))


def _read_date(request, *names):
    for name in names:
        raw = (request.GET.get(name) or "").strip()
        if not raw:
            continue
        parsed = parse_date(raw)
        if parsed is None:
            raise ValueError(f"La date « {name} » doit être au format AAAA-MM-JJ.")
        return parsed
    return None


def _scope(request):
    start = _read_date(request, "du", "date_debut")
    requested_end = _read_date(request, "au", "date_fin")
    today = timezone.localdate()
    if start and start > today:
        raise ValueError("La date de début ne peut pas être postérieure à aujourd'hui.")
    if start and requested_end and start > requested_end:
        raise ValueError("La date de début doit précéder la date de fin.")
    end = min(requested_end or today, today)

    school_id = (request.GET.get("ecole_id") or "").strip()
    class_id = (request.GET.get("classe_id") or request.GET.get("classe") or "").strip()
    classes = filter_by_user_school(
        Classe.objects.select_related("ecole").order_by("ecole__nom", "niveau", "nom"),
        request.user,
        "ecole",
    )
    if school_id:
        if not school_id.isdigit():
            raise ValueError("L'école sélectionnée est invalide.")
        classes = classes.filter(ecole_id=int(school_id))
    if class_id:
        if not class_id.isdigit():
            raise ValueError("La classe sélectionnée est invalide.")
        classes = classes.filter(pk=int(class_id))
    classes = list(classes)
    if school_id and not classes:
        raise ValueError("L'école sélectionnée est introuvable ou non autorisée.")
    if class_id and not classes:
        raise ValueError("La classe sélectionnée est introuvable ou non autorisée.")

    school = classes[0].ecole if classes else user_school(request.user)
    school_ids = {item.ecole_id for item in classes}
    requested_year = (request.GET.get("annee_scolaire") or "").strip()
    if requested_year and not re.fullmatch(r"\d{4}-\d{4}", requested_year):
        raise ValueError("L'année scolaire doit être au format AAAA-AAAA.")
    if requested_year:
        school_year = requested_year
    elif len(classes) == 1:
        school_year = classes[0].annee_scolaire
    elif school and len(school_ids) <= 1:
        school_year = get_annee_active(request, school) or ""
    else:
        school_year = ""
    if school_year:
        classes = [item for item in classes if item.annee_scolaire == school_year]

    if len(classes) == 1:
        scope_label = f"Classe : {classes[0].nom}"
    elif classes and len({item.ecole_id for item in classes}) == 1:
        scope_label = "Tout l'établissement"
    elif classes:
        scope_label = "Établissements autorisés"
    else:
        scope_label = "Aucune classe dans le périmètre"

    generated_at = timezone.localtime()
    return {
        "classes": classes,
        "class_ids": [item.pk for item in classes],
        "school": school if len(school_ids) <= 1 else None,
        "school_name": school.nom if school and len(school_ids) <= 1 else "ÉTABLISSEMENTS AUTORISÉS",
        "school_year": school_year,
        "scope_label": scope_label,
        "start": start,
        "end": end,
        "requested_end": requested_end,
        "period_adjusted": bool(requested_end and requested_end > today),
        "generated_at": generated_at,
        "generated_by": _display_user(request.user),
        "report_reference": (
            f"RCE-{generated_at:%Y%m%d-%H%M%S}-{uuid.uuid4().hex[:6].upper()}"
        ),
    }


def _period_label(data):
    if data["start"]:
        return f"Du {data['start']:%d/%m/%Y} au {data['end']:%d/%m/%Y}"
    return f"Situation arrêtée au {data['end']:%d/%m/%Y}"


def _payments(data):
    queryset = (
        Paiement.objects.filter(eleve__classe_id__in=data["class_ids"])
        .select_related(
            "eleve", "eleve__classe", "type_paiement", "mode_paiement",
            "cree_par", "valide_par",
        )
        .order_by("date_paiement", "numero_recu", "pk")
    )
    if data["school_year"]:
        queryset = queryset.filter(annee_scolaire=data["school_year"])
    if data["start"]:
        queryset = queryset.filter(date_paiement__gte=data["start"])
    return queryset.filter(date_paiement__lte=data["end"])


def _empty_allocation():
    return {key: ZERO for key, _label in COMPONENTS}


def _allocations(data, selected):
    """Rejoue l'historique pour ventiler chaque encaissement sans forfait."""
    selected_ids = {item.pk for item in selected}
    if not selected_ids:
        return {}
    student_ids = {item.eleve_id for item in selected}
    schedules = EcheancierPaiement.objects.filter(eleve_id__in=student_ids)
    if data["school_year"]:
        schedules = schedules.filter(annee_scolaire=data["school_year"])
    schedule_by_key = {
        (item.eleve_id, item.annee_scolaire): item for item in schedules
    }
    history = (
        Paiement.objects.filter(
            eleve_id__in=student_ids,
            statut="VALIDE",
            date_paiement__lte=data["end"],
        )
        .filter(filtre_types_scolarite())
        .select_related("type_paiement")
        .order_by("eleve_id", "annee_scolaire", "date_paiement", "date_creation", "pk")
    )
    if data["school_year"]:
        history = history.filter(annee_scolaire=data["school_year"])

    balances = {}
    result = {}
    for payment in history:
        key = (payment.eleve_id, payment.annee_scolaire)
        schedule = schedule_by_key.get(key)
        allocation = _empty_allocation()
        if not schedule:
            allocation["non_affecte"] = payment.montant or ZERO
        else:
            current = balances.setdefault(key, due_balances(schedule))
            raw, remaining, unapplied = allocate_amount_sequentially(
                payment.montant, current
            )
            balances[key] = remaining
            admission_key = (
                "reinscription"
                if schedule.nature_frais == "REINSCRIPTION"
                else "inscription"
            )
            allocation[admission_key] = raw["inscription"]
            for tranche in ("tranche_1", "tranche_2", "tranche_3"):
                allocation[tranche] = raw[tranche]
            allocation["non_affecte"] = unapplied
        if payment.pk in selected_ids:
            result[payment.pk] = allocation

    for payment in selected:
        if payment.pk in result:
            continue
        allocation = _empty_allocation()
        if est_type_scolarite(payment.type_paiement):
            allocation["non_affecte"] = payment.montant or ZERO
        else:
            allocation["autres"] = payment.montant or ZERO
        result[payment.pk] = allocation
    return result


def _student_situations(data):
    """Situation de chaque élève à la date d'arrêté du rapport."""
    schedules = EcheancierPaiement.objects.filter(
        eleve__classe_id__in=data["class_ids"],
    ).select_related("eleve", "eleve__classe")
    if data["school_year"]:
        schedules = schedules.filter(annee_scolaire=data["school_year"])
    schedules = list(schedules.order_by(
        "eleve__classe__nom", "eleve__nom", "eleve__prenom", "annee_scolaire",
    ))
    situations = calculer_situations_echeanciers(
        schedules,
        date_reference=data["end"],
        date_limite=data["end"],
    )
    rows = []
    by_key = {}
    for schedule in schedules:
        situation = situations[schedule.pk]
        tuition_due = sum((
            Decimal(str(schedule.tranche_1_due or 0)),
            Decimal(str(schedule.tranche_2_due or 0)),
            Decimal(str(schedule.tranche_3_due or 0)),
        ), ZERO)
        total_due = situation["total_du"]
        paid = situation["encaisse"]
        discount = situation["remises"]
        remaining = situation["reste"]
        coverage = max(ZERO, total_due - remaining)
        row = {
            "matricule": schedule.eleve.matricule,
            "student": schedule.eleve.nom_complet,
            "class": schedule.eleve.classe.nom,
            "school_year": schedule.annee_scolaire,
            "total_due": total_due,
            "tuition_due": tuition_due,
            "paid": paid,
            "discount": discount,
            "discount_rate": _percentage(discount, tuition_due),
            "coverage": coverage,
            "remaining": remaining,
            "situation": _settlement_label(total_due, coverage, discount),
        }
        rows.append(row)
        by_key[(schedule.eleve_id, schedule.annee_scolaire)] = row
    return rows, by_key


def collect_accounting_data(request):
    data = _scope(request)
    payments = list(_payments(data))
    validated = [item for item in payments if item.statut == "VALIDE"]
    allocations = _allocations(data, validated)
    student_rows, student_by_key = _student_situations(data)
    discounts = list(
        PaiementRemise.objects.filter(paiement__in=validated)
        .select_related("paiement", "remise")
    )
    discount_by_payment = defaultdict(lambda: ZERO)
    discount_by_reason = defaultdict(lambda: ZERO)
    for discount in discounts:
        amount = discount.montant_remise or ZERO
        discount_by_payment[discount.paiement_id] += amount
        discount_by_reason[discount.remise.get_motif_display()] += amount

    status_labels = dict(Paiement.STATUT_CHOICES)
    by_status = {}
    for code, label in Paiement.STATUT_CHOICES:
        rows = [item for item in payments if item.statut == code]
        by_status[code] = {
            "label": label,
            "count": len(rows),
            "amount": sum((item.montant or ZERO for item in rows), ZERO),
        }

    by_mode = defaultdict(lambda: {
        "count": 0, "amount": ZERO, "reference_required": 0,
        "reference_missing": 0, "reference_missing_amount": ZERO,
    })
    by_type = defaultdict(lambda: {"count": 0, "amount": ZERO})
    by_class = defaultdict(lambda: {
        "count": 0, "amount": ZERO, "discount": ZERO, "tuition_due": ZERO,
    })
    by_component = {
        key: {"label": label, "count": 0, "amount": ZERO}
        for key, label in COMPONENTS
    }
    rows = []
    for payment in validated:
        amount = payment.montant or ZERO
        discount = discount_by_payment[payment.pk]
        mode = payment.mode_paiement.nom if payment.mode_paiement_id else "Non précisé"
        payment_type = payment.type_paiement.nom if payment.type_paiement_id else "Non précisé"
        class_name = payment.eleve.classe.nom
        external_reference_required = _requires_reference(mode)
        has_reference = bool((payment.reference_externe or "").strip())
        mode_row = by_mode[mode]
        mode_row["count"] += 1
        mode_row["amount"] += amount
        if external_reference_required:
            mode_row["reference_required"] += 1
            if not has_reference:
                mode_row["reference_missing"] += 1
                mode_row["reference_missing_amount"] += amount
        by_type[payment_type]["count"] += 1
        by_type[payment_type]["amount"] += amount
        by_class[class_name]["count"] += 1
        by_class[class_name]["amount"] += amount
        by_class[class_name]["discount"] += discount
        allocation = allocations[payment.pk]
        student_situation = student_by_key.get(
            (payment.eleve_id, payment.annee_scolaire)
        )
        discount_rate = _percentage(
            discount,
            student_situation["tuition_due"] if student_situation else amount + discount,
        )
        for component, component_amount in allocation.items():
            if component_amount:
                by_component[component]["count"] += 1
                by_component[component]["amount"] += component_amount
        rows.append({
            "date": payment.date_paiement,
            "receipt": payment.numero_recu,
            "matricule": payment.eleve.matricule,
            "student": payment.eleve.nom_complet,
            "class": class_name,
            "type": payment_type,
            "mode": mode,
            "amount": amount,
            "discount": discount,
            "discount_rate": discount_rate,
            "situation": (
                student_situation["situation"]
                if student_situation and est_type_scolarite(payment.type_paiement)
                else (
                    "Autre service"
                    if not est_type_scolarite(payment.type_paiement)
                    else "Échéancier absent"
                )
            ),
            "reference": payment.reference_externe or "-",
            "reference_status": (
                "Non requise" if not external_reference_required
                else ("Complète" if has_reference else "À compléter")
            ),
            "cashier": _display_user(payment.cree_par),
            "validator": _display_user(payment.valide_par),
            "allocation": allocation,
        })

    for student_row in student_rows:
        by_class[student_row["class"]]["tuition_due"] += student_row["tuition_due"]
    for class_row in by_class.values():
        class_row["discount_rate"] = _percentage(
            class_row["discount"], class_row["tuition_due"]
        )

    total_validated = sum((item.montant or ZERO for item in validated), ZERO)
    total_discounts = sum(discount_by_payment.values(), ZERO)
    data.update({
        "period_label": _period_label(data),
        "payments": payments,
        "payment_rows": rows,
        "payment_count": len(payments),
        "validated_count": len(validated),
        "total_validated": total_validated,
        "total_discounts": total_discounts,
        "total_coverage": total_validated + total_discounts,
        "student_rows": student_rows,
        "total_tuition_due": sum(
            (item["tuition_due"] for item in student_rows), ZERO
        ),
        "by_status": by_status,
        "by_mode": dict(sorted(by_mode.items())),
        "by_type": dict(sorted(by_type.items())),
        "by_class": dict(sorted(by_class.items())),
        "by_component": by_component,
        "discount_by_reason": dict(sorted(discount_by_reason.items())),
        "reference_missing_count": sum(item["reference_missing"] for item in by_mode.values()),
        "reference_missing_amount": sum(
            (item["reference_missing_amount"] for item in by_mode.values()), ZERO
        ),
        "unallocated_total": by_component["non_affecte"]["amount"],
    })
    return data


def _pdf_table(rows, widths=None, numeric_columns=(), total_row=False):
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, Table, TableStyle

    body = getSampleStyleSheet()["BodyText"]
    body.fontSize = 6.5
    body.leading = 7.5
    prepared = []
    for row_index, row in enumerate(rows):
        prepared.append([
            value if row_index == 0 or column in numeric_columns
            else Paragraph(escape(str(value or "-")), body)
            for column, value in enumerate(row)
        ])
    style = [
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + BLUE)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 6.5),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#B7C4CC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7F9")]),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
    ]
    for column in numeric_columns:
        style.append(("ALIGN", (column, 1), (column, -1), "RIGHT"))
    if total_row and len(rows) > 1:
        style.extend([
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#" + LIGHT_BLUE)),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ])
    return Table(prepared, repeatRows=1, colWidths=widths, style=TableStyle(style))


def build_accounting_pdf(data):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.platypus import (
        KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    buffer = BytesIO()
    page_size = landscape(A4)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], fontSize=15, leading=18,
        textColor=colors.HexColor("#" + BLUE), spaceAfter=3,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle", parent=styles["Normal"], fontSize=8,
        textColor=colors.HexColor("#" + GREY), spaceAfter=6,
    )
    section_style = ParagraphStyle(
        "Section", parent=styles["Heading2"], fontSize=9.5, leading=11,
        textColor=colors.HexColor("#" + BLUE), spaceBefore=7, spaceAfter=3,
    )
    subsection_style = ParagraphStyle(
        "Subsection", parent=styles["Heading3"], fontSize=8, leading=9,
        textColor=colors.HexColor("#" + BLUE), spaceBefore=3, spaceAfter=2,
    )
    note_style = ParagraphStyle(
        "Note", parent=styles["Normal"], fontSize=7, leading=9,
        backColor=colors.HexColor("#FFF5E6"), borderPadding=4, spaceAfter=4,
    )
    document_options = {
        "pagesize": page_size,
        "leftMargin": 0.8 * cm,
        "rightMargin": 0.8 * cm,
        "topMargin": 1.55 * cm,
        "bottomMargin": 1.05 * cm,
        "title": "Rapport comptable des encaissements",
        "author": data["generated_by"],
    }
    document = SimpleDocTemplate(buffer, **document_options)

    logo_path = _get_logo_path(data.get("school"))
    report_title = "RAPPORT COMPTABLE DES ENCAISSEMENTS"

    def draw_chrome(pdf, page_number, page_count):
        width, height = page_size
        pdf.saveState()
        pdf.resetTransforms()
        if logo_path:
            # Filigrane de l'ecole, place derriere le contenu de chaque page.
            # Une opacite faible conserve la lisibilite des tableaux comptables.
            pdf.saveState()
            try:
                pdf.setFillAlpha(0.08)
                watermark_width = width * 1.10
                watermark_height = height * 0.90
                pdf.translate(width / 2.0, height / 2.0)
                pdf.rotate(22)
                pdf.translate(-width / 2.0, -height / 2.0)
                pdf.drawImage(
                    logo_path,
                    (width - watermark_width) / 2.0,
                    (height - watermark_height) / 2.0,
                    watermark_width,
                    watermark_height,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass
            finally:
                pdf.restoreState()

            try:
                pdf.drawImage(logo_path, 0.8 * cm, height - 1.25 * cm, 1.3 * cm, 0.7 * cm,
                              preserveAspectRatio=True, mask="auto")
            except Exception:
                pass
        pdf.setFillColor(colors.HexColor("#" + BLUE))
        pdf.setFont("Helvetica-Bold", 8)
        pdf.drawString(2.25 * cm, height - 0.88 * cm, data["school_name"])
        pdf.setFont("Helvetica", 6.5)
        pdf.setFillColor(colors.HexColor("#" + GREY))
        pdf.drawRightString(width - 0.8 * cm, height - 0.88 * cm, report_title)
        pdf.line(0.8 * cm, height - 1.3 * cm, width - 0.8 * cm, height - 1.3 * cm)
        pdf.line(0.8 * cm, 0.78 * cm, width - 0.8 * cm, 0.78 * cm)
        pdf.drawString(
            0.8 * cm, 0.45 * cm,
            f"Confidentiel - Réf. {data['report_reference']} - {data['generated_by']} - "
            f"{data['generated_at']:%d/%m/%Y à %H:%M}",
        )
        pdf.drawRightString(width - 0.8 * cm, 0.45 * cm, f"Page {page_number}/{page_count}")
        pdf.restoreState()

    elements = [
        Paragraph(report_title, title_style),
        Paragraph(
            escape(
                f"{data['scope_label']} | Année scolaire : {data['school_year'] or 'Toutes'} | "
                f"{data['period_label']} | Réf. {data['report_reference']}"
            ),
            subtitle_style,
        ),
    ]
    if data["period_adjusted"]:
        elements.append(Paragraph(
            escape(
                f"La date demandée ({data['requested_end']:%d/%m/%Y}) est future : "
                f"la situation a été automatiquement arrêtée au {data['end']:%d/%m/%Y}."
            ), note_style,
        ))

    kpis = [
        ["Encaissements validés", "Remises accordées", "Couverture", "En attente", "Remboursés", "Réf. manquantes"],
        [_money(data["total_validated"]), _money(data["total_discounts"]),
         _money(data["total_coverage"]), data["by_status"]["EN_ATTENTE"]["count"],
         data["by_status"]["REMBOURSE"]["count"], data["reference_missing_count"]],
    ]
    elements.append(Table(kpis, colWidths=[4.4 * cm] * 6, style=TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + BLUE)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#" + LIGHT_BLUE)),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B7C4CC")),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
    ])))

    elements.append(Paragraph("1. Synthèse par statut", section_style))
    status_rows = [["Statut", "Opérations", "Montant (GNF)"]]
    for code, _label in Paiement.STATUT_CHOICES:
        item = data["by_status"][code]
        status_rows.append([item["label"], item["count"], _money(item["amount"])])
    status_rows.append(["TOTAL", data["payment_count"], _money(sum(
        (item.montant or ZERO for item in data["payments"]), ZERO
    ))])
    elements.append(_pdf_table(status_rows, [9 * cm, 4 * cm, 6 * cm], (1, 2), True))

    elements.append(Paragraph("2. Affectation réelle des encaissements", section_style))
    allocation_rows = [["Affectation", "Opérations", "Montant (GNF)"]]
    for item in data["by_component"].values():
        allocation_rows.append([item["label"], item["count"], _money(item["amount"])])
    allocation_rows.append(["TOTAL", "-", _money(data["total_validated"])])
    elements.append(_pdf_table(allocation_rows, [10 * cm, 4 * cm, 6 * cm], (1, 2), True))
    if data["unallocated_total"]:
        elements.append(Paragraph(
            f"Alerte : {_money(data['unallocated_total'])} GNF restent non affectés et doivent être contrôlés.",
            note_style,
        ))

    elements.append(Paragraph("3. Contrôle par mode et justificatifs", section_style))
    mode_rows = [["Mode", "Opérations", "Montant", "Réf. requises", "Réf. manquantes", "Montant à justifier"]]
    for label, item in data["by_mode"].items():
        mode_rows.append([
            label, item["count"], _money(item["amount"]), item["reference_required"],
            item["reference_missing"], _money(item["reference_missing_amount"]),
        ])
    if len(mode_rows) == 1:
        mode_rows.append(["Aucun encaissement validé", 0, 0, 0, 0, 0])
    elements.append(_pdf_table(mode_rows, [5 * cm, 3 * cm, 4 * cm, 3 * cm, 3.5 * cm, 4.5 * cm], (1, 2, 3, 4, 5)))
    elements.append(Paragraph(
        "Ce contrôle vérifie les références enregistrées. Le rapprochement définitif doit être fait avec la caisse, les relevés Mobile Money, les chèques et les relevés bancaires.",
        note_style,
    ))

    elements.append(Paragraph("4. Ventilation par type et par classe", section_style))
    type_rows = [["Type", "Opérations", "Montant"]] + [
        [label, item["count"], _money(item["amount"])]
        for label, item in data["by_type"].items()
    ]
    class_rows = [["Classe", "Opérations", "Encaissé", "Remises", "Remise %", "Couverture"]] + [
        [label, item["count"], _money(item["amount"]), _money(item["discount"]),
         f"{item['discount_rate']:.1f} %", _money(item["amount"] + item["discount"])]
        for label, item in data["by_class"].items()
    ]
    # Keep these tables independent: the class table may contain many rows and
    # must be allowed to split across pages without overflowing the A4 frame.
    elements.append(KeepTogether([
        Paragraph("4.1. Par type de paiement", subsection_style),
        _pdf_table(
            type_rows or [["Type", "Opérations", "Montant"]],
            [10 * cm, 5 * cm, 7 * cm],
            (1, 2),
        ),
    ]))
    elements.append(Paragraph("4.2. Par classe", subsection_style))
    elements.append(_pdf_table(
        class_rows or [["Classe", "Opérations", "Encaissé", "Remises", "Remise %", "Couverture"]],
        [6.5 * cm, 3.2 * cm, 4.5 * cm, 4.2 * cm, 3 * cm, 4.5 * cm],
        (1, 2, 3, 4, 5),
    ))

    elements.append(Paragraph("5. Remises et réductions", section_style))
    discount_rows = [["Motif", "Montant (GNF)", "% de la scolarité"]] + [
        [label, _money(amount), f"{_percentage(amount, data['total_tuition_due']):.1f} %"]
        for label, amount in data["discount_by_reason"].items()
    ]
    if len(discount_rows) == 1:
        discount_rows.append(["Aucune remise", "0", "0.0 %"])
    discount_rows.append([
        "TOTAL", _money(data["total_discounts"]),
        f"{_percentage(data['total_discounts'], data['total_tuition_due']):.1f} %",
    ])
    elements.append(_pdf_table(
        discount_rows, [11 * cm, 6 * cm, 5 * cm], (1, 2), True
    ))

    elements.append(Paragraph("6. Situation des élèves et remises appliquées", section_style))
    student_details = [[
        "Élève", "Classe", "Total dû", "Base scolarité", "Encaissé",
        "Remise", "Remise %", "Couverture", "Reste", "Situation",
    ]]
    for item in data["student_rows"]:
        student_details.append([
            f"{item['matricule']}\n{item['student']}", item["class"],
            _money(item["total_due"]), _money(item["tuition_due"]),
            _money(item["paid"]), _money(item["discount"]),
            f"{item['discount_rate']:.1f} %", _money(item["coverage"]),
            _money(item["remaining"]), item["situation"],
        ])
    if len(student_details) == 1:
        student_details.append([
            "Aucun échéancier", "-", "0", "0", "0", "0", "0.0 %", "0", "0", "-",
        ])
    elements.append(_pdf_table(
        student_details,
        [4 * cm, 2.5 * cm, 2.3 * cm, 2.5 * cm, 2.3 * cm, 2.2 * cm,
         1.5 * cm, 2.3 * cm, 2.1 * cm, 3.3 * cm],
        (2, 3, 4, 5, 6, 7, 8),
    ))

    if data["payment_rows"]:
        elements.append(Paragraph("7. Journal détaillé des encaissements validés", section_style))
        details = [[
            "Date", "Reçu", "Élève", "Classe", "Type", "Mode", "Montant",
            "Remise", "Remise %", "Situation", "Référence", "Validation",
        ]]
        for item in data["payment_rows"]:
            details.append([
                item["date"].strftime("%d/%m/%Y"), item["receipt"],
                f"{item['matricule']}\n{item['student']}", item["class"], item["type"],
                item["mode"], _money(item["amount"]), _money(item["discount"]),
                f"{item['discount_rate']:.1f} %", item["situation"],
                f"{item['reference']}\n{item['reference_status']}",
                f"Caisse : {item['cashier']}\nVisa : {item['validator']}",
            ])
        elements.append(_pdf_table(
            details,
            [1.5 * cm, 1.8 * cm, 3.4 * cm, 2.2 * cm, 2.8 * cm, 2.1 * cm,
             2.2 * cm, 1.9 * cm, 1.4 * cm, 2.5 * cm, 2.5 * cm, 3 * cm],
            (6, 7, 8),
        ))
    else:
        elements.append(Paragraph(
            "Aucun encaissement validé pour les filtres sélectionnés.", note_style
        ))

    elements.extend([
        Spacer(1, 0.25 * cm),
        _pdf_table(
            [["Établi par", "Contrôlé par", "Validé par"],
             [data["generated_by"], "Nom / Signature", "Direction / Signature"]],
            [8.7 * cm, 8.7 * cm, 8.7 * cm],
        ),
    ])
    counted_pages = []

    class CountingCanvas(pdf_canvas.Canvas):
        def showPage(self):
            counted_pages.append(self.getPageNumber())
            super().showPage()

    # First pass: determine the final page count using the exact same layout.
    probe_document = SimpleDocTemplate(BytesIO(), **document_options)
    probe_document.build(deepcopy(elements), canvasmaker=CountingCanvas)
    page_count = len(counted_pages) or 1

    def draw_page(pdf, doc):
        # Page templates run outside any clipping path created by large tables.
        draw_chrome(pdf, doc.page, page_count)

    # Second pass: use the known total in each page template.
    document.build(elements, onFirstPage=draw_page, onLaterPages=draw_page)
    buffer.seek(0)
    return buffer


def _style_sheet(ws, headers_row, widths):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    border_side = Side(style="thin", color="B7C4CC")
    for cell in ws[headers_row]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    for column, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.freeze_panes = f"A{headers_row + 1}"
    ws.auto_filter.ref = ws.dimensions


def build_accounting_workbook(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    workbook.remove(workbook.active)

    summary = workbook.create_sheet("Synthèse")
    summary.append([data["school_name"], "Rapport comptable des encaissements"])
    summary.append(["Référence", data["report_reference"]])
    summary.append(["Portée", data["scope_label"]])
    summary.append(["Période", data["period_label"]])
    summary.append(["Année scolaire", data["school_year"] or "Toutes"])
    summary.append(["Généré par", data["generated_by"]])
    summary.append(["Généré le", data["generated_at"].strftime("%d/%m/%Y à %H:%M")])
    summary.append([])
    summary.append(["Indicateur", "Valeur"])
    summary.append(["Encaissements validés", int(data["total_validated"])])
    summary.append(["Remises accordées", int(data["total_discounts"])])
    summary.append(["Couverture totale", int(data["total_coverage"])])
    summary.append(["Références manquantes", data["reference_missing_count"]])
    summary.append(["Montant à justifier", int(data["reference_missing_amount"])])
    _style_sheet(summary, 9, [36, 34])
    summary["A1"].font = Font(size=14, bold=True, color=BLUE)

    journal = workbook.create_sheet("Journal validé")
    journal_headers = [
        "Date", "Reçu", "Matricule", "Élève", "Classe", "Type", "Mode",
        "Montant", "Remise", "Remise (%)", "Situation", "Référence",
        "État référence", "Caissier", "Validateur", "Inscription", "Réinscription", "Tranche 1",
        "Tranche 2", "Tranche 3", "Autres services", "Non affecté",
    ]
    journal.append(journal_headers)
    for item in data["payment_rows"]:
        allocation = item["allocation"]
        journal.append([
            item["date"], _excel_safe(item["receipt"]), _excel_safe(item["matricule"]),
            _excel_safe(item["student"]), _excel_safe(item["class"]),
            _excel_safe(item["type"]), _excel_safe(item["mode"]), int(item["amount"]),
            int(item["discount"]), float(item["discount_rate"] / Decimal("100")),
            item["situation"], _excel_safe(item["reference"]), item["reference_status"],
            _excel_safe(item["cashier"]), _excel_safe(item["validator"]),
            int(allocation["inscription"]), int(allocation["reinscription"]),
            int(allocation["tranche_1"]), int(allocation["tranche_2"]),
            int(allocation["tranche_3"]), int(allocation["autres"]),
            int(allocation["non_affecte"]),
        ])
    _style_sheet(
        journal, 1,
        [12, 18, 16, 26, 18, 25, 18, 15, 15, 13, 28, 20, 17, 20, 20] + [15] * 7,
    )
    for cell in journal["J"][1:]:
        cell.number_format = "0.0%"

    students = workbook.create_sheet("Situation élèves")
    students.append([
        "Matricule", "Élève", "Classe", "Année scolaire", "Total dû",
        "Base scolarité", "Encaissé", "Remise", "Remise (%)",
        "Couverture", "Reste", "Situation",
    ])
    for item in data["student_rows"]:
        students.append([
            _excel_safe(item["matricule"]), _excel_safe(item["student"]),
            _excel_safe(item["class"]), item["school_year"], int(item["total_due"]),
            int(item["tuition_due"]), int(item["paid"]), int(item["discount"]),
            float(item["discount_rate"] / Decimal("100")), int(item["coverage"]),
            int(item["remaining"]), item["situation"],
        ])
    _style_sheet(students, 1, [16, 28, 20, 16] + [16] * 7 + [30])
    for cell in students["I"][1:]:
        cell.number_format = "0.0%"

    affectations = workbook.create_sheet("Affectations")
    affectations.append(["Affectation réelle", "Opérations", "Montant (GNF)"])
    for item in data["by_component"].values():
        affectations.append([item["label"], item["count"], int(item["amount"])])
    affectations.append(["TOTAL", "", int(data["total_validated"])])
    _style_sheet(affectations, 1, [32, 15, 20])

    statuses = workbook.create_sheet("Statuts")
    statuses.append(["Statut", "Opérations", "Montant (GNF)"])
    for code, _label in Paiement.STATUT_CHOICES:
        item = data["by_status"][code]
        statuses.append([item["label"], item["count"], int(item["amount"])])
    _style_sheet(statuses, 1, [24, 15, 20])

    ventilations = workbook.create_sheet("Ventilations")
    ventilations.append(["PAR MODE", "Opérations", "Montant", "Réf. requises", "Réf. manquantes", "Montant à justifier"])
    for label, item in data["by_mode"].items():
        ventilations.append([label, item["count"], int(item["amount"]), item["reference_required"], item["reference_missing"], int(item["reference_missing_amount"])])
    ventilations.append([])
    ventilations.append(["PAR TYPE", "Opérations", "Montant"])
    for label, item in data["by_type"].items():
        ventilations.append([label, item["count"], int(item["amount"])])
    ventilations.append([])
    ventilations.append(["PAR CLASSE", "Opérations", "Encaissé", "Remises", "Remise (%)", "Couverture"])
    class_section_start = ventilations.max_row + 1
    for label, item in data["by_class"].items():
        ventilations.append([
            label, item["count"], int(item["amount"]), int(item["discount"]),
            float(item["discount_rate"] / Decimal("100")),
            int(item["amount"] + item["discount"]),
        ])
    _style_sheet(ventilations, 1, [32, 15, 20, 18, 16, 22])
    for cell in ventilations["E"][class_section_start - 1:]:
        cell.number_format = "0.0%"

    discounts = workbook.create_sheet("Remises")
    discounts.append(["Motif", "Montant (GNF)", "% de la scolarité"])
    for label, amount in data["discount_by_reason"].items():
        discounts.append([
            label, int(amount),
            float(_percentage(amount, data["total_tuition_due"]) / Decimal("100")),
        ])
    discounts.append([
        "TOTAL", int(data["total_discounts"]),
        float(_percentage(
            data["total_discounts"], data["total_tuition_due"]
        ) / Decimal("100")),
    ])
    _style_sheet(discounts, 1, [35, 22, 20])
    for cell in discounts["C"][1:]:
        cell.number_format = "0.0%"
    return workbook


def collect_payment_modes_data(request):
    """Agrège les montants par mode avec le même périmètre que les rapports."""
    data = _scope(request)
    status = (request.GET.get("statut") or "VALIDE").strip().upper()
    labels = dict(Paiement.STATUT_CHOICES)
    if status not in labels and status != "TOUS":
        status = "VALIDE"

    payments = _payments(data)
    if status != "TOUS":
        payments = payments.filter(statut=status)

    grouped = (
        payments.order_by()
        .values("mode_paiement__nom")
        .annotate(count=Count("id"), amount=Sum("montant"))
        .order_by("-amount", "mode_paiement__nom")
    )
    modes = []
    total_amount = ZERO
    payment_count = 0
    for item in grouped:
        amount = item["amount"] or ZERO
        count = item["count"] or 0
        total_amount += amount
        payment_count += count
        modes.append({
            "name": item["mode_paiement__nom"] or "Non précisé",
            "count": count,
            "amount": amount,
        })

    for item in modes:
        item["share"] = (
            item["amount"] / total_amount if total_amount else ZERO
        )
        item["average"] = (
            item["amount"] / item["count"] if item["count"] else ZERO
        )

    data.update({
        "status": status,
        "status_label": "Tous les statuts" if status == "TOUS" else labels[status],
        "modes": modes,
        "total_amount": total_amount,
        "payment_count": payment_count,
    })
    return data


def build_payment_modes_pdf(data):
    """Construit l'extrait PDF des montants par mode d'encaissement."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = BytesIO()
    page_size = A4
    document = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.8 * cm,
        bottomMargin=1.2 * cm,
        title="Montants par mode d'encaissement",
        author=data["generated_by"],
    )
    logo_path = _get_logo_path(data.get("school"))
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ModesTitle",
        parent=styles["Title"],
        fontSize=15,
        leading=18,
        textColor=colors.HexColor("#" + BLUE),
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "ModesSubtitle",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#" + GREY),
        spaceAfter=10,
    )

    def draw_page(pdf, doc):
        width, height = page_size
        pdf.saveState()
        if logo_path:
            pdf.saveState()
            try:
                pdf.setFillAlpha(0.07)
                watermark_width = width * 1.15
                watermark_height = height * 0.65
                pdf.translate(width / 2.0, height / 2.0)
                pdf.rotate(22)
                pdf.translate(-width / 2.0, -height / 2.0)
                pdf.drawImage(
                    logo_path,
                    (width - watermark_width) / 2.0,
                    (height - watermark_height) / 2.0,
                    watermark_width,
                    watermark_height,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass
            finally:
                pdf.restoreState()
            try:
                pdf.drawImage(
                    logo_path, 1.2 * cm, height - 1.35 * cm, 1.3 * cm, 0.75 * cm,
                    preserveAspectRatio=True, mask="auto",
                )
            except Exception:
                pass
        pdf.setFont("Helvetica-Bold", 8)
        pdf.setFillColor(colors.HexColor("#" + BLUE))
        pdf.drawString(2.7 * cm, height - 0.92 * cm, data["school_name"])
        pdf.setFont("Helvetica", 6.5)
        pdf.setFillColor(colors.HexColor("#" + GREY))
        pdf.drawRightString(
            width - 1.2 * cm, height - 0.92 * cm,
            "MONTANTS PAR MODE D'ENCAISSEMENT",
        )
        pdf.line(1.2 * cm, height - 1.4 * cm, width - 1.2 * cm, height - 1.4 * cm)
        pdf.line(1.2 * cm, 0.8 * cm, width - 1.2 * cm, 0.8 * cm)
        pdf.drawString(
            1.2 * cm, 0.48 * cm,
            f"Réf. {data['report_reference']} - {data['generated_by']} - "
            f"{data['generated_at']:%d/%m/%Y à %H:%M}",
        )
        pdf.drawRightString(width - 1.2 * cm, 0.48 * cm, f"Page {doc.page}")
        pdf.restoreState()

    elements = [
        Paragraph("MONTANTS PAR MODE D'ENCAISSEMENT", title_style),
        Paragraph(
            escape(
                f"{data['scope_label']} | Année scolaire : "
                f"{data['school_year'] or 'Toutes'} | {_period_label(data)} | "
                f"Statut : {data['status_label']}"
            ),
            subtitle_style,
        ),
    ]
    summary = Table(
        [
            ["Montant total", "Nombre d'encaissements", "Modes utilisés"],
            [_money(data["total_amount"]), data["payment_count"], len(data["modes"])],
        ],
        colWidths=[6.2 * cm, 6.2 * cm, 6.2 * cm],
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + BLUE)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#" + LIGHT_BLUE)),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B7C4CC")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]),
    )
    elements.extend([summary, Spacer(1, 0.45 * cm)])

    rows = [["Mode", "Opérations", "Montant (GNF)", "Part du total", "Montant moyen"]]
    for item in data["modes"]:
        rows.append([
            item["name"], item["count"], _money(item["amount"]),
            f"{item['share'] * Decimal('100'):.1f} %", _money(item["average"]),
        ])
    if not data["modes"]:
        rows.append(["Aucun encaissement", 0, "0", "0,0 %", "0"])
    rows.append([
        "TOTAL", data["payment_count"], _money(data["total_amount"]),
        "100,0 %" if data["total_amount"] else "0,0 %",
        _money(
            data["total_amount"] / data["payment_count"]
            if data["payment_count"] else ZERO
        ),
    ])
    elements.append(
        _pdf_table(
            rows,
            [5.2 * cm, 3.1 * cm, 4.1 * cm, 3.2 * cm, 3.7 * cm],
            (1, 2, 3, 4),
            True,
        )
    )
    elements.extend([
        Spacer(1, 0.4 * cm),
        Paragraph(
            "Les montants correspondent aux paiements du statut et de la période "
            "sélectionnés, dans le périmètre autorisé de l'utilisateur.",
            subtitle_style,
        ),
    ])
    document.build(elements, onFirstPage=draw_page, onLaterPages=draw_page)
    buffer.seek(0)
    return buffer


def build_payment_modes_workbook(data):
    """Construit le classeur Excel auditable des montants par mode."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Modes d'encaissement"
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:E1")
    sheet["A1"] = "MONTANTS PAR MODE D'ENCAISSEMENT"
    sheet["A1"].font = Font(size=15, bold=True, color=BLUE)
    sheet["A1"].alignment = Alignment(horizontal="center")
    metadata = [
        ("Établissement", data["school_name"]),
        ("Portée", data["scope_label"]),
        ("Période", _period_label(data)),
        ("Année scolaire", data["school_year"] or "Toutes"),
        ("Statut", data["status_label"]),
        ("Généré par", data["generated_by"]),
    ]
    for row, (label, value) in enumerate(metadata, 2):
        sheet.cell(row, 1, label).font = Font(bold=True, color=GREY)
        sheet.cell(row, 2, _excel_safe(value))

    header_row = 9
    headers = [
        "Mode d'encaissement", "Nombre d'encaissements", "Montant (GNF)",
        "Part du total", "Montant moyen (GNF)",
    ]
    for column, value in enumerate(headers, 1):
        sheet.cell(header_row, column, value)

    first_data_row = header_row + 1
    source_modes = data["modes"] or [{"name": "Aucun encaissement", "count": 0, "amount": ZERO}]
    for item in source_modes:
        sheet.append([
            _excel_safe(item["name"]), item["count"], int(item["amount"]), None, None,
        ])
    last_data_row = sheet.max_row
    total_row = last_data_row + 1
    sheet.cell(total_row, 1, "TOTAL")
    sheet.cell(total_row, 2, f"=SUM(B{first_data_row}:B{last_data_row})")
    sheet.cell(total_row, 3, f"=SUM(C{first_data_row}:C{last_data_row})")
    sheet.cell(total_row, 4, f'=IF(C{total_row}=0,0,SUM(D{first_data_row}:D{last_data_row}))')
    sheet.cell(total_row, 5, f'=IF(B{total_row}=0,0,C{total_row}/B{total_row})')
    for row in range(first_data_row, last_data_row + 1):
        sheet.cell(row, 4, f'=IF($C${total_row}=0,0,C{row}/$C${total_row})')
        sheet.cell(row, 5, f'=IF(B{row}=0,0,C{row}/B{row})')

    header_fill = PatternFill("solid", fgColor=BLUE)
    total_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    thin = Side(style="thin", color="B7C4CC")
    for cell in sheet[header_row]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row in range(first_data_row, total_row + 1):
        for cell in sheet[row]:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(
                horizontal="left" if cell.column == 1 else "right",
                vertical="center",
            )
    for cell in sheet[total_row]:
        cell.fill = total_fill
        cell.font = Font(bold=True, color=BLUE)

    for row in range(first_data_row, total_row + 1):
        sheet.cell(row, 2).number_format = "#,##0"
        sheet.cell(row, 3).number_format = '#,##0 "GNF"'
        sheet.cell(row, 4).number_format = "0.0%"
        sheet.cell(row, 5).number_format = '#,##0 "GNF"'
    for column, width in {"A": 30, "B": 22, "C": 20, "D": 17, "E": 22}.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[1].height = 25
    sheet.row_dimensions[header_row].height = 32
    sheet.freeze_panes = f"A{first_data_row}"
    sheet.auto_filter.ref = f"A{header_row}:E{last_data_row}"
    sheet.print_title_rows = f"1:{header_row}"
    sheet.print_area = f"A1:E{total_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    return workbook


def _bad_request(exc):
    return HttpResponse(str(exc), status=400, content_type="text/plain; charset=utf-8")


def _filename(data, extension):
    scope = data["classes"][0].nom if len(data["classes"]) == 1 else data["school_name"]
    return f"rapport_comptable_{_safe_filename(scope)}_{data['end']:%Y-%m-%d}.{extension}"


def _modes_filename(data, extension):
    scope = data["classes"][0].nom if len(data["classes"]) == 1 else data["school_name"]
    return (
        f"montants_par_mode_{_safe_filename(scope)}_"
        f"{data['end']:%Y-%m-%d}.{extension}"
    )


def _accounting_pdf_response(request, disposition):
    try:
        data = collect_accounting_data(request)
    except ValueError as exc:
        return _bad_request(exc)
    response = HttpResponse(
        build_accounting_pdf(data).getvalue(),
        content_type="application/pdf",
    )
    response["Content-Disposition"] = (
        f'{disposition}; filename="{_filename(data, "pdf")}"'
    )
    response["Cache-Control"] = "private, no-store"
    response["X-Content-Type-Options"] = "nosniff"
    return response


@can_view_reports
def apercu_comptabilite_pdf(request):
    """Affiche dans le navigateur le PDF exact proposé au téléchargement."""
    return _accounting_pdf_response(request, "inline")


@can_view_reports
def export_comptabilite_pdf(request):
    return _accounting_pdf_response(request, "attachment")


@can_view_reports
def export_comptabilite_excel(request):
    try:
        data = collect_accounting_data(request)
    except ValueError as exc:
        return _bad_request(exc)
    stream = BytesIO()
    build_accounting_workbook(data).save(stream)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_filename(data, "xlsx")}"'
    return response


@can_view_reports
def export_payment_modes_pdf(request):
    try:
        data = collect_payment_modes_data(request)
    except ValueError as exc:
        return _bad_request(exc)
    response = HttpResponse(
        build_payment_modes_pdf(data).getvalue(),
        content_type="application/pdf",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{_modes_filename(data, "pdf")}"'
    )
    return response


@can_view_reports
def export_payment_modes_excel(request):
    try:
        data = collect_payment_modes_data(request)
    except ValueError as exc:
        return _bad_request(exc)
    stream = BytesIO()
    build_payment_modes_workbook(data).save(stream)
    response = HttpResponse(
        stream.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{_modes_filename(data, "xlsx")}"'
    )
    return response
