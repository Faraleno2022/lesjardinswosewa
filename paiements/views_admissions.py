"""Listes et exports des élèves inscrits et réinscrits."""

import io

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import F, Q, Sum
from django.http import Http404, HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas

from ecole_moderne.pdf_utils import draw_logo_watermark
from utilisateurs.utils import filter_by_user_school

from .models import EcheancierPaiement


NATURES = {
    'inscription': ('INSCRIPTION', 'Élèves inscrits'),
    'reinscription': ('REINSCRIPTION', 'Élèves réinscrits'),
}


def _nature(nature):
    try:
        return NATURES[nature]
    except KeyError as exc:
        raise Http404("Type d'admission inconnu") from exc


def _queryset(request, nature):
    code, _label = _nature(nature)
    queryset = (
        EcheancierPaiement.objects
        .select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
        .filter(
            nature_frais=code,
            annee_scolaire=F('eleve__classe__annee_scolaire'),
            eleve__est_dans_corbeille=False,
        )
    )
    queryset = filter_by_user_school(
        queryset, request.user, 'eleve__classe__ecole'
    )
    recherche = (request.GET.get('q') or '').strip()
    if recherche:
        queryset = queryset.filter(
            Q(eleve__matricule__icontains=recherche)
            | Q(eleve__nom__icontains=recherche)
            | Q(eleve__prenom__icontains=recherche)
            | Q(eleve__classe__nom__icontains=recherche)
        )
    return queryset.order_by('-date_creation', '-id')


@login_required
def liste_admissions(request, nature):
    _code, label = _nature(nature)
    queryset = _queryset(request, nature)
    totaux = queryset.aggregate(
        montant_du=Sum('frais_inscription_du'),
        montant_paye=Sum('frais_inscription_paye'),
    )
    page_obj = Paginator(queryset, 30).get_page(request.GET.get('page'))
    for echeancier in page_obj.object_list:
        echeancier.reste_admission = max(
            0,
            int(echeancier.frais_inscription_du or 0)
            - int(echeancier.frais_inscription_paye or 0),
        )
    return render(request, 'paiements/liste_admissions.html', {
        'titre_page': label,
        'nature': nature,
        'page_obj': page_obj,
        'q': (request.GET.get('q') or '').strip(),
        'montant_du': totaux['montant_du'] or 0,
        'montant_paye': totaux['montant_paye'] or 0,
    })


@login_required
def export_admissions_excel(request, nature):
    _code, label = _nature(nature)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = label[:31]
    headers = [
        'Matricule', 'Prénom', 'Nom', 'Sexe', 'Classe', 'École',
        'Année scolaire', "Date d'inscription", 'Frais dus (GNF)',
        'Frais payés (GNF)', 'Reste admission (GNF)', 'Statut échéancier',
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E78')
    for echeancier in _queryset(request, nature):
        eleve = echeancier.eleve
        montant_du = int(echeancier.frais_inscription_du or 0)
        montant_paye = int(echeancier.frais_inscription_paye or 0)
        sheet.append([
            eleve.matricule or '', eleve.prenom, eleve.nom,
            eleve.get_sexe_display(), eleve.classe.nom,
            eleve.classe.ecole.nom, echeancier.annee_scolaire,
            eleve.date_inscription.strftime('%d/%m/%Y') if eleve.date_inscription else '',
            montant_du, montant_paye, max(0, montant_du - montant_paye),
            echeancier.get_statut_display(),
        ])
    widths = [18, 20, 20, 12, 18, 28, 15, 18, 18, 18, 20, 20]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    stream = io.BytesIO()
    workbook.save(stream)
    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="eleves_{nature}.xlsx"'
    return response


@login_required
def export_admissions_pdf(request, nature):
    _code, label = _nature(nature)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="eleves_{nature}.pdf"'
    page_width, page_height = landscape(A4)
    pdf = canvas.Canvas(response, pagesize=(page_width, page_height))

    def header():
        draw_logo_watermark(pdf, page_width, page_height, opacity=0.035, scale=1.3)
        pdf.setFont('Helvetica-Bold', 17)
        pdf.drawCentredString(page_width / 2, page_height - 35, label.upper())
        pdf.setFont('Helvetica-Bold', 8)
        columns = [
            (30, 'N°'), (55, 'Matricule'), (145, 'Prénom et nom'),
            (300, 'Classe'), (390, 'École'), (560, 'Année'),
            (625, 'Dû'), (690, 'Payé'), (755, 'Reste'),
        ]
        for x, title in columns:
            pdf.drawString(x, page_height - 58, title)
        pdf.setStrokeColor(colors.HexColor('#1F4E78'))
        pdf.line(30, page_height - 63, page_width - 25, page_height - 63)
        return page_height - 78

    y = header()
    pdf.setFont('Helvetica', 8)
    for index, echeancier in enumerate(_queryset(request, nature), 1):
        if y < 35:
            pdf.showPage()
            y = header()
            pdf.setFont('Helvetica', 8)
        eleve = echeancier.eleve
        du = int(echeancier.frais_inscription_du or 0)
        paye = int(echeancier.frais_inscription_paye or 0)
        values = [
            (30, str(index)), (55, (eleve.matricule or '')[:16]),
            (145, eleve.nom_complet[:30]), (300, eleve.classe.nom[:17]),
            (390, eleve.classe.ecole.nom[:30]), (560, echeancier.annee_scolaire),
            (625, f'{du:,}'.replace(',', ' ')),
            (690, f'{paye:,}'.replace(',', ' ')),
            (755, f'{max(0, du - paye):,}'.replace(',', ' ')),
        ]
        for x, value in values:
            pdf.drawString(x, y, value)
        y -= 17
    pdf.save()
    return response
