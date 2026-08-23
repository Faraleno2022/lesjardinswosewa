from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse, HttpResponseForbidden
from datetime import datetime
from decimal import Decimal

from eleves.models import Classe, Ecole
from eleves.utils_annee import get_annee_active
from paiements.models import Paiement, PaiementRemise
from paiements.calculs import filtre_types_scolarite, normaliser_libelle
from utilisateurs.utils import user_is_admin, user_school
from rapports.utils import _draw_header_and_watermark

# ReportLab
# ReportLab: fera l'objet d'un import différé dans la vue PDF


TRANCHES_EXPORT_HEADERS = [
    'Élève',
    'Inscription payée',
    'Réinscription payée',
    'Tranche 1 payée',
    'Tranche 2 payée',
    'Tranche 3 payée',
    'Total dû',
    'Total payé',
    'Remise (GNF)',
    'Remise (%)',
    'Total couvert',
    'Reste',
    'Situation',
]


def _pourcentages_remise_selectionnes(eleve, annee_scolaire=None):
    """Retourne les taux réellement sélectionnés pour les remises validées.

    Le montant accordé peut être plafonné par le reste dû et ne permet donc
    pas de retrouver fidèlement le taux choisi par l'utilisateur. La source de
    vérité est ``RemiseReduction.valeur`` pour les remises de type pourcentage.
    Une remise fixe ne doit jamais produire artificiellement un pourcentage.
    """
    lignes = PaiementRemise.objects.filter(
        paiement__eleve=eleve,
        paiement__statut='VALIDE',
        montant_remise__gt=0,
        remise__type_remise='POURCENTAGE',
    ).filter(filtre_types_scolarite(prefix='paiement__type_paiement'))
    if annee_scolaire:
        lignes = lignes.filter(paiement__annee_scolaire=annee_scolaire)

    taux = {
        Decimal(str(valeur))
        for valeur in lignes.values_list('remise__valeur', flat=True)
        if valeur is not None
    }
    return tuple(sorted(taux))


def _pourcentages_pdf(taux):
    """Formate sans recalcul les taux sélectionnés pour le PDF."""
    if not taux:
        return ''
    return ' + '.join(
        f"{format(valeur.normalize(), 'f')} %" for valeur in taux
    )


def _pourcentages_excel(taux):
    """Conserve un vrai pourcentage Excel quand un seul taux est présent."""
    if not taux:
        return None
    if len(taux) == 1:
        return float(taux[0] / Decimal('100'))
    return ' + '.join(
        f"{format(valeur.normalize(), 'f')} %" for valeur in taux
    )


def _paiements_fallback_par_poste(eleve, annee_scolaire=None):
    """Totalise les reçus sans compter plusieurs fois un paiement combiné.

    Sans échéancier la ventilation d'un reçu combiné est indémontrable : son
    montant est conservé dans le total, mais pas inventé dans plusieurs
    colonnes de tranches.
    """
    paiements = Paiement.objects.filter(eleve=eleve, statut='VALIDE').filter(
        filtre_types_scolarite()
    )
    if annee_scolaire:
        paiements = paiements.filter(annee_scolaire=annee_scolaire)
    postes = {
        'inscription': Decimal('0'),
        'reinscription': Decimal('0'),
        'tranche_1': Decimal('0'),
        'tranche_2': Decimal('0'),
        'tranche_3': Decimal('0'),
    }
    total = Decimal('0')
    for paiement in paiements.select_related('type_paiement'):
        montant = Decimal(str(paiement.montant or 0))
        total += montant
        nom = normaliser_libelle(paiement.type_paiement.nom)
        cibles = []
        compact = ''.join(nom.split())
        if 'reinscription' in compact:
            cibles.append('reinscription')
        elif 'inscription' in nom:
            cibles.append('inscription')
        for numero in (1, 2, 3):
            if f'tranche {numero}' in nom or f'tranche{numero}' in nom:
                cibles.append(f'tranche_{numero}')
        if len(cibles) == 1:
            postes[cibles[0]] += montant
    remises = PaiementRemise.objects.filter(
        paiement__in=paiements,
    ).aggregate(total=Sum('montant_remise'))['total'] or Decimal('0')
    return postes, total, Decimal(str(remises))


def _donnees_tranches_eleve(eleve, annee_scolaire=None):
    """Retourne une ligne commune aux exports PDF et Excel.

    Les frais d'admission sont placés dans une seule des deux colonnes selon
    ``nature_frais``. Ils ne sont donc jamais comptés deux fois dans le total.
    """
    if annee_scolaire:
        echeancier = eleve.echeanciers.filter(
            annee_scolaire=annee_scolaire
        ).first()
    else:
        echeancier = getattr(eleve, 'echeancier', None)

    inscription = reinscription = Decimal('0')
    tranche_1 = tranche_2 = tranche_3 = Decimal('0')
    total_du = total_paye = remise = reste = Decimal('0')
    total_couvert = Decimal('0')
    pourcentages_remise = _pourcentages_remise_selectionnes(
        eleve, annee_scolaire
    )
    situation = 'Échéancier absent'

    if echeancier is not None:
        admission_payee = Decimal(str(echeancier.frais_inscription_paye or 0))
        if echeancier.nature_frais == 'REINSCRIPTION':
            reinscription = admission_payee
        else:
            inscription = admission_payee
        tranche_1 = Decimal(str(echeancier.tranche_1_payee or 0))
        tranche_2 = Decimal(str(echeancier.tranche_2_payee or 0))
        tranche_3 = Decimal(str(echeancier.tranche_3_payee or 0))
        total_du = Decimal(str(echeancier.total_du or 0))
        total_paye = inscription + reinscription + tranche_1 + tranche_2 + tranche_3
        remise = Decimal(str(echeancier.total_remises_valides or 0))
        total_couvert = min(total_du, total_paye + remise)
        reste = max(Decimal('0'), total_du - total_couvert)
        if total_du <= 0:
            situation = 'Échéancier vide'
        elif reste == 0:
            situation = 'Soldé - remise appliquée' if remise else 'Soldé'
        elif total_couvert > 0:
            situation = 'Partiel - remise appliquée' if remise else 'Partiel'
        else:
            situation = 'À payer'
    else:
        postes, total_paye, remise = _paiements_fallback_par_poste(eleve, annee_scolaire)
        inscription = postes['inscription']
        reinscription = postes['reinscription']
        tranche_1 = postes['tranche_1']
        tranche_2 = postes['tranche_2']
        tranche_3 = postes['tranche_3']
        total_couvert = total_paye + remise

    return {
        'inscription': inscription,
        'reinscription': reinscription,
        'tranche_1': tranche_1,
        'tranche_2': tranche_2,
        'tranche_3': tranche_3,
        'total_du': total_du,
        'total_paye': Decimal(str(total_paye or 0)),
        'remise': remise,
        'pourcentages_remise': pourcentages_remise,
        'total_couvert': total_couvert,
        'reste': reste,
        'situation': situation,
    }


@login_required
def export_tranches_par_classe_pdf(request):
    """Export PDF des tranches par classe avec logo entête et filigrane.

    Filtres GET:
    - ecole: id de l'école
    - classe: id de la classe
    - annee_scolaire: ex '2024-2025'

    Respecte la séparation par école pour les non-admins.
    """
    # Contrôle d'accès: Admin ou Comptable uniquement
    is_admin = user_is_admin(request.user)
    is_comptable = False
    try:
        if hasattr(request.user, 'profil'):
            is_comptable = (getattr(request.user.profil, 'role', None) == 'COMPTABLE')
    except Exception:
        is_comptable = False
    if not (is_admin or is_comptable):
        return HttpResponseForbidden("Accès refusé: vous n'avez pas l'autorisation d'exporter ce rapport.")

    # Lecture et validation des paramètres
    raw_ecole = (request.GET.get('ecole') or '').strip()
    raw_classe = (request.GET.get('classe') or request.GET.get('classe_id') or '').strip()
    annee_scolaire = (request.GET.get('annee_scolaire') or '').strip()

    def parse_int(value):
        try:
            return int(value)
        except Exception:
            return None

    ecole_id = parse_int(raw_ecole) if raw_ecole else None
    classe_id = parse_int(raw_classe) if raw_classe else None

    # Scope classes (filtrées par année active)
    classes = Classe.objects.select_related('ecole').all()
    ecole_user = user_school(request.user)
    annee_active = get_annee_active(request, ecole_user) if ecole_user else None
    restreindre = not user_is_admin(request.user) and ecole_user is not None
    if restreindre:
        classes = classes.filter(ecole=ecole_user)
    elif ecole_id:
        classes = classes.filter(ecole_id=ecole_id)
    if classe_id:
        classes = classes.filter(id=classe_id)
    if annee_active and not annee_scolaire:
        classes = classes.filter(annee_scolaire=annee_active)
    elif annee_scolaire:
        classes = classes.filter(annee_scolaire=annee_scolaire)

    # Anti-abus: limiter le nombre de classes exportées en une requête
    classes = list(classes.order_by('ecole__nom', 'niveau', 'nom')[:200])

    # L'en-tete doit recevoir l'ecole reelle du perimetre exporte. Sans ce
    # parametre, le helper utilisait seulement le logo statique par defaut.
    ecole_export = ecole_user if restreindre else None
    ecole_ids = {classe.ecole_id for classe in classes}
    if len(ecole_ids) == 1:
        ecole_export = classes[0].ecole
    elif not classes and ecole_id and not restreindre:
        ecole_export = Ecole.objects.filter(pk=ecole_id).first()

    # Préparer réponse PDF
    response = HttpResponse(content_type='application/pdf')
    suffix = datetime.now().strftime('%Y%m%d')
    response['Content-Disposition'] = f'attachment; filename="tranches_par_classe_{suffix}.pdf"'

    # Import différé de ReportLab pour éviter les erreurs si non installé
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab n'est pas installé. Veuillez exécuter: pip install reportlab", status=500)

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=20, leftMargin=20, topMargin=80, bottomMargin=30
    )
    elements = []
    styles = getSampleStyleSheet()
    cell = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=6.2, leading=7.2)
    header_cell = ParagraphStyle(
        'HeaderCell', parent=cell, fontName='Helvetica-Bold', textColor=colors.white,
        alignment=1,
    )

    titre = 'Tranches par classe'
    if annee_scolaire:
        titre += f" - Année {annee_scolaire}"
    elements.append(Paragraph(titre, styles['Title']))
    elements.append(Spacer(1, 0.5*cm))

    header = TRANCHES_EXPORT_HEADERS

    def P(x):
        return Paragraph(str(x or ''), cell)

    # Parcours des classes
    for classe in classes:
        # Titre de la classe
        titre_classe = f"Classe: {classe.nom} - {getattr(classe.ecole, 'nom', '')}"
        elements.append(Paragraph(titre_classe, styles['Heading2']))
        elements.append(Spacer(1, 0.2*cm))

        data = [[Paragraph(str(label), header_cell) for label in header]]

        # Élèves de la classe
        # Utiliser le related_name défini sur Eleve.classe = 'eleves'
        eleves = getattr(classe, 'eleves', None)
        if eleves is not None:
            eleves = eleves.all().order_by('nom', 'prenom')
        else:
            eleves = []

        for e in eleves:
            ligne = _donnees_tranches_eleve(e, annee_scolaire)

            # Construire le nom de l'élève sans déclencher d'erreur si un attribut manque
            nom_affiche = getattr(e, 'nom_complet', None) or f"{getattr(e, 'prenom', '')} {getattr(e, 'nom', '')}".strip()
            data.append([
                P(nom_affiche),
                f"{ligne['inscription']:,}".replace(',', ' '),
                f"{ligne['reinscription']:,}".replace(',', ' '),
                f"{ligne['tranche_1']:,}".replace(',', ' '),
                f"{ligne['tranche_2']:,}".replace(',', ' '),
                f"{ligne['tranche_3']:,}".replace(',', ' '),
                f"{ligne['total_du']:,}".replace(',', ' '),
                f"{ligne['total_paye']:,}".replace(',', ' '),
                f"{ligne['remise']:,}".replace(',', ' '),
                _pourcentages_pdf(ligne['pourcentages_remise']),
                f"{ligne['total_couvert']:,}".replace(',', ' '),
                f"{ligne['reste']:,}".replace(',', ' '),
                P(ligne['situation']),
            ])

        # Construire la table pour la classe
        col_widths = [
            3.8*cm, 1.7*cm, 1.8*cm, 1.65*cm, 1.65*cm, 1.65*cm,
            1.8*cm, 1.8*cm, 1.8*cm, 1.45*cm, 1.9*cm, 1.7*cm, 2.8*cm,
        ]
        table = Table(data, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#174A6E')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 6.2),
            ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
            ('ALIGN', (0,0), (0,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 2),
            ('RIGHTPADDING', (0,0), (-1,-1), 2),
            ('TOPPADDING', (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.6*cm))

    # Construire le document avec en-tête + filigrane logo
    def dessiner_entete(canvas, document):
        _draw_header_and_watermark(
            canvas,
            document,
            ecole=ecole_export,
            titre_override='Tranches par classe',
        )

    doc.build(
        elements,
        onFirstPage=dessiner_entete,
        onLaterPages=dessiner_entete,
    )
    return response

@login_required
def export_tranches_par_classe_excel(request):
    """Export Excel (XLSX) des tranches par classe, inscription et réinscription séparées.

    Filtres GET facultatifs: ecole, classe/classe_id, annee_scolaire.
    Respecte la séparation par école pour non-admin.
    """
    # Contrôle d'accès
    is_admin = user_is_admin(request.user)
    is_comptable = False
    try:
        if hasattr(request.user, 'profil'):
            is_comptable = (getattr(request.user.profil, 'role', None) == 'COMPTABLE')
    except Exception:
        is_comptable = False
    if not (is_admin or is_comptable):
        return HttpResponseForbidden("Accès refusé: vous n'avez pas l'autorisation d'exporter ce rapport.")

    # Import openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return HttpResponse("OpenPyXL n'est pas installé. Veuillez exécuter: pip install openpyxl", status=500)

    raw_ecole = (request.GET.get('ecole') or '').strip()
    raw_classe = (request.GET.get('classe') or request.GET.get('classe_id') or '').strip()
    annee_scolaire = (request.GET.get('annee_scolaire') or '').strip()

    def parse_int(value):
        try:
            return int(value)
        except Exception:
            return None

    ecole_id = parse_int(raw_ecole) if raw_ecole else None
    classe_id = parse_int(raw_classe) if raw_classe else None

    classes = Classe.objects.select_related('ecole').all()
    ecole_user = user_school(request.user)
    annee_active_xl = get_annee_active(request, ecole_user) if ecole_user else None
    restreindre = not user_is_admin(request.user) and ecole_user is not None
    if restreindre:
        classes = classes.filter(ecole=ecole_user)
    elif ecole_id:
        classes = classes.filter(ecole_id=ecole_id)
    if classe_id:
        classes = classes.filter(id=classe_id)
    if annee_active_xl and not annee_scolaire:
        classes = classes.filter(annee_scolaire=annee_active_xl)
    elif annee_scolaire:
        classes = classes.filter(annee_scolaire=annee_scolaire)
    classes = classes.order_by('ecole__nom', 'niveau', 'nom')[:200]

    wb = Workbook()
    ws_index = wb.active
    ws_index.title = 'Index'
    ws_index.append(['Tranches par classe', f"Année: {annee_scolaire}" if annee_scolaire else ''])
    ws_index.append(['Écoles / Classes listées:'])

    headers = TRANCHES_EXPORT_HEADERS

    for classe in classes:
        sheet_name = f"{classe.nom[:25]}"  # Limite Excel <=31
        ws = wb.create_sheet(title=sheet_name)
        ws.append([f"Classe: {classe.nom} - {getattr(classe.ecole, 'nom', '')}"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
        ws['A1'].fill = PatternFill('solid', fgColor='DCEAF3')
        ws['A1'].font = Font(color='174A6E', bold=True, size=13)
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.append(headers)

        eleves_mgr = getattr(classe, 'eleves', None)
        eleves = eleves_mgr.all().order_by('nom', 'prenom') if eleves_mgr is not None else []

        for e in eleves:
            ligne = _donnees_tranches_eleve(e, annee_scolaire)

            ws.append([
                getattr(e, 'nom_complet', f"{e.prenom} {e.nom}"),
                int(ligne['inscription']),
                int(ligne['reinscription']),
                int(ligne['tranche_1']),
                int(ligne['tranche_2']),
                int(ligne['tranche_3']),
                int(ligne['total_du']),
                int(ligne['total_paye']),
                int(ligne['remise']),
                _pourcentages_excel(ligne['pourcentages_remise']),
                int(ligne['total_couvert']),
                int(ligne['reste']),
                ligne['situation'],
            ])

        for cell_header in ws[2]:
            cell_header.fill = PatternFill('solid', fgColor='174A6E')
            cell_header.font = Font(color='FFFFFF', bold=True)
            cell_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f"A2:M{ws.max_row}"
        for row in ws.iter_rows(min_row=3, min_col=2, max_col=12):
            for cell_value in row:
                cell_value.number_format = '#,##0'
        for cell_value in ws['J'][2:]:
            cell_value.number_format = '0.0%'
        for col in range(1, 14):
            if col == 1:
                width = 27
            elif col == 13:
                width = 29
            else:
                width = 16
            ws.column_dimensions[get_column_letter(col)].width = width

        # Index line
        ws_index.append([getattr(classe.ecole, 'nom', ''), classe.nom, sheet_name])

    # Supprimer la feuille par défaut si vide
    if ws_index.max_row == 2:
        ws_index.append(['Aucune classe'])
    ws_index['A1'].fill = PatternFill('solid', fgColor='174A6E')
    ws_index['A1'].font = Font(color='FFFFFF', bold=True, size=13)
    ws_index['B1'].fill = PatternFill('solid', fgColor='174A6E')
    ws_index['B1'].font = Font(color='FFFFFF', bold=True)
    ws_index['A2'].font = Font(color='174A6E', bold=True)
    ws_index.column_dimensions['A'].width = 36
    ws_index.column_dimensions['B'].width = 27
    ws_index.column_dimensions['C'].width = 29
    ws_index.freeze_panes = 'A3'

    from io import BytesIO
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    resp = HttpResponse(stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    suffix = datetime.now().strftime('%Y%m%d')
    filename = f'tranches_par_classe_{suffix}.xlsx'
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
