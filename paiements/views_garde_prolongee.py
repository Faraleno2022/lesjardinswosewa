"""Rapport détaillé des élèves bénéficiant de la garde prolongée."""

from decimal import Decimal
from io import BytesIO
from xml.sax.saxutils import escape

from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.utils.text import slugify

from eleves.models import Classe, Ecole, Eleve
from eleves.tarification import (
    MONTANT_GARDE_PROLONGEE_COLLEGE_10,
    MONTANT_GARDE_PROLONGEE_MATERNELLE_GARDERIE,
    MONTANT_GARDE_PROLONGEE_PRIMAIRE,
    NIVEAUX_COLLEGE_FIN_REVISION,
    NIVEAUX_MATERNELLE_GARDERIE,
    NIVEAUX_PRIMAIRE,
    montant_scolarite_garde_prolongee,
)
from eleves.utils_annee import get_annee_active
from rapports.utils import _draw_header_and_watermark
from utilisateurs.permissions import can_view_reports
from utilisateurs.utils import filter_by_user_school, user_school

from .models import EcheancierPaiement
from .services import calculer_situations_echeanciers


ZERO = Decimal("0")

CYCLE_CHOICES = (
    ("MATERNELLE_GARDERIE", "Maternelle et garderie"),
    ("PRIMAIRE", "Primaire"),
    ("COLLEGE_10", "10ème année"),
    ("NON_COUVERT", "Niveau non couvert par le forfait"),
)

BAREME_GARDE_PROLONGEE = (
    ("Maternelle et garderie", MONTANT_GARDE_PROLONGEE_MATERNELLE_GARDERIE),
    ("Primaire (1ère à 6ème)", MONTANT_GARDE_PROLONGEE_PRIMAIRE),
    ("10ème année", MONTANT_GARDE_PROLONGEE_COLLEGE_10),
)


def _money(value):
    return f"{int(value or 0):,}".replace(",", " ")


def _excel_safe(value):
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _cycle_info(niveau):
    if niveau in NIVEAUX_MATERNELLE_GARDERIE:
        return "MATERNELLE_GARDERIE", "Maternelle et garderie"
    if niveau in NIVEAUX_PRIMAIRE:
        return "PRIMAIRE", "Primaire"
    if niveau in NIVEAUX_COLLEGE_FIN_REVISION:
        return "COLLEGE_10", "10ème année"
    return "NON_COUVERT", "Niveau non couvert"


def _situation_label(situation):
    if situation is None:
        return "Sans échéancier", "sans_echeancier"
    total_due = Decimal(str(situation["total_du"] or 0))
    paid = Decimal(str(situation["encaisse"] or 0))
    discount = Decimal(str(situation["remises"] or 0))
    remaining = Decimal(str(situation["reste"] or 0))
    if total_due <= 0:
        return "Échéancier vide", "a_payer"
    if remaining <= 0:
        return ("Soldé - remise appliquée" if discount else "Soldé"), "solde"
    if paid + discount > 0:
        return (
            "Partiel - remise appliquée" if discount else "Partiel"
        ), "partiel"
    return "À payer", "a_payer"


def _valid_id(value):
    value = (value or "").strip()
    return int(value) if value.isdigit() else None


def collect_garde_prolongee_data(request):
    """Retourne le périmètre filtré commun à l'écran, au PDF et à Excel."""
    accessible_classes = filter_by_user_school(
        Classe.objects.select_related("ecole").order_by(
            "ecole__nom", "-annee_scolaire", "niveau", "nom"
        ),
        request.user,
        "ecole",
    )
    classes = list(accessible_classes)
    schools = list(
        Ecole.objects.filter(pk__in={item.ecole_id for item in classes}).order_by("nom")
    )
    years = sorted({item.annee_scolaire for item in classes}, reverse=True)

    class_id = _valid_id(request.GET.get("classe_id") or request.GET.get("classe"))
    selected_class = next((item for item in classes if item.pk == class_id), None)

    requested_school_id = _valid_id(request.GET.get("ecole_id"))
    allowed_school_ids = {item.pk for item in schools}
    if selected_class:
        selected_school_id = selected_class.ecole_id
    elif requested_school_id in allowed_school_ids:
        selected_school_id = requested_school_id
    else:
        school = user_school(request.user)
        selected_school_id = school.pk if school and school.pk in allowed_school_ids else None
    selected_school = next(
        (item for item in schools if item.pk == selected_school_id), None
    )

    requested_year = (request.GET.get("annee_scolaire") or "").strip()
    if selected_class:
        selected_year = selected_class.annee_scolaire
    elif requested_year in years:
        selected_year = requested_year
    else:
        selected_year = get_annee_active(request, selected_school) or (
            years[0] if years else ""
        )

    cycle_filter = (request.GET.get("cycle") or "").strip().upper()
    if cycle_filter not in {"", *(code for code, _label in CYCLE_CHOICES)}:
        cycle_filter = ""
    nature_filter = (request.GET.get("nature") or "").strip().upper()
    if nature_filter not in {"", "INSCRIPTION", "REINSCRIPTION", "SANS_ECHEANCIER"}:
        nature_filter = ""
    situation_filter = (request.GET.get("situation") or "").strip().lower()
    if situation_filter not in {
        "", "solde", "partiel", "a_payer", "sans_echeancier", "a_verifier"
    }:
        situation_filter = ""
    student_status = (request.GET.get("statut_eleve") or "").strip().upper()
    status_labels = dict(Eleve.STATUT_CHOICES)
    if student_status not in {"", *status_labels}:
        student_status = ""
    search = (request.GET.get("q") or "").strip()

    scope_classes = accessible_classes
    if selected_school_id:
        scope_classes = scope_classes.filter(ecole_id=selected_school_id)
    if selected_year:
        scope_classes = scope_classes.filter(annee_scolaire=selected_year)
    if selected_class:
        scope_classes = scope_classes.filter(pk=selected_class.pk)

    students = (
        Eleve.objects.select_related(
            "classe", "classe__ecole", "responsable_principal"
        )
        .filter(
            garde_prolongee=True,
            est_dans_corbeille=False,
            classe__in=scope_classes,
        )
        .order_by("classe__ecole__nom", "classe__nom", "nom", "prenom")
    )
    if student_status:
        students = students.filter(statut=student_status)
    if cycle_filter:
        if cycle_filter == "MATERNELLE_GARDERIE":
            students = students.filter(classe__niveau__in=NIVEAUX_MATERNELLE_GARDERIE)
        elif cycle_filter == "PRIMAIRE":
            students = students.filter(classe__niveau__in=NIVEAUX_PRIMAIRE)
        elif cycle_filter == "COLLEGE_10":
            students = students.filter(classe__niveau__in=NIVEAUX_COLLEGE_FIN_REVISION)
        else:
            covered = (
                NIVEAUX_MATERNELLE_GARDERIE
                | NIVEAUX_PRIMAIRE
                | NIVEAUX_COLLEGE_FIN_REVISION
            )
            students = students.exclude(classe__niveau__in=covered)
    if search:
        students = students.filter(
            Q(matricule__icontains=search)
            | Q(nom__icontains=search)
            | Q(prenom__icontains=search)
            | Q(responsable_principal__nom__icontains=search)
            | Q(responsable_principal__prenom__icontains=search)
            | Q(responsable_principal__telephone__icontains=search)
        )

    students = list(students)
    student_ids = [item.pk for item in students]
    schedule_years = {item.classe.annee_scolaire for item in students}
    schedules = list(
        EcheancierPaiement.objects.filter(
            eleve_id__in=student_ids,
            annee_scolaire__in=schedule_years,
        ).select_related("eleve", "eleve__classe")
    )
    schedule_by_key = {
        (item.eleve_id, item.annee_scolaire): item for item in schedules
    }
    situations = calculer_situations_echeanciers(schedules)

    rows = []
    for student in students:
        school_year = student.classe.annee_scolaire
        schedule = schedule_by_key.get((student.pk, school_year))
        situation = situations.get(schedule.pk) if schedule else None
        cycle_code, cycle_label = _cycle_info(student.classe.niveau)
        reference_forfait = montant_scolarite_garde_prolongee(student.classe.niveau)
        applied_forfait = None
        admission_due = None
        admission_label = "Non défini"
        nature_code = "SANS_ECHEANCIER"
        total_due = paid = discount = remaining = None
        if schedule:
            applied_forfait = sum(
                (
                    Decimal(str(schedule.tranche_1_due or 0)),
                    Decimal(str(schedule.tranche_2_due or 0)),
                    Decimal(str(schedule.tranche_3_due or 0)),
                ),
                ZERO,
            )
            admission_due = Decimal(str(schedule.frais_inscription_du or 0))
            admission_label = schedule.libelle_frais_admission
            nature_code = schedule.nature_frais
            total_due = Decimal(str(situation["total_du"] or 0))
            paid = Decimal(str(situation["encaisse"] or 0))
            discount = Decimal(str(situation["remises"] or 0))
            remaining = Decimal(str(situation["reste"] or 0))

        situation_label, situation_code = _situation_label(situation)
        if schedule is None:
            conformity_label, conformity_code = "Sans échéancier", "sans_echeancier"
        elif reference_forfait is None:
            conformity_label, conformity_code = "Niveau non couvert", "a_verifier"
        elif applied_forfait != reference_forfait:
            conformity_label, conformity_code = "Forfait à vérifier", "a_verifier"
        else:
            conformity_label, conformity_code = "Forfait conforme", "conforme"

        if nature_filter and nature_code != nature_filter:
            continue
        if situation_filter == "a_verifier":
            if conformity_code != "a_verifier":
                continue
        elif situation_filter and situation_code != situation_filter:
            continue

        responsible = student.responsable_principal
        rows.append(
            {
                "student_id": student.pk,
                "matricule": student.matricule,
                "student": f"{student.nom} {student.prenom}".strip(),
                "student_status": student.get_statut_display(),
                "school": student.classe.ecole.nom,
                "class": student.classe.nom,
                "level": student.classe.get_niveau_display(),
                "school_year": school_year,
                "cycle_code": cycle_code,
                "cycle": cycle_label,
                "responsible": responsible.nom_complet if responsible else "Non renseigné",
                "phone": responsible.telephone if responsible else "",
                "nature_code": nature_code,
                "admission_label": admission_label,
                "admission_due": admission_due,
                "reference_forfait": reference_forfait,
                "applied_forfait": applied_forfait,
                "total_due": total_due,
                "paid": paid,
                "discount": discount,
                "remaining": remaining,
                "situation": situation_label,
                "situation_code": situation_code,
                "conformity": conformity_label,
                "conformity_code": conformity_code,
            }
        )

    query = request.GET.copy()
    query.pop("page", None)
    query.pop("fragment", None)
    amount_rows = [row for row in rows if row["total_due"] is not None]
    return {
        "titre_page": "Élèves en garde prolongée",
        "rows": rows,
        "schools": schools,
        "classes": classes,
        "years": years,
        "cycle_choices": CYCLE_CHOICES,
        "student_status_choices": Eleve.STATUT_CHOICES,
        "selected_school_id": selected_school_id,
        "selected_school": selected_school,
        "selected_class_id": selected_class.pk if selected_class else None,
        "selected_year": selected_year,
        "cycle_filter": cycle_filter,
        "nature_filter": nature_filter,
        "situation_filter": situation_filter,
        "student_status": student_status,
        "q": search,
        "query_string": query.urlencode(),
        "student_count": len(rows),
        "forfait_reference_total": sum(
            (row["reference_forfait"] or ZERO for row in rows), ZERO
        ),
        "admission_total": sum(
            (row["admission_due"] or ZERO for row in rows), ZERO
        ),
        "total_due": sum((row["total_due"] or ZERO for row in amount_rows), ZERO),
        "paid_total": sum((row["paid"] or ZERO for row in amount_rows), ZERO),
        "discount_total": sum(
            (row["discount"] or ZERO for row in amount_rows), ZERO
        ),
        "remaining_total": sum(
            (row["remaining"] or ZERO for row in amount_rows), ZERO
        ),
        "settled_count": sum(1 for row in rows if row["situation_code"] == "solde"),
        "without_schedule_count": sum(
            1 for row in rows if row["nature_code"] == "SANS_ECHEANCIER"
        ),
        "to_check_count": sum(
            1 for row in rows if row["conformity_code"] == "a_verifier"
        ),
        "bareme": BAREME_GARDE_PROLONGEE,
        "generated_at": timezone.localtime(),
    }


@can_view_reports
def garde_prolongee_report(request):
    context = collect_garde_prolongee_data(request)
    paginator = Paginator(context["rows"], 50)
    context["page_obj"] = paginator.get_page(request.GET.get("page") or 1)
    context["rows"] = context["page_obj"].object_list
    if request.GET.get("fragment") == "1":
        return render(request, "paiements/_garde_prolongee_resultats.html", context)
    return render(request, "paiements/garde_prolongee_report.html", context)


def _report_filename(data, extension):
    scope = data["selected_school"].nom if data["selected_school"] else "toutes_ecoles"
    period = data["selected_year"] or str(timezone.localdate())
    return (
        f"garde_prolongee_{slugify(scope) or 'etablissement'}_"
        f"{slugify(period) or 'toutes-annees'}.{extension}"
    )


@can_view_reports
def export_garde_prolongee_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    data = collect_garde_prolongee_data(request)
    stream = BytesIO()
    page_size = landscape(A4)
    document = SimpleDocTemplate(
        stream,
        pagesize=page_size,
        leftMargin=0.6 * cm,
        rightMargin=0.6 * cm,
        topMargin=1.7 * cm,
        bottomMargin=0.8 * cm,
        title="Rapport détaillé des élèves en garde prolongée",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "GardeTitle", parent=styles["Title"], fontSize=15, leading=18,
        textColor=colors.HexColor("#174A6E"), spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "GardeSubtitle", parent=styles["Normal"], fontSize=8, leading=10,
        textColor=colors.HexColor("#5D6973"), spaceAfter=6,
    )
    cell_style = ParagraphStyle(
        "GardeCell", parent=styles["Normal"], fontSize=5.7, leading=6.8,
    )
    header_style = ParagraphStyle(
        "GardeHeader", parent=cell_style, fontName="Helvetica-Bold",
        textColor=colors.white, alignment=1,
    )

    story = [
        Paragraph("RAPPORT DÉTAILLÉ — GARDE PROLONGÉE", title_style),
        Paragraph(
            escape(
                f"Année scolaire : {data['selected_year'] or 'Toutes'} | "
                f"École : {data['selected_school'].nom if data['selected_school'] else 'Toutes les écoles'} | "
                f"Généré le {data['generated_at']:%d/%m/%Y à %H:%M}"
            ),
            subtitle_style,
        ),
    ]

    summary = Table(
        [
            ["Élèves", "Forfaits de référence", "Admission", "Total dû", "Encaissé", "Remises", "Reste"],
            [
                data["student_count"], _money(data["forfait_reference_total"]),
                _money(data["admission_total"]), _money(data["total_due"]),
                _money(data["paid_total"]), _money(data["discount_total"]),
                _money(data["remaining_total"]),
            ],
        ],
        colWidths=[3.8 * cm] * 7,
    )
    summary.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#174A6E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#AAB7C0")),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#EAF2F7")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([summary, Spacer(1, 0.18 * cm)])

    scale = Table(
        [["Barème annuel de garde prolongée", "Montant (GNF)"]] + [
            [label, _money(amount)] for label, amount in data["bareme"]
        ],
        colWidths=[8 * cm, 3 * cm],
    )
    scale.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#207A54")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#AAB7C0")),
    ]))
    story.extend([scale, Spacer(1, 0.25 * cm)])

    headers = [
        "Matricule", "Élève / responsable", "Classe / école", "Cycle",
        "Admission", "Frais", "Forfait réf.", "Forfait appliqué", "Total dû",
        "Encaissé", "Remise", "Reste", "Situation / contrôle",
    ]
    table_data = [[Paragraph(label, header_style) for label in headers]]
    for row in data["rows"]:
        table_data.append([
            Paragraph(escape(row["matricule"]), cell_style),
            Paragraph(
                f"{escape(row['student'])}<br/>"
                f"{escape((row['responsible'] + ' ' + row['phone']).strip())}",
                cell_style,
            ),
            Paragraph(
                f"{escape(row['class'])}<br/>{escape(row['school'])}", cell_style
            ),
            Paragraph(
                f"{escape(row['cycle'])}<br/>{escape(row['level'])}", cell_style
            ),
            Paragraph(escape(row["admission_label"]), cell_style),
            "—" if row["admission_due"] is None else _money(row["admission_due"]),
            "—" if row["reference_forfait"] is None else _money(row["reference_forfait"]),
            "—" if row["applied_forfait"] is None else _money(row["applied_forfait"]),
            "—" if row["total_due"] is None else _money(row["total_due"]),
            "—" if row["paid"] is None else _money(row["paid"]),
            "—" if row["discount"] is None else _money(row["discount"]),
            "—" if row["remaining"] is None else _money(row["remaining"]),
            Paragraph(
                f"{escape(row['situation'])}<br/>{escape(row['conformity'])}",
                cell_style,
            ),
        ])
    if len(table_data) == 1:
        table_data.append([
            Paragraph("Aucun élève ne correspond aux filtres sélectionnés.", cell_style)
        ] + [""] * (len(headers) - 1))

    detail = Table(
        table_data,
        repeatRows=1,
        colWidths=[
            1.6 * cm, 2.7 * cm, 2.6 * cm, 2.3 * cm, 1.8 * cm, 1.45 * cm,
            1.6 * cm, 1.7 * cm, 1.55 * cm, 1.55 * cm, 1.4 * cm, 1.55 * cm,
            2.5 * cm,
        ],
    )
    detail.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#174A6E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (5, 1), (11, -1), "RIGHT"),
        ("FONTSIZE", (0, 1), (-1, -1), 5.7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B7C4CC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FA")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(detail)

    def draw_page(canvas, doc):
        _draw_header_and_watermark(
            canvas,
            doc,
            ecole=data["selected_school"],
            titre_override="Garde prolongée",
        )
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#5D6973"))
        canvas.drawRightString(page_size[0] - 22, 16, f"Page {doc.page}")
        canvas.restoreState()

    document.build(story, onFirstPage=draw_page, onLaterPages=draw_page)
    response = HttpResponse(stream.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="{_report_filename(data, "pdf")}"'
    )
    response["Cache-Control"] = "private, no-store"
    return response


@can_view_reports
def export_garde_prolongee_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    data = collect_garde_prolongee_data(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Garde prolongée"
    sheet.merge_cells("A1:P1")
    sheet["A1"] = "RAPPORT DÉTAILLÉ — ÉLÈVES EN GARDE PROLONGÉE"
    sheet["A1"].font = Font(size=15, bold=True, color="174A6E")
    sheet["A2"] = "École"
    sheet["B2"] = data["selected_school"].nom if data["selected_school"] else "Toutes les écoles"
    sheet["A3"] = "Année scolaire"
    sheet["B3"] = data["selected_year"] or "Toutes"
    sheet["A4"] = "Généré le"
    sheet["B4"] = data["generated_at"].strftime("%d/%m/%Y à %H:%M")
    sheet["D2"] = "Élèves"
    sheet["E2"] = data["student_count"]
    sheet["D3"] = "Total dû"
    sheet["E3"] = int(data["total_due"])
    sheet["F2"] = "Encaissé"
    sheet["G2"] = int(data["paid_total"])
    sheet["F3"] = "Remises"
    sheet["G3"] = int(data["discount_total"])
    sheet["H2"] = "Reste"
    sheet["I2"] = int(data["remaining_total"])

    headers = [
        "Matricule", "Élève", "Statut élève", "Responsable", "Téléphone",
        "École", "Classe", "Niveau", "Cycle", "Année scolaire",
        "Nature admission", "Frais admission (GNF)", "Forfait référence (GNF)",
        "Forfait appliqué (GNF)", "Total dû (GNF)", "Encaissé (GNF)",
        "Remises (GNF)", "Reste (GNF)", "Situation", "Contrôle forfait",
    ]
    header_row = 6
    for column, label in enumerate(headers, 1):
        sheet.cell(header_row, column, label)
    for row_number, row in enumerate(data["rows"], header_row + 1):
        values = [
            row["matricule"], row["student"], row["student_status"],
            row["responsible"], row["phone"], row["school"], row["class"],
            row["level"], row["cycle"], row["school_year"],
            row["admission_label"],
            None if row["admission_due"] is None else int(row["admission_due"]),
            None if row["reference_forfait"] is None else int(row["reference_forfait"]),
            None if row["applied_forfait"] is None else int(row["applied_forfait"]),
            None if row["total_due"] is None else int(row["total_due"]),
            None if row["paid"] is None else int(row["paid"]),
            None if row["discount"] is None else int(row["discount"]),
            None if row["remaining"] is None else int(row["remaining"]),
            row["situation"], row["conformity"],
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row_number, column, _excel_safe(value))

    thin = Side(style="thin", color="B7C4CC")
    for cell in sheet[header_row]:
        cell.fill = PatternFill("solid", fgColor="174A6E")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in sheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(12, 19):
        for cell in sheet.iter_cols(
            min_col=column, max_col=column, min_row=header_row + 1
        ):
            for item in cell:
                item.number_format = '#,##0 "GNF"'
    widths = [
        16, 25, 14, 24, 16, 24, 21, 19, 23, 15,
        22, 20, 22, 22, 19, 19, 18, 18, 25, 22,
    ]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:T{max(header_row, sheet.max_row)}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.print_title_rows = f"{header_row}:{header_row}"

    scale = workbook.create_sheet("Barème")
    scale.append(["Cycle", "Forfait annuel de scolarité (GNF)"])
    for label, amount in data["bareme"]:
        scale.append([label, int(amount)])
    scale.append([])
    scale.append([
        "Règle",
        "Les frais d'inscription ou de réinscription s'ajoutent au forfait du cycle.",
    ])
    scale.column_dimensions["A"].width = 36
    scale.column_dimensions["B"].width = 70
    scale["A1"].fill = PatternFill("solid", fgColor="207A54")
    scale["B1"].fill = PatternFill("solid", fgColor="207A54")
    scale["A1"].font = scale["B1"].font = Font(color="FFFFFF", bold=True)
    for cell in scale["B"][1:4]:
        cell.number_format = '#,##0 "GNF"'

    stream = BytesIO()
    workbook.save(stream)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{_report_filename(data, "xlsx")}"'
    )
    response["Cache-Control"] = "private, no-store"
    return response
