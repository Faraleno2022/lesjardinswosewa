from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook
from pypdf import PdfReader

from eleves.models import Classe, Ecole, Eleve, Responsable
from paiements.models import EcheancierPaiement

from .support import TEST_MIDDLEWARE


@override_settings(MIDDLEWARE=TEST_MIDDLEWARE)
class RapportGardeProlongeeTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username="rapport-garde",
            email="rapport-garde@example.com",
            password="mot-de-passe-test",
        )
        self.client.force_login(self.user)
        self.ecole = Ecole.objects.create(
            nom="École garde détaillée",
            adresse="Conakry",
            telephone="+224620000750",
            directeur="Direction",
        )
        self.responsable = Responsable.objects.create(
            prenom="Parent",
            nom="Disponible",
            relation="MERE",
            telephone="+224620000751",
            adresse="Conakry",
        )
        self.maternelle = self._classe("Grande section", "GRANDE_SECTION")
        self.primaire = self._classe("Primaire 1", "PRIMAIRE_1")
        self.primaire_sans_echeancier = self._classe("Primaire 2", "PRIMAIRE_2")
        self.dixieme = self._classe("10ème année", "COLLEGE_10")

        self.eleve_maternelle = self._eleve(
            self.maternelle, "GARDE-MAT", "Aminata", "Camara"
        )
        self.eleve_primaire = self._eleve(
            self.primaire, "GARDE-PRI", "Mamadou", "Diallo"
        )
        self.eleve_sans_echeancier = self._eleve(
            self.primaire_sans_echeancier, "GARDE-SANS", "Fatou", "Barry"
        )
        self.eleve_dixieme = self._eleve(
            self.dixieme, "GARDE-10", "Ibrahima", "Sylla"
        )
        self._eleve(
            self.primaire, "SANS-GARDE", "Élève", "Ordinaire",
            garde_prolongee=False,
        )

        self._echeancier(
            self.eleve_maternelle,
            nature="REINSCRIPTION",
            admission="30000",
            tranches=("900000", "900000", "900000"),
            payes=("30000", "900000", "900000", "900000"),
        )
        self._echeancier(
            self.eleve_primaire,
            nature="INSCRIPTION",
            admission="50000",
            tranches=("1000000", "900000", "900000"),
            payes=("50000", "500000", "0", "0"),
        )
        self._echeancier(
            self.eleve_dixieme,
            nature="INSCRIPTION",
            admission="70000",
            tranches=("950000", "950000", "950000"),
            payes=("0", "0", "0", "0"),
        )

    def _classe(self, nom, niveau):
        return Classe.objects.create(
            ecole=self.ecole,
            nom=nom,
            niveau=niveau,
            annee_scolaire="2025-2026",
        )

    def _eleve(self, classe, matricule, prenom, nom, garde_prolongee=True):
        return Eleve.objects.create(
            matricule=matricule,
            prenom=prenom,
            nom=nom,
            sexe="F",
            date_naissance=date(2015, 1, 1),
            classe=classe,
            date_inscription=date(2025, 9, 1),
            statut="ACTIF",
            garde_prolongee=garde_prolongee,
            responsable_principal=self.responsable,
        )

    def _echeancier(self, eleve, nature, admission, tranches, payes):
        return EcheancierPaiement.objects.create(
            eleve=eleve,
            annee_scolaire="2025-2026",
            nature_frais=nature,
            frais_inscription_du=Decimal(admission),
            tranche_1_due=Decimal(tranches[0]),
            tranche_2_due=Decimal(tranches[1]),
            tranche_3_due=Decimal(tranches[2]),
            frais_inscription_paye=Decimal(payes[0]),
            tranche_1_payee=Decimal(payes[1]),
            tranche_2_payee=Decimal(payes[2]),
            tranche_3_payee=Decimal(payes[3]),
            date_echeance_inscription=date(2025, 9, 1),
            date_echeance_tranche_1=date(2025, 10, 1),
            date_echeance_tranche_2=date(2026, 1, 15),
            date_echeance_tranche_3=date(2026, 3, 15),
        )

    def test_rapport_affiche_tous_les_forfaits_et_frais_admission(self):
        response = self.client.get(
            reverse("paiements:garde_prolongee_report"),
            {"annee_scolaire": "2025-2026"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["student_count"], 4)
        self.assertNotContains(response, "SANS-GARDE")
        rows = {row["matricule"]: row for row in response.context["page_obj"].object_list}

        maternelle = rows["GARDE-MAT"]
        self.assertEqual(maternelle["reference_forfait"], Decimal("2700000"))
        self.assertEqual(maternelle["admission_due"], Decimal("30000"))
        self.assertEqual(maternelle["total_due"], Decimal("2730000"))
        self.assertEqual(maternelle["nature_code"], "REINSCRIPTION")
        self.assertEqual(maternelle["situation_code"], "solde")

        primaire = rows["GARDE-PRI"]
        self.assertEqual(primaire["reference_forfait"], Decimal("2800000"))
        self.assertEqual(primaire["admission_due"], Decimal("50000"))
        self.assertEqual(primaire["total_due"], Decimal("2850000"))
        self.assertEqual(primaire["remaining"], Decimal("2300000"))

        dixieme = rows["GARDE-10"]
        self.assertEqual(dixieme["reference_forfait"], Decimal("2850000"))
        self.assertEqual(dixieme["admission_due"], Decimal("70000"))
        self.assertEqual(dixieme["total_due"], Decimal("2920000"))

        absent = rows["GARDE-SANS"]
        self.assertIsNone(absent["total_due"])
        self.assertEqual(absent["reference_forfait"], Decimal("2800000"))
        self.assertEqual(absent["situation_code"], "sans_echeancier")
        self.assertEqual(response.context["without_schedule_count"], 1)

    def test_filtres_dynamiques_cycle_nature_recherche_et_fragment(self):
        response = self.client.get(
            reverse("paiements:garde_prolongee_report"),
            {
                "annee_scolaire": "2025-2026",
                "cycle": "PRIMAIRE",
                "nature": "INSCRIPTION",
                "q": "GARDE-PRI",
                "fragment": "1",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "paiements/_garde_prolongee_resultats.html")
        self.assertEqual(response.context["student_count"], 1)
        self.assertContains(response, "GARDE-PRI")
        self.assertNotContains(response, "GARDE-MAT")

        missing = self.client.get(
            reverse("paiements:garde_prolongee_report"),
            {
                "annee_scolaire": "2025-2026",
                "nature": "SANS_ECHEANCIER",
                "fragment": "1",
            },
        )
        self.assertEqual(missing.context["student_count"], 1)
        self.assertContains(missing, "GARDE-SANS")

    def test_exports_pdf_et_excel_reprennent_les_memes_details(self):
        params = {"annee_scolaire": "2025-2026"}
        pdf = self.client.get(reverse("paiements:export_garde_prolongee_pdf"), params)
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf["Content-Type"], "application/pdf")
        self.assertTrue(pdf.content.startswith(b"%PDF"))
        pdf_text = "\n".join(
            page.extract_text() or "" for page in PdfReader(BytesIO(pdf.content)).pages
        )
        self.assertIn("GARDE-MAT", pdf_text)
        self.assertIn("2 700 000", pdf_text)
        self.assertIn("30 000", pdf_text)

        excel = self.client.get(
            reverse("paiements:export_garde_prolongee_excel"), params
        )
        self.assertEqual(excel.status_code, 200)
        workbook = load_workbook(BytesIO(excel.content), data_only=True)
        self.assertEqual(workbook.sheetnames, ["Garde prolongée", "Barème"])
        sheet = workbook["Garde prolongée"]
        headers = [sheet.cell(6, column).value for column in range(1, 21)]
        self.assertIn("Frais admission (GNF)", headers)
        self.assertIn("Forfait référence (GNF)", headers)
        self.assertIn("Contrôle forfait", headers)
        exported = {
            sheet.cell(row, 1).value: [sheet.cell(row, column).value for column in range(1, 21)]
            for row in range(7, sheet.max_row + 1)
        }
        self.assertEqual(exported["GARDE-MAT"][10], "Frais de réinscription")
        self.assertEqual(exported["GARDE-MAT"][11], 30000)
        self.assertEqual(exported["GARDE-MAT"][12], 2700000)
        self.assertEqual(exported["GARDE-MAT"][14], 2730000)
        self.assertEqual(workbook["Barème"]["B4"].value, 2850000)

    def test_acces_depuis_le_tableau_de_bord(self):
        response = self.client.get(reverse("paiements:tableau_bord"))
        self.assertContains(response, reverse("paiements:garde_prolongee_report"))

    def test_rapport_et_exports_exigent_la_permission_rapports(self):
        simple_user = get_user_model().objects.create_user(
            username="sans-permission-garde",
            password="mot-de-passe-test",
        )
        simple_user.profil.role = "ENSEIGNANT"
        simple_user.profil.peut_consulter_rapports = False
        simple_user.profil.save(update_fields=["role", "peut_consulter_rapports"])
        self.client.force_login(simple_user)

        for url_name in (
            "paiements:garde_prolongee_report",
            "paiements:export_garde_prolongee_pdf",
            "paiements:export_garde_prolongee_excel",
        ):
            with self.subTest(url_name=url_name):
                response = self.client.get(reverse(url_name))
                self.assertEqual(response.status_code, 403)
