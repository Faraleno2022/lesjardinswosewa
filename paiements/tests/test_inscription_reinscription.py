from datetime import date
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook
from pypdf import PdfReader

from eleves.models import Classe, Ecole, Eleve, GrilleTarifaire, Responsable
from paiements.models import (
    EcheancierPaiement,
    ModePaiement,
    Paiement,
    PaiementRemise,
    RemiseReduction,
    TypePaiement,
)
from paiements.views import _allocate_payment_to_echeancier
from paiements.views_tranches import (
    _donnees_tranches_eleve,
    _pourcentages_pdf,
)

from .support import TEST_MIDDLEWARE


@override_settings(MIDDLEWARE=TEST_MIDDLEWARE)
class InscriptionReinscriptionReportingTests(TestCase):
    def setUp(self):
        self.ecole = Ecole.objects.create(
            nom="École ventilation",
            adresse="Conakry",
            telephone="+224620000111",
            directeur="Direction",
        )
        self.classe = Classe.objects.create(
            nom="6ème A",
            ecole=self.ecole,
            niveau="PRIMAIRE_6",
            annee_scolaire="2025-2026",
        )
        self.responsable = Responsable.objects.create(
            prenom="Parent",
            nom="Test",
            relation="PERE",
            telephone="+224620000112",
            adresse="Conakry",
        )
        GrilleTarifaire.objects.create(
            ecole=self.ecole,
            niveau=self.classe.niveau,
            annee_scolaire=self.classe.annee_scolaire,
            frais_inscription=Decimal("30000"),
            frais_reinscription=Decimal("20000"),
            tranche_1=Decimal("100000"),
            tranche_2=Decimal("100000"),
            tranche_3=Decimal("100000"),
        )
        self.user = get_user_model().objects.create_superuser(
            username="admin_ventilation",
            email="admin@example.com",
            password="pass12345",
        )
        self.client.force_login(self.user)

    def _student(self, matricule, prenom):
        return Eleve.objects.create(
            nom="Camara",
            prenom=prenom,
            matricule=matricule,
            classe=self.classe,
            sexe="F",
            date_naissance=date(2015, 1, 1),
            lieu_naissance="Conakry",
            date_inscription=date(2025, 9, 1),
            responsable_principal=self.responsable,
        )

    def _schedule(self, student, nature, fee, tranche_1=0, admission_paye=0):
        return EcheancierPaiement.objects.create(
            eleve=student,
            annee_scolaire="2025-2026",
            nature_frais=nature,
            frais_inscription_du=Decimal(str(fee)),
            tranche_1_due=Decimal(str(tranche_1)),
            tranche_2_due=0,
            tranche_3_due=0,
            frais_inscription_paye=Decimal(str(admission_paye)),
            date_echeance_inscription=date(2025, 9, 1),
            date_echeance_tranche_1=date(2026, 1, 15),
            date_echeance_tranche_2=date(2026, 3, 15),
            date_echeance_tranche_3=date(2026, 5, 15),
        )

    def test_report_columns_are_disjoint_and_total_is_not_duplicated(self):
        self._schedule(self._student("VENT-001", "Aïssata"), "INSCRIPTION", 30000)
        self._schedule(self._student("VENT-002", "Mariam"), "REINSCRIPTION", 20000)

        response = self.client.get(reverse("paiements:liste_paiements"))

        self.assertEqual(response.status_code, 200)
        totals = response.context["totaux_du"]
        self.assertEqual(totals["frais_inscription_total"], 30000)
        self.assertEqual(totals["frais_reinscription_total"], 20000)
        self.assertEqual(totals["du_global_net"], 50000)
        self.assertEqual(totals["frais_reinscription_pct"], 40.0)
        row = response.context["totaux_du_detail_classes"][0]
        self.assertEqual(row["frais_inscription_total"], 30000)
        self.assertEqual(row["frais_reinscription_total"], 20000)
        self.assertEqual(row["du_global_net"], 50000)

    def test_equal_tariffs_remain_distinguishable_by_nature(self):
        self._schedule(self._student("VENT-003", "Fatou"), "INSCRIPTION", 20000)
        self._schedule(self._student("VENT-004", "Hawa"), "REINSCRIPTION", 20000)

        response = self.client.get(reverse("paiements:liste_paiements"))
        totals = response.context["totaux_du"]

        self.assertEqual(totals["frais_inscription_total"], 20000)
        self.assertEqual(totals["frais_reinscription_total"], 20000)
        self.assertEqual(totals["du_global_net"], 40000)

    def test_excel_export_uses_the_same_disjoint_distribution(self):
        self._schedule(self._student("VENT-005", "Kadiatou"), "INSCRIPTION", 30000)
        self._schedule(self._student("VENT-006", "Nènè"), "REINSCRIPTION", 20000)

        response = self.client.get(reverse("paiements:export_recap_par_classe_excel"))

        self.assertEqual(response.status_code, 200)
        worksheet = load_workbook(BytesIO(response.content), data_only=True).active
        values = [cell.value for cell in worksheet[2]]
        self.assertEqual(values[4], 30000)
        self.assertEqual(values[5], 20000)
        self.assertEqual(values[6], 40.0)
        self.assertEqual(values[7], 50000)

    def test_tranches_exports_separate_reenrollment_from_enrollment(self):
        student = self._student("VENT-009", "Saran")
        self._schedule(
            student,
            "REINSCRIPTION",
            20000,
            tranche_1=100000,
            admission_paye=20000,
        )

        response = self.client.get(
            reverse("paiements:export_tranches_par_classe_excel"),
            {"annee_scolaire": "2025-2026", "classe": self.classe.pk},
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        worksheet = next(ws for ws in workbook.worksheets if ws.title != "Index")
        self.assertEqual(
            [cell.value for cell in worksheet[2]],
            [
                "Élève", "Inscription payée", "Réinscription payée",
                "Tranche 1 payée", "Tranche 2 payée", "Tranche 3 payée",
                "Total dû", "Encaissé", "Remise (GNF)", "Remise (%)",
                "Total couvert", "Reste", "Situation",
            ],
        )
        values = [cell.value for cell in worksheet[3]]
        self.assertEqual(values[1], 0)
        self.assertEqual(values[2], 20000)
        self.assertEqual(values[6], 120000)
        self.assertEqual(values[7], 20000)
        self.assertEqual(values[8], 0)
        self.assertIsNone(values[9])
        self.assertEqual(values[10], 20000)
        self.assertEqual(values[11], 100000)
        self.assertEqual(values[12], "Partiel")

    def test_tranches_exporte_remise_pourcentage_et_statut_solde(self):
        student = self._student("VENT-010", "M'Mah")
        schedule = self._schedule(
            student, "REINSCRIPTION", 20000, tranche_1=100000, admission_paye=20000,
        )
        schedule.tranche_1_payee = Decimal("90000")
        schedule.save(update_fields=["tranche_1_payee"])
        payment_type = TypePaiement.objects.create(
            nom="Tranche 1 avec remise", categorie="SCOLARITE",
        )
        mode = ModePaiement.objects.create(nom="Espèces")
        payment = Paiement.objects.create(
            eleve=student,
            type_paiement=payment_type,
            mode_paiement=mode,
            numero_recu="VENT-REM-001",
            montant=Decimal("110000"),
            annee_scolaire="2025-2026",
            date_paiement=date(2026, 1, 15),
            statut="VALIDE",
        )
        discount = RemiseReduction.objects.create(
            nom="Remise fratrie 15 %",
            type_remise="POURCENTAGE",
            valeur=Decimal("15"),
            motif="FRATRIE",
            date_debut=date(2025, 9, 1),
            date_fin=date(2026, 6, 30),
        )
        PaiementRemise.objects.create(
            paiement=payment,
            remise=discount,
            montant_remise=Decimal("10000"),
            motif="GESTE_COMMERCIAL",
            tranches_concernees="1",
            base_calcul="TRANCHES_DUES",
        )

        params = {"annee_scolaire": "2025-2026", "classe": self.classe.pk}
        excel = self.client.get(reverse("paiements:export_tranches_par_classe_excel"), params)
        workbook = load_workbook(BytesIO(excel.content), data_only=True)
        worksheet = next(ws for ws in workbook.worksheets if ws.title != "Index")
        values = [cell.value for cell in worksheet[3]]
        self.assertEqual(values[8], 10000)
        # Le taux exporté est exactement celui choisi dans le système. Il ne
        # doit pas être recalculé à 10 % depuis 10 000 / 100 000.
        self.assertEqual(values[9], 0.15)
        self.assertEqual(values[10], 120000)
        self.assertEqual(values[11], 0)
        self.assertEqual(values[12], "Soldé - remise appliquée")

        ligne = _donnees_tranches_eleve(student, "2025-2026")
        self.assertEqual(ligne["pourcentages_remise"], (Decimal("15"),))
        self.assertEqual(_pourcentages_pdf(ligne["pourcentages_remise"]), "15 %")

        pdf = self.client.get(reverse("paiements:export_tranches_par_classe_pdf"), params)
        self.assertEqual(pdf.status_code, 200)
        self.assertTrue(pdf.content.startswith(b"%PDF"))

    def test_tranches_exporte_remise_fixe_sans_inventer_de_pourcentage(self):
        student = self._student("VENT-011", "Aïssatou")
        schedule = self._schedule(
            student, "REINSCRIPTION", 20000, tranche_1=100000,
            admission_paye=20000,
        )
        schedule.tranche_1_payee = Decimal("90000")
        schedule.save(update_fields=["tranche_1_payee"])
        payment_type = TypePaiement.objects.create(
            nom="Tranche 1 avec remise fixe", categorie="SCOLARITE",
        )
        mode = ModePaiement.objects.create(nom="Virement")
        payment = Paiement.objects.create(
            eleve=student,
            type_paiement=payment_type,
            mode_paiement=mode,
            numero_recu="VENT-REM-002",
            montant=Decimal("110000"),
            annee_scolaire="2025-2026",
            date_paiement=date(2026, 1, 15),
            statut="VALIDE",
        )
        discount = RemiseReduction.objects.create(
            nom="Geste commercial 10 000 GNF",
            type_remise="MONTANT_FIXE",
            valeur=Decimal("10000"),
            motif="SOCIALE",
            date_debut=date(2025, 9, 1),
            date_fin=date(2026, 6, 30),
        )
        PaiementRemise.objects.create(
            paiement=payment,
            remise=discount,
            montant_remise=Decimal("10000"),
            motif="GESTE_COMMERCIAL",
            tranches_concernees="1",
            base_calcul="TRANCHES_DUES",
        )

        params = {"annee_scolaire": "2025-2026", "classe": self.classe.pk}
        excel = self.client.get(
            reverse("paiements:export_tranches_par_classe_excel"), params
        )
        workbook = load_workbook(BytesIO(excel.content), data_only=True)
        worksheet = next(ws for ws in workbook.worksheets if ws.title != "Index")
        values = [cell.value for cell in worksheet[3]]

        self.assertEqual(values[7], 110000)
        self.assertEqual(values[8], 10000)
        self.assertIsNone(values[9])
        self.assertEqual(values[10], 120000)

        ligne = _donnees_tranches_eleve(student, "2025-2026")
        self.assertEqual(ligne["pourcentages_remise"], ())
        self.assertEqual(_pourcentages_pdf(ligne["pourcentages_remise"]), "")

    def test_remise_t1_ne_s_affiche_jamais_comme_paiement_t2_ou_t3(self):
        """Reproduit le reçu REC20260001 décrit par l'utilisateur."""
        student = self._student("PN1-001", "Fara")
        schedule = self._schedule(
            student, "INSCRIPTION", 50000, tranche_1=500000,
            admission_paye=50000,
        )
        schedule.tranche_1_payee = Decimal("475000")
        schedule.tranche_2_due = Decimal("500000")
        schedule.tranche_2_payee = Decimal("25000")
        schedule.tranche_3_due = Decimal("500000")
        schedule.save(update_fields=[
            "tranche_1_payee", "tranche_2_due", "tranche_2_payee",
            "tranche_3_due",
        ])
        payment_type = TypePaiement.objects.create(
            nom="Inscription + Tranche 1", categorie="SCOLARITE",
        )
        mode = ModePaiement.objects.create(nom="Cache")
        payment = Paiement.objects.create(
            eleve=student,
            type_paiement=payment_type,
            mode_paiement=mode,
            numero_recu="REC20260001",
            montant=Decimal("550000"),
            annee_scolaire="2025-2026",
            date_paiement=date(2026, 8, 23),
            statut="VALIDE",
        )
        discount = RemiseReduction.objects.create(
            nom="Remise scolarité 5% (T1)",
            type_remise="POURCENTAGE",
            valeur=Decimal("5"),
            motif="AUTRE",
            date_debut=date(2025, 9, 1),
            date_fin=date(2026, 8, 31),
        )
        PaiementRemise.objects.create(
            paiement=payment,
            remise=discount,
            montant_remise=Decimal("25000"),
            motif="GESTE_COMMERCIAL",
            tranches_concernees="1",
            base_calcul="TRANCHES_DUES",
            deduite_du_paiement=False,
        )

        ligne = _donnees_tranches_eleve(student, "2025-2026")
        self.assertEqual(ligne["inscription"], Decimal("50000"))
        self.assertEqual(ligne["tranche_1"], Decimal("475000"))
        self.assertEqual(ligne["tranche_2"], Decimal("0"))
        self.assertEqual(ligne["tranche_3"], Decimal("0"))
        self.assertEqual(ligne["total_paye"], Decimal("550000"))
        self.assertEqual(ligne["remise"], Decimal("25000"))
        self.assertEqual(ligne["reste"], Decimal("975000"))

        params = {"annee_scolaire": "2025-2026", "classe": self.classe.pk}
        excel = self.client.get(
            reverse("paiements:export_tranches_par_classe_excel"), params
        )
        workbook = load_workbook(BytesIO(excel.content), data_only=True)
        worksheet = next(ws for ws in workbook.worksheets if ws.title != "Index")
        headers = [cell.value for cell in worksheet[2]]
        values = [cell.value for cell in worksheet[3]]
        self.assertEqual(headers[7], "Encaissé")
        self.assertEqual(values[1:6], [50000, 0, 475000, 0, 0])
        self.assertEqual(values[7:12], [550000, 25000, 0.05, 575000, 975000])

        pdf_tranches = self.client.get(
            reverse("paiements:export_tranches_par_classe_pdf"), params
        )
        self.assertEqual(pdf_tranches.status_code, 200)
        self.assertTrue(pdf_tranches.content.startswith(b"%PDF"))

        receipt = self.client.get(
            reverse("paiements:generer_recu_pdf", args=[payment.pk])
        )
        self.assertEqual(receipt.status_code, 200)
        receipt_text = "\n".join(
            page.extract_text() or ""
            for page in PdfReader(BytesIO(receipt.content)).pages
        )
        self.assertIn("Montant encaissé : 550 000 GNF", receipt_text)
        self.assertIn(
            "Montant affecté aux échéances : 525 000 GNF", receipt_text
        )
        self.assertIn("1ère tranche: 475 000 GNF", receipt_text)
        self.assertIn("2ème tranche: 0 GNF", receipt_text)
        self.assertIn("3ème tranche: 0 GNF", receipt_text)
        self.assertIn("Remises appliquées", receipt_text)
        self.assertIn("Remise scolarité 5% (T1)", receipt_text)
        self.assertNotIn("Remise accordée", receipt_text)

    @patch("paiements.views.timezone.localdate", return_value=date(2025, 10, 1))
    def test_reenrollment_plus_first_installment_sets_nature_and_allocates(self, _localdate):
        student = self._student("VENT-007", "Binta")
        schedule = self._schedule(student, "INSCRIPTION", 30000, tranche_1=100000)
        payment_type = TypePaiement.objects.create(nom="Réinscription + Tranche 1")
        mode = ModePaiement.objects.create(nom="Espèces")
        payment = Paiement.objects.create(
            eleve=student,
            type_paiement=payment_type,
            mode_paiement=mode,
            montant=Decimal("120000"),
            date_paiement=date(2025, 10, 1),
            statut="VALIDE",
        )

        _allocate_payment_to_echeancier(payment)
        schedule.refresh_from_db()

        self.assertEqual(schedule.nature_frais, "REINSCRIPTION")
        self.assertEqual(schedule.frais_inscription_du, Decimal("20000"))
        self.assertEqual(schedule.frais_inscription_paye, Decimal("20000"))
        self.assertEqual(schedule.tranche_1_payee, Decimal("100000"))

    def test_schedule_page_displays_reenrollment_label(self):
        student = self._student("VENT-008", "M'Mah")
        self._schedule(student, "REINSCRIPTION", 20000)

        response = self.client.get(
            reverse("paiements:echeancier_eleve", args=[student.pk])
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Frais de réinscription")
        self.assertNotContains(response, ">Frais d'inscription<", html=False)
