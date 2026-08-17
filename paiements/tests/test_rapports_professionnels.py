from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import RequestFactory, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from pypdf import PdfReader

from eleves.models import Classe, Ecole, Eleve, Responsable
from paiements.models import (
    EcheancierPaiement,
    ModePaiement,
    Paiement,
    PaiementRemise,
    RemiseReduction,
    TypePaiement,
)
from paiements.rapports_professionnels import collect_accounting_data

from .support import TEST_MIDDLEWARE


@override_settings(MIDDLEWARE=TEST_MIDDLEWARE)
class RapportComptableProfessionnelTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username="rapport-professionnel",
            email="rapport@example.com",
            password="mot-de-passe-test",
            first_name="Aïssatou",
            last_name="Camara",
        )
        self.client.force_login(self.user)
        self.factory = RequestFactory()
        self.ecole = Ecole.objects.create(
            nom="École rapport professionnel",
            adresse="Conakry",
            telephone="+224620000601",
            directeur="Direction",
        )
        self.classe = Classe.objects.create(
            ecole=self.ecole,
            nom="11 Série littéraire",
            niveau="LYCEE_11",
            annee_scolaire="2025-2026",
        )
        responsable = Responsable.objects.create(
            prenom="Parent",
            nom="Rapport",
            relation="PERE",
            telephone="+224620000602",
        )
        self.eleve = Eleve.objects.create(
            matricule="RCE-001",
            prenom="Mariama",
            nom="Diallo",
            sexe="F",
            date_naissance=date(2010, 1, 1),
            classe=self.classe,
            date_inscription=date(2025, 9, 1),
            responsable_principal=responsable,
        )
        self.echeancier = EcheancierPaiement.objects.create(
            eleve=self.eleve,
            annee_scolaire="2025-2026",
            nature_frais="REINSCRIPTION",
            frais_inscription_du=Decimal("20000"),
            tranche_1_due=Decimal("100000"),
            tranche_2_due=0,
            tranche_3_due=0,
            date_echeance_inscription=date(2025, 9, 30),
            date_echeance_tranche_1=date(2026, 1, 10),
            date_echeance_tranche_2=date(2026, 3, 5),
            date_echeance_tranche_3=date(2026, 5, 5),
        )
        self.scolarite = TypePaiement.objects.create(
            nom="Réinscription + Tranche 1",
            categorie="SCOLARITE",
        )
        self.cantine = TypePaiement.objects.create(
            nom="Cantine mensuelle",
            categorie="CANTINE",
        )
        self.mobile = ModePaiement.objects.create(nom="Mobile Money")
        self.especes = ModePaiement.objects.create(nom="Espèces")
        self.paiement_scolarite = self._payment(
            "RCE-REC-001", "80000", "VALIDE", self.scolarite, self.mobile,
            reference="MM-2026-001",
        )
        self._payment(
            "RCE-REC-002", "15000", "VALIDE", self.cantine, self.especes,
        )
        self._payment(
            "RCE-REC-003", "20000", "EN_ATTENTE", self.scolarite, self.especes,
        )
        remise = RemiseReduction.objects.create(
            nom="Remise sociale rapport",
            type_remise="MONTANT_FIXE",
            valeur=Decimal("10000"),
            motif="SOCIALE",
            date_debut=date(2025, 9, 1),
            date_fin=date(2026, 6, 30),
        )
        PaiementRemise.objects.create(
            paiement=self.paiement_scolarite,
            remise=remise,
            montant_remise=Decimal("10000"),
            motif="GESTE_COMMERCIAL",
            tranches_concernees="1",
        )

    def _payment(self, receipt, amount, status, payment_type, mode, reference=""):
        return Paiement.objects.create(
            eleve=self.eleve,
            type_paiement=payment_type,
            mode_paiement=mode,
            numero_recu=receipt,
            montant=Decimal(amount),
            annee_scolaire="2025-2026",
            date_paiement=date(2026, 1, 15),
            statut=status,
            reference_externe=reference,
            cree_par=self.user,
            valide_par=self.user if status == "VALIDE" else None,
        )

    def _request(self, **params):
        values = {
            "classe_id": str(self.classe.pk),
            "du": "2026-01-01",
            "au": "2026-01-31",
        }
        values.update(params)
        request = self.factory.get("/paiements/export/comptabilite/pdf/", values)
        request.user = self.user
        return request

    def test_ventilation_separe_reinscription_tranches_et_autres_services(self):
        data = collect_accounting_data(self._request())

        self.assertEqual(data["validated_count"], 2)
        self.assertEqual(data["total_validated"], Decimal("95000"))
        self.assertEqual(data["total_discounts"], Decimal("10000"))
        self.assertEqual(data["by_component"]["inscription"]["amount"], 0)
        self.assertEqual(data["by_component"]["reinscription"]["amount"], 20000)
        self.assertEqual(data["by_component"]["tranche_1"]["amount"], 60000)
        self.assertEqual(data["by_component"]["autres"]["amount"], 15000)
        self.assertEqual(data["unallocated_total"], 0)
        self.assertEqual(data["by_status"]["EN_ATTENTE"]["amount"], 20000)

    def test_date_future_est_automatiquement_limitee_a_aujourdhui(self):
        future = timezone.localdate() + timedelta(days=30)

        data = collect_accounting_data(
            self._request(du="", au=future.isoformat())
        )

        self.assertTrue(data["period_adjusted"])
        self.assertEqual(data["end"], timezone.localdate())
        self.assertIn(timezone.localdate().strftime("%d/%m/%Y"), data["period_label"])

    def test_reference_mobile_money_manquante_est_signalee(self):
        self.paiement_scolarite.reference_externe = ""
        self.paiement_scolarite.save(update_fields=["reference_externe"])

        data = collect_accounting_data(self._request())

        self.assertEqual(data["reference_missing_count"], 1)
        self.assertEqual(data["reference_missing_amount"], Decimal("80000"))

    def test_exports_pdf_et_excel_contiennent_les_elements_professionnels(self):
        params = {
            "classe_id": self.classe.pk,
            "du": "2026-01-01",
            "au": "2026-01-31",
        }
        pdf = self.client.get(reverse("paiements:export_comptabilite_pdf"), params)
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf["Content-Type"], "application/pdf")
        self.assertTrue(pdf.content.startswith(b"%PDF"))
        self.assertGreater(len(pdf.content), 5000)

        excel = self.client.get(reverse("paiements:export_comptabilite_excel"), params)
        self.assertEqual(excel.status_code, 200)
        workbook = load_workbook(BytesIO(excel.content), data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            [
                "Synthèse", "Journal validé", "Situation élèves", "Affectations",
                "Statuts", "Ventilations", "Remises",
            ],
        )
        headers = [cell.value for cell in workbook["Journal validé"][1]]
        self.assertIn("Réinscription", headers)
        self.assertIn("Tranche 1", headers)
        self.assertIn("Remise (%)", headers)
        self.assertIn("Situation", headers)
        self.assertIn("Validateur", headers)
        student_row = [cell.value for cell in workbook["Situation élèves"][2]]
        self.assertEqual(student_row[7], 10000)
        self.assertEqual(student_row[8], 0.1)
        self.assertEqual(student_row[11], "Partiel - remise appliquée")
        summary_values = [workbook["Synthèse"].cell(row, 1).value for row in range(1, 15)]
        self.assertIn("Référence", summary_values)

    def test_exports_par_mode_contiennent_les_montants_et_totaux(self):
        params = {
            "classe_id": self.classe.pk,
            "du": "2026-01-01",
            "au": "2026-01-31",
            "statut": "VALIDE",
        }
        pdf = self.client.get(
            reverse("paiements:export_modes_encaissement_pdf"), params
        )
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf["Content-Type"], "application/pdf")
        self.assertIn("montants_par_mode_", pdf["Content-Disposition"])
        pdf_text = "\n".join(
            page.extract_text() or "" for page in PdfReader(BytesIO(pdf.content)).pages
        )
        self.assertIn("Mobile Money", pdf_text)
        self.assertIn("Espèces", pdf_text)
        self.assertIn("95 000", pdf_text)

        excel = self.client.get(
            reverse("paiements:export_modes_encaissement_excel"), params
        )
        self.assertEqual(excel.status_code, 200)
        self.assertEqual(
            excel["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(excel.content), data_only=False)
        self.assertEqual(workbook.sheetnames, ["Modes d'encaissement"])
        sheet = workbook["Modes d'encaissement"]
        self.assertEqual(
            [sheet.cell(9, column).value for column in range(1, 6)],
            [
                "Mode d'encaissement", "Nombre d'encaissements", "Montant (GNF)",
                "Part du total", "Montant moyen (GNF)",
            ],
        )
        self.assertEqual([sheet["A10"].value, sheet["B10"].value, sheet["C10"].value],
                         ["Mobile Money", 1, 80000])
        self.assertEqual([sheet["A11"].value, sheet["B11"].value, sheet["C11"].value],
                         ["Espèces", 1, 15000])
        self.assertEqual(sheet["A12"].value, "TOTAL")
        self.assertEqual(sheet["B12"].value, "=SUM(B10:B11)")
        self.assertEqual(sheet["C12"].value, "=SUM(C10:C11)")
        self.assertEqual(sheet["D10"].value, "=IF($C$12=0,0,C10/$C$12)")

    def test_export_par_mode_respecte_le_statut_selectionne(self):
        response = self.client.get(
            reverse("paiements:export_modes_encaissement_excel"),
            {
                "classe_id": self.classe.pk,
                "du": "2026-01-01",
                "au": "2026-01-31",
                "statut": "EN_ATTENTE",
            },
        )

        workbook = load_workbook(BytesIO(response.content), data_only=False)
        sheet = workbook["Modes d'encaissement"]
        self.assertEqual(sheet["B6"].value, "En attente")
        self.assertEqual([sheet["A10"].value, sheet["B10"].value, sheet["C10"].value],
                         ["Espèces", 1, 20000])
        self.assertEqual(sheet["A11"].value, "TOTAL")

    def test_tableau_par_mode_affiche_eleves_montants_et_soldes(self):
        self.assertEqual(
            reverse("paiements:modes_encaissement_eleves"),
            "/paiements/rapport/modes-encaissement/",
        )
        response = self.client.get(
            reverse("paiements:modes_encaissement_eleves"),
            {
                "classe_id": self.classe.pk,
                "date_debut": "2026-01-01",
                "date_fin": "2026-01-31",
                "statut": "VALIDE",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "DIALLO MARIAMA")
        self.assertContains(response, "RCE-001")
        self.assertContains(response, "Mobile Money")
        self.assertContains(response, "Espèces")
        self.assertEqual(response.context["student_count"], 1)
        self.assertEqual(response.context["operation_count"], 2)
        self.assertEqual(response.context["total_amount"], Decimal("95000"))
        self.assertEqual(response.context["remaining_total"], Decimal("30000"))
        mobile_row = next(
            row for row in response.context["rows"] if row["mode"] == "Mobile Money"
        )
        self.assertEqual(mobile_row["period_amount"], Decimal("80000"))
        self.assertEqual(mobile_row["paid"], Decimal("80000"))
        self.assertEqual(mobile_row["discount"], Decimal("10000"))
        self.assertEqual(mobile_row["total_due"], Decimal("120000"))
        self.assertEqual(mobile_row["remaining"], Decimal("30000"))
        self.assertEqual(mobile_row["situation"], "Partiel - remise appliquée")

        legacy_response = self.client.get(
            reverse("paiements:modes_encaissement_eleves_legacy"),
            {
                "classe_id": self.classe.pk,
                "date_debut": "2026-01-01",
                "date_fin": "2026-01-31",
            },
        )
        self.assertEqual(legacy_response.status_code, 200)

    def test_filtres_dynamiques_par_mode_statut_recherche_et_fragment(self):
        response = self.client.get(
            reverse("paiements:modes_encaissement_eleves"),
            {
                "classe_id": self.classe.pk,
                "date_debut": "2026-01-01",
                "date_fin": "2026-01-31",
                "statut": "EN_ATTENTE",
                "mode_id": self.especes.pk,
                "q": "RCE-001",
                "fragment": "1",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(
            response, "paiements/_modes_encaissement_resultats.html"
        )
        self.assertEqual(response.context["operation_count"], 1)
        self.assertEqual(response.context["total_amount"], Decimal("20000"))
        self.assertEqual(len(response.context["rows"]), 1)
        self.assertEqual(response.context["rows"][0]["mode"], "Espèces")
        self.assertNotContains(response, "Mobile Money")

        no_match = self.client.get(
            reverse("paiements:modes_encaissement_eleves"),
            {
                "classe_id": self.classe.pk,
                "date_debut": "2026-01-01",
                "date_fin": "2026-01-31",
                "q": "élève inexistant",
                "fragment": "1",
            },
        )
        self.assertEqual(no_match.context["student_count"], 0)
        self.assertContains(no_match, "Aucun élève ne correspond")

    def test_filtre_soldes_utilise_remises_et_paiements_valides(self):
        self._payment(
            "RCE-REC-SOLDE", "30000", "VALIDE", self.scolarite, self.especes,
        )

        response = self.client.get(
            reverse("paiements:modes_encaissement_eleves"),
            {
                "classe_id": self.classe.pk,
                "date_debut": "2026-01-01",
                "date_fin": "2026-01-31",
                "statut": "VALIDE",
                "situation": "solde",
            },
        )

        self.assertEqual(response.context["student_count"], 1)
        self.assertEqual(response.context["settled_count"], 1)
        self.assertEqual(response.context["remaining_total"], Decimal("0"))
        self.assertContains(response, "Soldé - remise appliquée")

    def test_interfaces_affichent_les_boutons_export_par_mode(self):
        dashboard = self.client.get(reverse("paiements:tableau_bord"))
        self.assertContains(
            dashboard, reverse("paiements:export_modes_encaissement_pdf")
        )
        self.assertContains(
            dashboard, reverse("paiements:export_modes_encaissement_excel")
        )
        self.assertContains(
            dashboard, reverse("paiements:modes_encaissement_eleves")
        )

        report = self.client.get(reverse("paiements:rapport_comptable"))
        self.assertContains(
            report, reverse("paiements:export_modes_encaissement_pdf")
        )
        self.assertContains(
            report, reverse("paiements:export_modes_encaissement_excel")
        )
        self.assertContains(
            report, reverse("paiements:modes_encaissement_eleves")
        )

    def test_remise_solde_la_scolarite_et_le_rapport_le_precise(self):
        self._payment(
            "RCE-REC-004", "30000", "VALIDE", self.scolarite, self.especes,
        )

        data = collect_accounting_data(self._request())

        situation = data["student_rows"][0]
        self.assertEqual(situation["total_due"], Decimal("120000"))
        self.assertEqual(situation["paid"], Decimal("110000"))
        self.assertEqual(situation["discount"], Decimal("10000"))
        self.assertEqual(situation["discount_rate"], Decimal("10"))
        self.assertEqual(situation["coverage"], Decimal("120000"))
        self.assertEqual(situation["remaining"], Decimal("0"))
        self.assertEqual(situation["situation"], "Soldé - remise appliquée")

        response = self.client.get(
            reverse("paiements:export_comptabilite_excel"),
            {"classe_id": self.classe.pk, "du": "2026-01-01", "au": "2026-01-31"},
        )
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        exported = [cell.value for cell in workbook["Situation élèves"][2]]
        self.assertEqual(exported[7], 10000)
        self.assertEqual(exported[8], 0.1)
        self.assertEqual(exported[10], 0)
        self.assertEqual(exported[11], "Soldé - remise appliquée")

    def test_export_exige_la_permission_de_consulter_les_rapports(self):
        simple_user = get_user_model().objects.create_user(
            username="sans-permission-rapport",
            password="mot-de-passe-test",
        )
        simple_user.profil.role = "ENSEIGNANT"
        simple_user.profil.peut_consulter_rapports = False
        simple_user.profil.save(update_fields=["role", "peut_consulter_rapports"])
        self.client.force_login(simple_user)

        for url_name in (
            "paiements:export_comptabilite_pdf",
            "paiements:export_modes_encaissement_pdf",
            "paiements:export_modes_encaissement_excel",
            "paiements:modes_encaissement_eleves",
        ):
            with self.subTest(url_name=url_name):
                response = self.client.get(reverse(url_name))
                self.assertEqual(response.status_code, 403)
